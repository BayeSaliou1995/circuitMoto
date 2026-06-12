# -*- coding: utf-8 -*-
"""Vues back-office liées aux personnes, inscriptions et emailing."""
from __future__ import annotations

import datetime as dt
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.core.mail import EmailMultiAlternatives, get_connection
from django.db import IntegrityError, transaction
from django.db.models import Prefetch, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.html import strip_tags
from django.urls import reverse
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..emails import _abs_url, notify_document_refuse, notify_inscription_validee, notify_paiement_resume_compose
from ..models import (
    Assurance,
    Circuit,
    ContactUrgence,
    Document,
    InfosMedicales,
    Inscription,
    JournalAudit,
    LibellePaiement,
    Moto,
    Paiement,
    Personne,
    SelectionOption,
    StatutDocument,
    StatutInscription,
)
from .common import is_ajax, json_err, json_ok, ordered_payments_for_display
from .forms import (
    ContactUrgenceFormSet,
    EmailingForm,
    InfosMedicalesForm,
    MotoFormSet,
    PaymentSummaryComposeForm,
    PersonneForm,
)
from .payments import apply_manual_payments_to_inscription, build_email_payment_rows, build_payment_summary_defaults


@staff_member_required
def inscription_list(request):
    q = (request.GET.get("q") or "").strip()[:200]
    statut = (request.GET.get("statut") or "").strip().upper()
    order = (request.GET.get("order") or "-cree_le").strip()
    page_size = int(request.GET.get("page_size", 20) or 20)

    qs = Inscription.objects.select_related("circuit", "pilote", "passager", "assurance")
    if q:
        terms = [t for t in q.split() if t]
        for t in terms:
            qs = qs.filter(
                Q(pilote__nom__icontains=t) |
                Q(pilote__prenom__icontains=t) |
                Q(passager__nom__icontains=t) |
                Q(passager__prenom__icontains=t) |
                Q(circuit__nom__icontains=t) |
                Q(circuit__code__icontains=t) |
                Q(id_public__icontains=t)
            )
    if statut:
        qs = qs.filter(statut=statut)

    allowed = {"-cree_le", "cree_le", "pilote__nom", "-pilote__nom", "circuit__nom", "-circuit__nom", "circuit__code", "-circuit__code", "circuit__date_debut", "-circuit__date_debut"}
    if order not in allowed:
        order = "-cree_le"
    qs = qs.order_by(order, "id")

    from django.core.paginator import Paginator
    page_obj = Paginator(qs, max(1, min(page_size, 200))).get_page(request.GET.get("page", 1))
    from django.utils.http import urlencode
    preserved = urlencode({k: v for k, v in request.GET.items() if k.lower() != "page"})
    circuits = Circuit.objects.order_by("-date_debut")
    return render(request, "circuitMoto/admin/inscriptions_list.html", {
        "page_obj": page_obj,
        "q": q,
        "statut": statut,
        "order": order,
        "preserved": preserved,
        "circuits": circuits,
        "export_years": [],
        "export_circuits": [],
        "exp_year": (request.GET.get("exp_year") or request.GET.get("year") or "").strip(),
        "exp_circuit": (request.GET.get("exp_circuit") or request.GET.get("circuit") or "").strip(),
    })


@staff_member_required
def personne_list(request):
    q = request.GET.get("q", "").strip()
    qs = Personne.objects.all().order_by("nom", "prenom").prefetch_related("motos", "contacts_urgence", "inscriptions_pilote__circuit", "inscriptions_passager__circuit")
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(email__icontains=q) | Q(telephone__icontains=q))
    circuits = Circuit.objects.order_by("-date_debut")
    return render(request, "circuitMoto/admin/personnes_list.html", {"personnes": qs, "q": q, "circuits": circuits})


@staff_member_required
@transaction.atomic
def personne_create(request):
    circuits = Circuit.objects.order_by("-date_debut")
    if request.method == "POST":
        form = PersonneForm(request.POST)
        med_form = InfosMedicalesForm(request.POST, prefix="med")
        motos = MotoFormSet(request.POST, prefix="motos")
        urg = ContactUrgenceFormSet(request.POST, prefix="urg")
        circuit_id = request.POST.get("circuit_id")
        if form.is_valid() and med_form.is_valid() and motos.is_valid() and urg.is_valid():
            personne = form.save()
            med_cd = med_form.cleaned_data
            if any(med_cd.get(k) for k in med_cd.keys()):
                InfosMedicales.objects.update_or_create(personne=personne, defaults=med_cd)
            for f in motos.forms:
                if not f.cleaned_data or f.cleaned_data.get("DELETE"):
                    continue
                if f.has_any_data():
                    Moto.objects.create(proprietaire=personne, marque=f.cleaned_data.get("marque") or "", modele=f.cleaned_data.get("modele") or "", immatriculation=f.cleaned_data.get("immatriculation") or "")
            for f in urg.forms:
                if not f.cleaned_data or f.cleaned_data.get("DELETE"):
                    continue
                if f.has_any_data():
                    ContactUrgence.objects.create(personne=personne, nom=f.cleaned_data.get("nom") or "", lien_parente=f.cleaned_data.get("lien_parente") or "", telephone=f.cleaned_data.get("telephone") or "")
            if circuit_id:
                circuit = get_object_or_404(Circuit, pk=circuit_id)
                ins = Inscription.objects.create(circuit=circuit, pilote=personne, devise=circuit.devise, prix_pilote_unitaire=circuit.prix_pilote_unitaire, prix_passager_unitaire=0)
                messages.success(request, "Dossier créé et inscrit sur le circuit sélectionné. Complétez l’inscription.")
                return redirect("inscription_edit_start", id_public=ins.id_public)
            messages.success(request, "Dossier personne créé.")
            return redirect("bo_personne_list")
    else:
        form = PersonneForm()
        med_form = InfosMedicalesForm(prefix="med")
        motos = MotoFormSet(prefix="motos")
        urg = ContactUrgenceFormSet(prefix="urg")
    return render(request, "circuitMoto/admin/personnes_form.html", {"form": form, "med_form": med_form, "motos": motos, "urg": urg, "circuits": circuits})


@staff_member_required
@transaction.atomic
def inscrire_personne(request, pk):
    if request.method != "POST":
        return redirect("bo_personne_list")
    personne = get_object_or_404(Personne, pk=pk)
    circuit = get_object_or_404(Circuit, pk=request.POST.get("circuit_id"))
    try:
        ins, created = Inscription.objects.get_or_create(circuit=circuit, pilote=personne, defaults={"devise": circuit.devise, "prix_pilote_unitaire": circuit.prix_pilote_unitaire, "prix_passager_unitaire": 0})
    except IntegrityError:
        ins = Inscription.objects.get(circuit=circuit, pilote=personne)
        created = False
    if created:
        messages.success(request, f"{personne.prenom} {personne.nom} inscrit sur {circuit.nom}.")
    else:
        messages.info(request, f"{personne.prenom} {personne.nom} est déjà inscrit sur « {circuit.nom} ».")
    return redirect("inscription_edit_start", id_public=ins.id_public)


@staff_member_required
@transaction.atomic
def inscription_delete(request, pk: int):
    if request.method != "POST":
        return redirect("bo_inscription_list")
    ins = get_object_or_404(Inscription, pk=pk)
    ref = ins.id_public
    ins.delete()
    messages.success(request, f"Inscription {ref} supprimée.")
    return redirect("bo_inscription_list")


@staff_member_required
def personne_detail(request, pk):
    p = get_object_or_404(Personne, pk=pk)
    if request.method == "POST":
        action = request.POST.get("action")
        ajax = is_ajax(request)
        try:
            ins_qs = Inscription.objects.filter(Q(pilote=p) | Q(passager=p))
            if action == "start_inscription":
                circuit_id = (request.POST.get("circuit_id") or "").strip()
                if not circuit_id:
                    msg = "Sélectionnez un circuit."
                    return json_err(msg) if ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))
                circuit = Circuit.objects.filter(pk=circuit_id).first()
                if not circuit:
                    msg = "Circuit introuvable."
                    return json_err(msg) if ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))
                url = f"{reverse('inscription_create_start')}?seed={p.pk}&circuit={circuit.pk}"
                return json_ok("Redirection vers l'assistant d'inscription…", redirect_url=url) if ajax else redirect(url)

            if action in {"doc_validate", "doc_refuse", "doc_reset"}:
                doc = Document.objects.filter(pk=request.POST.get("doc_id"), inscription__in=ins_qs).select_related("inscription", "verifie_par").first()
                if not doc:
                    msg = "Document introuvable pour cette personne."
                    return json_err(msg) if ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))
                if action == "doc_validate":
                    doc.valider(getattr(request, "user", None))
                    JournalAudit.objects.create(inscription=doc.inscription, acteur=request.user, action="DOCUMENT_VALIDE", donnees={"document_id": doc.pk, "type": doc.type_document})
                    msg = "Document validé."
                elif action == "doc_refuse":
                    raison = request.POST.get("raison", "")
                    notify_document_refuse(doc)
                    doc.refuser(raison, getattr(request, "user", None))
                    JournalAudit.objects.create(inscription=doc.inscription, acteur=request.user, action="DOCUMENT_REFUSE", donnees={"document_id": doc.pk, "raison": raison})
                    msg = "Document refusé."
                else:
                    doc.statut = StatutDocument.EN_ATTENTE
                    doc.verifie_par = None
                    doc.verifie_le = None
                    doc.note = ""
                    doc.save(update_fields=["statut", "verifie_par", "verifie_le", "note", "modifie_le"])
                    JournalAudit.objects.create(inscription=doc.inscription, acteur=request.user, action="DOCUMENT_REINITIALISE", donnees={"document_id": doc.pk})
                    msg = "Document réinitialisé en 'En attente'."
                payload = {"doc": {"id": doc.pk, "statut": doc.statut, "statut_label": doc.get_statut_display(), "note": doc.note or "", "verifie_par": ((doc.verifie_par.get_full_name() or doc.verifie_par.username or doc.verifie_par.email) if doc.verifie_par else ""), "verifie_le": timezone.localtime(doc.verifie_le).strftime("%d/%m/%Y %H:%M") if doc.verifie_le else ""}}
                return json_ok(msg, **payload) if ajax else (messages.success(request, msg) or redirect("bo_personne_detail", pk=p.pk))

            if action == "ins_validate":
                ins = ins_qs.filter(pk=request.POST.get("ins_id")).select_related("circuit").first()
                if not ins:
                    msg = "Inscription introuvable pour cette personne."
                    return json_err(msg) if ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))
                ins.statut = StatutInscription.VALIDE
                ins.save(update_fields=["statut", "modifie_le"])
                from ..models import initialiser_plan_paiement_par_defaut
                initialiser_plan_paiement_par_defaut(ins)
                notify_inscription_validee(ins)
                JournalAudit.objects.create(inscription=ins, acteur=request.user, action="INSCRIPTION_VALIDEE", donnees={"inscription_id": ins.pk})
                payload = {"ins": {"id": ins.pk, "statut": ins.statut, "statut_label": ins.get_statut_display()}, "hide_submitter": True}
                return json_ok("Inscription validée.", **payload) if ajax else (messages.success(request, "Inscription validée.") or redirect("bo_personne_detail", pk=p.pk))

            if action == "ins_payment_summary_preview":
                ins = ins_qs.filter(pk=request.POST.get("ins_id")).select_related("circuit", "pilote", "passager").prefetch_related("paiements", "selections_options__option").first()
                if not ins:
                    msg = "Inscription introuvable pour cette personne."
                    return json_err(msg) if ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))
                defaults = build_payment_summary_defaults(ins, p)
                form = PaymentSummaryComposeForm(initial=defaults)
                if ajax:
                    html = render_to_string("circuitMoto/admin/partials/_payment_summary_modal_form.html", {"form": form, "ins": ins, "personne": p, "recipient_email": defaults["recipient_email"], "recipient_name": defaults["recipient_name"], "paiements": defaults["paiements"], "has_acompte1": defaults["has_acompte1"], "has_acompte2": defaults["has_acompte2"], "has_options_added": defaults["has_options_added"]}, request=request)
                    return json_ok("Formulaire chargé.", modal_html=html)
                return redirect("bo_personne_detail", pk=p.pk)

            if action == "ins_payment_summary_send":
                ins = ins_qs.filter(pk=request.POST.get("ins_id")).select_related("circuit", "pilote", "passager").prefetch_related("paiements", "selections_options__option").first()
                if not ins:
                    msg = "Inscription introuvable pour cette personne."
                    return json_err(msg) if ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))
                form = PaymentSummaryComposeForm(request.POST)
                if not form.is_valid():
                    defaults = build_payment_summary_defaults(ins, p)
                    html = render_to_string("circuitMoto/admin/partials/_payment_summary_modal_form.html", {"form": form, "ins": ins, "personne": p, "recipient_email": defaults["recipient_email"], "recipient_name": defaults["recipient_name"], "paiements": defaults["paiements"], "has_acompte1": defaults["has_acompte1"], "has_acompte2": defaults["has_acompte2"], "has_options_added": defaults["has_options_added"]}, request=request)
                    return json_err("Merci de corriger le formulaire.", modal_html=html, status=422) if ajax else redirect("bo_personne_detail", pk=p.pk)
                recipient_role = form.cleaned_data["recipient_role"]
                recipient = ins.pilote if recipient_role == "pilote" else ins.passager
                if not recipient or not recipient.email:
                    msg = "Le destinataire n’a pas d’adresse email."
                    return json_err(msg) if ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                montant_options_ajoutees = form.cleaned_data.get("montant_options_ajoutees") or Decimal("0")
                paiement_options = form.cleaned_data.get("paiement_recu_pour_options") or Decimal("0")
                nouveau_total_attendu = form.cleaned_data.get("nouveau_total_attendu") or Decimal("0")
                if montant_options_ajoutees > 0:
                    ins.montant_options_ajoutees = int(montant_options_ajoutees)
                if paiement_options > 0:
                    ins.paiement_options_recu = int(paiement_options)
                if montant_options_ajoutees > 0 or paiement_options > 0:
                    ins.save(update_fields=["montant_options_ajoutees", "paiement_options_recu", "modifie_le"])

                existing_acompte1 = Decimal("0")
                existing_acompte2 = Decimal("0")
                existing_solde = Decimal("0")
                for pmt in ins.paiements.all():
                    enc = Decimal(pmt.montant_encaisse or 0)
                    if pmt.libelle == LibellePaiement.ACOMPTE1:
                        existing_acompte1 = enc
                    elif pmt.libelle == LibellePaiement.ACOMPTE2:
                        existing_acompte2 = enc
                    elif pmt.libelle == LibellePaiement.SOLDE:
                        existing_solde = enc
                new_acompte1 = max(Decimal(form.cleaned_data.get("acompte1_encaisse") or 0) - existing_acompte1, Decimal("0"))
                new_acompte2 = max(Decimal(form.cleaned_data.get("acompte2_encaisse") or 0) - existing_acompte2, Decimal("0"))
                new_solde = max(Decimal(form.cleaned_data.get("solde_encaisse") or 0) - existing_solde, Decimal("0"))
                updated_rows, remaining_unallocated = apply_manual_payments_to_inscription(ins, acompte1_amount=new_acompte1, acompte2_amount=new_acompte2, solde_amount=new_solde + paiement_options, payment_date=form.cleaned_data.get("date_paiement"))
                ins.refresh_from_db()
                paiements_after = ordered_payments_for_display(ins)
                total_paye = Decimal(sum((p.montant_encaisse or 0) for p in paiements_after))
                reste_a_payer = max(nouveau_total_attendu - total_paye, Decimal("0"))
                trop_percu = max(total_paye - nouveau_total_attendu, Decimal("0"))
                payment_rows = build_email_payment_rows(ins) if form.cleaned_data["inclure_detail_paiements"] else []
                preview_ctx = {
                    "ins": ins,
                    "recipient": recipient,
                    "dest_name": recipient.prenom or recipient.nom or "client",
                    "circuit": ins.circuit,
                    "payment_rows": payment_rows,
                    "intro_message": form.cleaned_data["intro_message"],
                    "note_client": form.cleaned_data["note_client"],
                    "total_attendu_initial": form.cleaned_data["total_attendu"],
                    "montant_options_ajoutees": montant_options_ajoutees,
                    "nouveau_total_attendu": nouveau_total_attendu,
                    "paiement_recu_pour_options": paiement_options,
                    "total_attendu": nouveau_total_attendu,
                    "total_paye": total_paye,
                    "reste_a_payer": reste_a_payer,
                    "trop_percu": trop_percu,
                    "situation_label": form.cleaned_data["situation_label"] or "",
                    "edit_url": _abs_url(reverse("inscription_edit_start", args=[ins.id_public])),
                    "infos_paiement": form.cleaned_data["infos_paiement_custom"] if form.cleaned_data["inclure_infos_paiement"] else "",
                    "has_infos_paiement": bool(form.cleaned_data["inclure_infos_paiement"] and form.cleaned_data["infos_paiement_custom"]),
                    "paiement_libelle": "Paiement manuel",
                    "subject": form.cleaned_data["sujet"],
                    "sent_acompte1_amount": new_acompte1,
                    "sent_acompte2_amount": new_acompte2,
                    "sent_solde_amount": new_solde,
                    "sent_options_amount": paiement_options,
                    "has_sent_acompte1": new_acompte1 > 0,
                    "has_sent_acompte2": new_acompte2 > 0,
                    "has_sent_solde": new_solde > 0,
                    "has_sent_options": paiement_options > 0,
                }
                notify_paiement_resume_compose(recipient_email=recipient.email, subject=form.cleaned_data["sujet"], ctx=preview_ctx, role=recipient_role)
                JournalAudit.objects.create(inscription=ins, acteur=request.user, action="PAIEMENT_RESUME_ENVOYE", donnees={"inscription_id": ins.pk, "destinataire_id": recipient.pk, "destinataire_email": recipient.email, "destinataire_role": recipient_role, "total_attendu_initial": str(form.cleaned_data.get("total_attendu") or 0), "montant_options_ajoutees": str(montant_options_ajoutees), "nouveau_total_attendu": str(nouveau_total_attendu), "acompte1_encaisse": str(form.cleaned_data.get("acompte1_encaisse") or 0), "acompte2_encaisse": str(form.cleaned_data.get("acompte2_encaisse") or 0), "solde_encaisse": str(form.cleaned_data.get("solde_encaisse") or 0), "paiement_recu_pour_options": str(paiement_options), "date_paiement": str(form.cleaned_data.get("date_paiement")), "total_paye": str(total_paye), "montant_restant": str(reste_a_payer), "trop_percu": str(trop_percu), "reste_non_impute": str(remaining_unallocated)})
                payments_payload = [{"id": pmt.pk, "montant_encaisse": str(pmt.montant_encaisse or 0), "encaisse_le": pmt.encaisse_le.strftime("%d/%m/%Y") if pmt.encaisse_le else "—", "statut": pmt.statut, "statut_label": pmt.get_statut_display()} for pmt in paiements_after]
                payment_summary_payload = {"ins_id": ins.pk, "options_added": str(montant_options_ajoutees or 0), "total_attendu": str(nouveau_total_attendu), "total_paye": str(total_paye), "reste_a_payer": str(reste_a_payer), "trop_percu": str(trop_percu), "payment_state": ("reste" if reste_a_payer > 0 else "trop_percu" if trop_percu > 0 else "solde"), "payment_state_label": ("Reste à payer" if reste_a_payer > 0 else "Trop-perçu" if trop_percu > 0 else "Soldé"), "devise": ins.devise}
                msg = f"Récapitulatif envoyé à {recipient.email}."
                return json_ok(msg, payments=payments_payload, payment_summary=payment_summary_payload) if ajax else (messages.success(request, msg) or redirect("bo_personne_detail", pk=p.pk))

            if action == "assurance_set_tel_urgence":
                ins = ins_qs.filter(pk=request.POST.get("ins_id")).select_related("assurance").first()
                tel = (request.POST.get("telephone_urgence") or "").strip()
                if not ins:
                    msg = "Inscription introuvable pour cette personne."
                    return json_err(msg) if ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))
                if not ins.assurance:
                    ins.assurance = Assurance.objects.create(inscription=ins)
                ins.assurance.telephone_urgence = tel
                ins.assurance.save(update_fields=["telephone_urgence", "modifie_le"])
                JournalAudit.objects.create(inscription=ins, acteur=request.user, action="ASSURANCE_TEL_URGENCE_MAJ", donnees={"telephone_urgence": tel})
                return json_ok("Numéro d’urgence mis à jour.", assurance={"ins_id": ins.pk, "telephone_urgence": ins.assurance.telephone_urgence}) if ajax else (messages.success(request, "Numéro d’urgence mis à jour.") or redirect("bo_personne_detail", pk=p.pk))
        except Exception as e:
            messages.error(request, f"Erreur lors du traitement: {e}")
            return redirect("bo_personne_detail", pk=p.pk)

    selections_prefetch = Prefetch("selections_options", queryset=SelectionOption.objects.select_related("option", "inscription"))
    documents_prefetch = Prefetch("documents", queryset=Document.objects.select_related("verifie_par").order_by("type_document", "-cree_le"))
    ins_pilote = Inscription.objects.filter(pilote=p).select_related("circuit", "assurance", "decharge", "pilote", "passager").prefetch_related(selections_prefetch, documents_prefetch, "paiements", "journaux_rappel", "journaux_audit").order_by("-cree_le")
    ins_passager = Inscription.objects.filter(passager=p).select_related("circuit", "assurance", "decharge", "pilote", "passager").prefetch_related(selections_prefetch, documents_prefetch, "paiements", "journaux_rappel", "journaux_audit").order_by("-cree_le")

    def enrich(ins):
        base = (ins.prix_pilote_unitaire or 0) + ((ins.prix_passager_unitaire or 0) if ins.passager_id else 0)
        opts = sum(sel.prix_total() for sel in ins.selections_options.all())
        total_paye = sum((pmt.montant_encaisse or 0) for pmt in ins.paiements.all())
        ins.total_base = base
        ins.total_options = opts
        ins.total_attendu = base + opts
        ins.total_paye = total_paye
        ins.reste_a_payer = max(ins.total_attendu - total_paye, 0)
        ins.trop_percu = max(total_paye - ins.total_attendu, 0)
        ins.options_added = getattr(ins, "montant_options_ajoutees", 0) or 0
        ins.paiements_display = ordered_payments_for_display(ins)
        ins.payment_state = "reste" if ins.reste_a_payer > 0 else "trop_percu" if ins.trop_percu > 0 else "solde"
        return ins

    ctx = {
        "personne": p,
        "infos_medicales": getattr(p, "infos_medicales", None),
        "motos": list(p.motos.all()),
        "contacts": list(p.contacts_urgence.all()),
        "ins_pilote": [enrich(i) for i in ins_pilote],
        "ins_passager": [enrich(i) for i in ins_passager],
        "circuits": Circuit.objects.order_by("-date_debut"),
    }
    return render(request, "circuitMoto/admin/personnes_detail.html", ctx)


@staff_member_required
def emailing(request):
    if request.method == "POST":
        form = EmailingForm(request.POST, request.FILES)
        if form.is_valid():
            circuit = form.cleaned_data["circuit"]
            statut = form.cleaned_data["statut"]
            incl_ps = form.cleaned_data["inclure_passagers"]
            sujet = form.cleaned_data["sujet"]
            message = form.cleaned_data["message"]
            test_only = form.cleaned_data["test_only"]
            files = request.FILES.getlist("pieces_jointes")
            ins_qs = Inscription.objects.select_related("pilote", "passager")
            if circuit:
                ins_qs = ins_qs.filter(circuit=circuit)
            if statut:
                ins_qs = ins_qs.filter(statut=statut)
            emails = set()
            for ins in ins_qs:
                if ins.pilote.email:
                    emails.add(ins.pilote.email)
                if incl_ps and ins.passager and ins.passager.email:
                    emails.add(ins.passager.email)
            if test_only:
                emails = {request.user.email} if request.user.email else set()
            if not emails:
                messages.warning(request, "Aucun destinataire.")
            else:
                attachments = []
                total_bytes = 0
                for f in files:
                    data = f.read()
                    total_bytes += len(data)
                    attachments.append((f.name, data, f.content_type or "application/octet-stream"))
                if total_bytes > 25 * 1024 * 1024 and len(emails) > 5:
                    messages.warning(request, "Beaucoup de pièces jointes volumineuses et de destinataires : pense à tester d’abord ou à réduire la taille.")
                sent = 0
                with get_connection(fail_silently=False) as conn:
                    for to in emails:
                        msg = EmailMultiAlternatives(subject=sujet, body=strip_tags(message), from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None), to=[to], connection=conn)
                        if "<" in message and ">" in message:
                            msg.attach_alternative(message, "text/html")
                        for name, data, ctype in attachments:
                            msg.attach(name, data, ctype)
                        msg.send()
                        sent += 1
                messages.success(request, f"Email envoyé à {sent} destinataire(s).")
                return redirect("bo_emailing")
    else:
        form = EmailingForm()
    return render(request, "circuitMoto/admin/emailing.html", {"form": form})


def _safe_str(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Oui" if v else "Non"
    return str(v)


@staff_member_required
def personnes_export_xlsx(request):
    year = (request.GET.get("year") or "").strip()
    circuit_id = (request.GET.get("circuit") or "").strip()
    statut = (request.GET.get("statut") or "").strip()
    qs = Inscription.objects.select_related("circuit", "assurance", "pilote", "passager", "pilote__infos_medicales", "passager__infos_medicales").prefetch_related(Prefetch("selections_options", queryset=SelectionOption.objects.select_related("option")), Prefetch("pilote__contacts_urgence"), Prefetch("passager__contacts_urgence")).order_by("circuit__date_debut", "pilote__nom", "pilote__prenom")
    if year:
        qs = qs.filter(circuit__date_debut__year=year)
    if circuit_id:
        qs = qs.filter(circuit_id=circuit_id)
    if statut:
        qs = qs.filter(statut=statut)

    person_cols = [("nom", "Nom"), ("prenom", "Prénom"), ("email", "Email"), ("telephone", "Téléphone"), ("date_naissance", "Date de naissance"), ("age", "Âge"), ("numero_carte_identite", "N° carte d’identité"), ("adresse", "Adresse"), ("code_postal", "Code postal"), ("localite", "Localité"), ("pays", "Pays")]
    context_cols = [("role", "Rôle"), ("statut", "Statut inscription"), ("annee", "Année"), ("circuit_code", "Code circuit"), ("circuit_nom", "Nom circuit"), ("circuit_debut", "Début"), ("circuit_fin", "Fin"), ("inscription_id", "ID interne"), ("id_public", "ID public"), ("cree_le", "Inscription créée le")]
    medical_cols = [("med_groupe", "Groupe sanguin"), ("med_hta", "HTA"), ("med_asthme", "Asthme"), ("med_epilepsie", "Épilepsie"), ("med_peau", "Problèmes de peau"), ("med_vertiges", "Vertiges"), ("med_notes", "Notes médicales")]
    assurance_cols = [("ass_type", "Assurance - Type"), ("ass_compagnie", "Assurance - Compagnie"), ("ass_police", "Assurance - N° police"), ("ass_valide_du", "Ass. valable du"), ("ass_valide_au", "Ass. valable au"), ("ass_tel_urg", "Ass. N° d’urgence")]
    MAX_CONTACTS = 2
    contact_cols = []
    for i in range(1, MAX_CONTACTS + 1):
        contact_cols += [(f"cu{i}_nom", f"Contact {i} - Nom"), (f"cu{i}_lien", f"Contact {i} - Lien"), (f"cu{i}_tel", f"Contact {i} - Téléphone")]
    options_cols = [("opts_synthese", "Options (code×qte)"), ("opts_detail", "Options (intitulés)"), ("opts_total", "Montant options")]

    wb = Workbook()
    ws = wb.active
    ws.title = "Inscrits"
    header = [label for _, label in (person_cols + context_cols + medical_cols + assurance_cols + contact_cols + options_cols)]
    ws.append(header)
    header_fill = PatternFill("solid", fgColor="EEF2FF")
    header_font = Font(bold=True)
    for col_idx in range(1, len(header) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    def _as_date(v):
        if v is None:
            return None
        if isinstance(v, dt.datetime):
            return v.date()
        if isinstance(v, dt.date):
            return v
        return v

    def _medical_vals(person):
        m = getattr(person, "infos_medicales", None)
        return {"med_groupe": (m.groupe_sanguin if m and m.groupe_sanguin else ""), "med_hta": bool(m and m.hta), "med_asthme": bool(m and m.asthme), "med_epilepsie": bool(m and m.epilepsie), "med_peau": bool(m and m.problemes_peau), "med_vertiges": bool(m and m.vertiges), "med_notes": (m.notes or "") if m else ""}

    def _assurance_vals(ins):
        a = getattr(ins, "assurance", None)
        return {"ass_type": (a.get_type_display() if a else ""), "ass_compagnie": (a.compagnie if a else ""), "ass_police": (a.numero_police if a else ""), "ass_valide_du": _as_date(a.valide_du) if a else None, "ass_valide_au": _as_date(a.valide_au) if a else None, "ass_tel_urg": (a.telephone_urgence if a else "")}

    def _contact_vals(person):
        rows = {"cu1_nom": "", "cu1_lien": "", "cu1_tel": "", "cu2_nom": "", "cu2_lien": "", "cu2_tel": ""}
        contacts = list(getattr(person, "contacts_urgence", []).all()) if person and hasattr(person, "contacts_urgence") else []
        try:
            contacts.sort(key=lambda c: getattr(c, "cree_le", dt.datetime.min), reverse=True)
        except Exception:
            pass
        for idx, c in enumerate(contacts[:MAX_CONTACTS], start=1):
            rows[f"cu{idx}_nom"] = c.nom or ""
            rows[f"cu{idx}_lien"] = c.lien_parente or ""
            rows[f"cu{idx}_tel"] = c.telephone or ""
        return rows

    def _options_vals(ins):
        sels = list(getattr(ins, "selections_options", []).all()) if hasattr(ins, "selections_options") else []
        if not sels:
            return {"opts_synthese": "", "opts_detail": "", "opts_total": 0}
        synth, detail, total = [], [], 0
        for s in sels:
            code = s.option.code if s.option_id else ""
            intitule = s.option.intitule if s.option_id else ""
            q = s.quantite or 0
            synth.append(f"{code}×{q}")
            who = " (passager)" if s.pour_passager else ""
            detail.append(f"{intitule}×{q}{who}")
            total += (s.prix_unitaire_fige or 0) * q
        return {"opts_synthese": " | ".join(synth)[:500], "opts_detail": " | ".join(detail)[:500], "opts_total": total}

    def add_row(ins, person, role: str):
        if not person:
            return
        circ = ins.circuit
        pvals = {"nom": person.nom or "", "prenom": person.prenom or "", "email": person.email or "", "telephone": person.telephone or "", "date_naissance": _as_date(getattr(person, "date_naissance", None)), "age": getattr(person, "age", None) or "", "numero_carte_identite": person.numero_carte_identite or "", "adresse": person.adresse or "", "code_postal": person.code_postal or "", "localite": person.localite or "", "pays": person.pays or ""}
        cvals = {"role": role, "statut": ins.statut, "annee": circ.date_debut.year if circ and circ.date_debut else "", "circuit_code": getattr(circ, "code", ""), "circuit_nom": getattr(circ, "nom", ""), "circuit_debut": _as_date(getattr(circ, "date_debut", None)), "circuit_fin": _as_date(getattr(circ, "date_fin", None)), "inscription_id": ins.id, "id_public": str(ins.id_public), "cree_le": _as_date(getattr(ins, "cree_le", None))}
        mvals, avals, cuvals, ovals = _medical_vals(person), _assurance_vals(ins), _contact_vals(person), _options_vals(ins)
        row = []
        for key, _ in person_cols: row.append(pvals.get(key, ""))
        for key, _ in context_cols: row.append(cvals.get(key, ""))
        for key, _ in medical_cols: row.append(_safe_str(mvals.get(key, "")))
        for key, _ in assurance_cols: row.append(avals.get(key, ""))
        for key, _ in contact_cols: row.append(cuvals.get(key, ""))
        for key, _ in options_cols: row.append(ovals.get(key, ""))
        ws.append(row)

    for ins in qs:
        add_row(ins, ins.pilote, "Pilote")
        if ins.passager_id:
            add_row(ins, ins.passager, "Passager")

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for cell in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            max_len = max(max_len, len(_safe_str(cell[0].value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 50)

    now = timezone.now().strftime("%Y%m%d-%H%M")
    filename = f"inscrits_{year or 'toutes-annees'}_{circuit_id or 'tous-circuits'}_{now}.xlsx"
    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)
    resp = HttpResponse(buff.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
