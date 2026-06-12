# circuitMoto/admin_views.py
# -*- coding: utf-8 -*-
from typing import List, Tuple
import re
import time
from django.conf import settings
from django import forms # type: ignore
from django.forms import formset_factory
from django.utils.text import slugify
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger # type: ignore
from django.contrib import messages # type: ignore
from django.contrib.admin.views.decorators import staff_member_required # type: ignore
from django.db.models import Prefetch
from django.core.mail import send_mass_mail # type: ignore
from django.shortcuts import render, redirect, get_object_or_404 # type: ignore
from django.utils import timezone # type: ignore
from django.core.paginator import Paginator # type: ignore
from django.utils.http import urlencode # type: ignore
from django.forms import inlineformset_factory
from django.http import JsonResponse
import datetime
from django.urls import reverse
from django.contrib.auth.hashers import check_password
import secrets, string
from django.db.models import Count, Q, Sum, F, IntegerField
from django.db.models.functions import TruncMonth, Coalesce
from datetime import date
from django.contrib.auth import update_session_auth_hash, get_user_model
from django.contrib.auth.password_validation import validate_password, ValidationError as PwValidationError
from django.contrib.auth.decorators import login_required
from django.db import transaction, IntegrityError
from django.shortcuts import render, redirect, get_object_or_404
from django.core.mail import EmailMultiAlternatives, get_connection
from django.utils.html import strip_tags
from django.views.decorators.http import require_POST
from django.core.cache import cache
# circuitMoto/admin_views.py
from django.db.models.deletion import ProtectedError
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from io import BytesIO
from django.http import HttpResponse
from django.db.models import Field
from django.db.models.functions import ExtractYear
from openpyxl import Workbook, load_workbook # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
import datetime as dt 
# import render_string
from django.template.loader import render_to_string
import os
from pathlib import Path
from django.db.models.functions import Lower
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.http import FileResponse, Http404
from django.utils import timezone

from .models import JournalAudit
from .backup_service import create_full_backup_zip, BackupError


from decimal import Decimal

from .models import (
    Circuit, Inscription, Document, SelectionOption, Personne,
    StatutInscription, OptionCircuit, InfosMedicales, Moto,
    ContactUrgence, StatutDocument, JournalAudit, StatutCircuit,
    initialiser_plan_paiement_par_defaut, SitePrefs,
    Paiement, ContactMessage, NewsletterSubscriber,
    LibellePaiement, StatutPaiement, MethodePaiement, BaladeJour, DemandeProgrammeBalade,  
    NiveauInteret2027, Programme2027Circuit, Programme2027Interest, Programme2027Selection,
)

from .emails import notify_inscription_validee, notify_document_refuse

def _ordered_payments_for_display(ins):
    order_map = {
        LibellePaiement.ACOMPTE1: 1,
        LibellePaiement.ACOMPTE2: 2,
        LibellePaiement.SOLDE: 3,
        LibellePaiement.AUTRE: 4,
    }
    paiements = list(ins.paiements.all())
    paiements.sort(key=lambda p: (order_map.get(p.libelle, 99), p.echeance_le or date.max, p.pk))
    return paiements

# =========================================================
# OUTILS NEWSLETTER
# =========================================================

EMAIL_SPLIT_RE = re.compile(r"[,\n;\t ]+")


def _parse_emails_blob(raw: str) -> Tuple[List[str], List[str]]:
    """
    Transforme un texte collé en liste d'emails valides.
    Accepte séparateurs: virgule, point-virgule, retour ligne, espace, tabulation.
    Retourne (emails_valides_uniques, emails_invalides).
    """
    if not raw:
        return [], []

    candidates = [item.strip() for item in EMAIL_SPLIT_RE.split(raw) if item.strip()]

    valid_emails = []
    invalid_emails = []
    seen = set()

    for email in candidates:
        email_norm = email.lower()
        try:
            validate_email(email_norm)
            if email_norm not in seen:
                seen.add(email_norm)
                valid_emails.append(email_norm)
        except ValidationError:
            invalid_emails.append(email)

    return valid_emails, invalid_emails


def _get_newsletter_config() -> dict:
    """
    Charge la configuration depuis settings, avec possibilité d'extension plus tard
    via SitePrefs si tu ajoutes ces champs dans la base.
    """
    return {
        "enabled": getattr(settings, "NEWSLETTER_ENABLED", True),
        "max_recipients": int(getattr(settings, "NEWSLETTER_MAX_RECIPIENTS_PER_SEND", 1500)),
        "batch_size": int(getattr(settings, "NEWSLETTER_BATCH_SIZE", 50)),
        "sleep_seconds": float(getattr(settings, "NEWSLETTER_SLEEP_SECONDS", 1)),
        "max_attachment_bytes": int(getattr(settings, "NEWSLETTER_MAX_ATTACHMENT_BYTES", 25 * 1024 * 1024)),
        "hide_old_emailing": bool(getattr(settings, "NEWSLETTER_HIDE_OLD_EMAILING", True)),
    }


def _send_newsletter_in_batches(
    *,
    subject: str,
    message: str,
    recipients: List[str],
    attachments: List[Tuple[str, bytes, str]],
    from_email: str,
    batch_size: int,
    sleep_seconds: float,
) -> int:
    """
    Envoi par lots avec une seule connexion SMTP réutilisée.
    Retourne le nombre total envoyé.
    """
    sent = 0

    if not recipients:
        return sent

    with get_connection(fail_silently=False) as conn:
        for start in range(0, len(recipients), batch_size):
            batch = recipients[start:start + batch_size]

            for to_email in batch:
                msg = EmailMultiAlternatives(
                    subject=subject,
                    body=strip_tags(message),
                    from_email=from_email,
                    to=[to_email],
                    connection=conn,
                )

                # Si le message ressemble à du HTML, on l'ajoute en version HTML
                if "<" in message and ">" in message:
                    msg.attach_alternative(message, "text/html")

                for name, data, ctype in attachments:
                    msg.attach(name, data, ctype)

                msg.send()
                sent += 1

            # Petite pause entre lots pour éviter de saturer SMTP
            if sleep_seconds > 0 and (start + batch_size) < len(recipients):
                time.sleep(sleep_seconds)

    return sent

def _get_payment_priority_rows(ins):
    return list(ins.paiements.all().order_by("echeance_le", "pk"))


def _apply_manual_payments_to_inscription(
    ins,
    *,
    acompte1_amount=Decimal("0"),
    acompte2_amount=Decimal("0"),
    solde_amount=Decimal("0"),
    payment_date=None,
):
    """
    Met à jour les paiements de l'inscription SANS mélanger les champs.
    - acompte1_amount s'applique uniquement à ACOMPTE1
    - acompte2_amount s'applique uniquement à ACOMPTE2
    - solde_amount s'applique uniquement à SOLDE
    """
    payment_date = payment_date or timezone.localdate()

    mapping = {
        LibellePaiement.ACOMPTE1: Decimal(acompte1_amount or 0),
        LibellePaiement.ACOMPTE2: Decimal(acompte2_amount or 0),
        LibellePaiement.SOLDE: Decimal(solde_amount or 0),
    }

    updated_rows = []
    remaining_unallocated = Decimal("0")

    for libelle, amount in mapping.items():
        if amount <= 0:
            continue

        p = ins.paiements.filter(libelle=libelle).first()
        if not p:
            remaining_unallocated += amount
            continue

        attendu = Decimal(p.montant_du or 0)
        deja_encaisse = Decimal(p.montant_encaisse or 0)
        restant_sur_ligne = max(attendu - deja_encaisse, Decimal("0"))

        a_imputer = min(amount, restant_sur_ligne)
        surplus = max(amount - a_imputer, Decimal("0"))

        if a_imputer > 0:
            nouveau_total = deja_encaisse + a_imputer
            p.montant_encaisse = nouveau_total
            p.encaisse_le = payment_date

            if nouveau_total >= attendu and attendu > 0:
                p.statut = StatutPaiement.PAYE
            elif nouveau_total > 0:
                p.statut = StatutPaiement.PARTIEL
            else:
                p.statut = StatutPaiement.A_PAYER

            p.save(update_fields=["montant_encaisse", "encaisse_le", "statut", "modifie_le"])
            updated_rows.append(p)

        if surplus > 0:
            remaining_unallocated += surplus

    return list(ins.paiements.all().order_by("echeance_le", "pk")), remaining_unallocated


def _build_payment_reminder_defaults(ins, personne):
    paiements = _ordered_payments_for_display(ins)

    total_attendu = Decimal(
        (ins.prix_pilote_unitaire or 0)
        + ((ins.prix_passager_unitaire or 0) if ins.passager_id else 0)
        + sum(sel.prix_total() for sel in ins.selections_options.all())
    )

    total_paye = Decimal(sum((p.montant_encaisse or 0) for p in paiements))
    montant_restant = max(total_attendu - total_paye, Decimal("0"))
    trop_percu = max(total_paye - total_attendu, Decimal("0"))

    if montant_restant > 0:
        situation_label = "Reste à payer"
    elif trop_percu > 0:
        situation_label = "Trop-perçu"
    else:
        situation_label = "Soldé"

    recipient_role = "pilote"
    recipient = ins.pilote or ins.passager

    recipient = ins.pilote if recipient_role == "pilote" else ins.passager

    return {
        "ins_id": ins.pk,
        "recipient_role": recipient_role,
        "sujet": f"[{ins.circuit.code}] Rappel de paiement – {ins.circuit.nom}",
        "intro_message": (
            "Bonjour,\n\n"
            "Nous vous rappelons qu’un montant reste à régler pour votre inscription."
        ),
        "total_attendu": total_attendu,
        "total_paye": total_paye,
        "montant_restant": montant_restant,
        "situation_label": situation_label,
        "inclure_detail_paiements": True,
        "inclure_infos_paiement": True,
        "infos_paiement_custom": (getattr(ins.circuit, "infos_paiement", "") or "").strip(),
        "note_client": (
            "Merci de bien vouloir procéder au règlement du montant restant dans les meilleurs délais."
        ),
        "recipient_email": getattr(recipient, "email", "") if recipient else "",
        "recipient_name": getattr(recipient, "nom_complet", "") if recipient else "",
        "paiements": paiements,
    }

# =========================================================
# FORMULAIRE NEWSLETTER
# =========================================================

class MultipleFilesInput(forms.ClearableFileInput):
    allow_multiple_selected = True


class NewsletterBroadcastForm(forms.Form):
    emails_blob = forms.CharField(
        label="Emails des destinataires",
        widget=forms.Textarea(attrs={
            "rows": 12,
            "placeholder": (
                "Collez ici les emails séparés par virgules, points-virgules ou retours à la ligne.\n"
                "Exemple :\n"
                "client1@email.com\n"
                "client2@email.com; client3@email.com, client4@email.com"
            ),
        }),
        help_text="Les doublons seront supprimés automatiquement.",
    )

    sujet = forms.CharField(
        max_length=180,
        label="Sujet",
        widget=forms.TextInput(attrs={"placeholder": "Objet de la newsletter"})
    )

    message = forms.CharField(
        label="Message",
        widget=forms.Textarea(attrs={
            "rows": 12,
            "placeholder": "Écrivez votre message ici. Le HTML simple est accepté."
        }),
        help_text="Vous pouvez coller du texte simple ou du HTML léger.",
    )

    test_only = forms.BooleanField(
        required=False,
        initial=False,
        label="N’envoyer qu’à moi (test)"
    )

    pieces_jointes = forms.FileField(
        label="Pièces jointes",
        required=False,
        widget=MultipleFilesInput(attrs={
            "multiple": True,
            "accept": ".pdf,.doc,.docx,.jpg,.jpeg,.png,.zip",
        }),
        help_text="Plusieurs fichiers possibles. Évitez des pièces jointes trop lourdes.",
    )

    recipients_count = forms.IntegerField(required=False, widget=forms.HiddenInput())
    invalid_count = forms.IntegerField(required=False, widget=forms.HiddenInput())

    def clean_pieces_jointes(self):
        config = _get_newsletter_config()
        files = self.files.getlist("pieces_jointes")

        total_bytes = 0
        for f in files:
            total_bytes += f.size

        if total_bytes > config["max_attachment_bytes"]:
            raise forms.ValidationError(
                f"Le poids total des pièces jointes dépasse la limite autorisée "
                f"({config['max_attachment_bytes'] // (1024 * 1024)} Mo)."
            )

        return files

    def clean_emails_blob(self):
        raw = self.cleaned_data.get("emails_blob", "")
        config = _get_newsletter_config()

        valid_emails, invalid_emails = _parse_emails_blob(raw)

        self._parsed_valid_emails = valid_emails
        self._parsed_invalid_emails = invalid_emails

        if not valid_emails:
            raise forms.ValidationError("Aucune adresse e-mail valide détectée.")

        if len(valid_emails) > config["max_recipients"]:
            raise forms.ValidationError(
                f"Le nombre d’adresses ({len(valid_emails)}) dépasse la limite autorisée "
                f"({config['max_recipients']})."
            )

        return raw

    def get_valid_emails(self) -> List[str]:
        return getattr(self, "_parsed_valid_emails", [])

    def get_invalid_emails(self) -> List[str]:
        return getattr(self, "_parsed_invalid_emails", [])


# =========================================================
# VUE NEWSLETTER
# =========================================================

@staff_member_required
def newsletter_broadcast(request):
    config = _get_newsletter_config()

    if not config["enabled"]:
        messages.error(request, "Le module newsletter est actuellement désactivé.")
        return redirect("bo_dashboard")

    preview_stats = {
        "valid_count": 0,
        "invalid_count": 0,
        "sample_invalids": [],
        "batch_size": config["batch_size"],
        "max_recipients": config["max_recipients"],
    }

    if request.method == "POST":
        form = NewsletterBroadcastForm(request.POST, request.FILES)

        if form.is_valid():
            subject = form.cleaned_data["sujet"].strip()
            message = form.cleaned_data["message"]
            test_only = form.cleaned_data["test_only"]

            recipients = form.get_valid_emails()
            invalids = form.get_invalid_emails()

            preview_stats["valid_count"] = len(recipients)
            preview_stats["invalid_count"] = len(invalids)
            preview_stats["sample_invalids"] = invalids[:10]

            if test_only:
                recipients = [request.user.email] if request.user.email else []
                if not recipients:
                    messages.error(
                        request,
                        "Votre compte administrateur n’a pas d’adresse email. Impossible d’exécuter le test."
                    )
                    return render(
                        request,
                        "circuitMoto/admin/newsletter_broadcast.html",
                        {"form": form, "preview_stats": preview_stats, "config": config},
                    )

            # Protection anti double clic / double soumission très rapprochée
            lock_key = f"newsletter_send_lock_user_{request.user.pk}"
            if cache.get(lock_key):
                messages.warning(request, "Un envoi est déjà en cours ou vient d’être lancé. Réessayez dans un instant.")
                return render(
                    request,
                    "circuitMoto/admin/newsletter_broadcast.html",
                    {"form": form, "preview_stats": preview_stats, "config": config},
                )

            cache.set(lock_key, True, timeout=60)

            try:
                uploaded_files = request.FILES.getlist("pieces_jointes")
                attachments = []

                for f in uploaded_files:
                    attachments.append((
                        f.name,
                        f.read(),
                        f.content_type or "application/octet-stream",
                    ))

                sent = _send_newsletter_in_batches(
                    subject=subject,
                    message=message,
                    recipients=recipients,
                    attachments=attachments,
                    from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
                    batch_size=config["batch_size"],
                    sleep_seconds=config["sleep_seconds"],
                )

                msg = (
                    f"Newsletter envoyée à {sent} destinataire(s). "
                    f"Lots de {config['batch_size']}."
                )
                if invalids and not test_only:
                    msg += f" {len(invalids)} adresse(s) invalide(s) ignorée(s)."

                messages.success(request, msg)
                return redirect("bo_newsletter_broadcast")

            finally:
                cache.delete(lock_key)

        else:
            raw = request.POST.get("emails_blob", "")
            valid_emails, invalid_emails = _parse_emails_blob(raw)
            preview_stats["valid_count"] = len(valid_emails)
            preview_stats["invalid_count"] = len(invalid_emails)
            preview_stats["sample_invalids"] = invalid_emails[:10]

    else:
        form = NewsletterBroadcastForm()

    return render(
        request,
        "circuitMoto/admin/newsletter_broadcast.html",
        {
            "form": form,
            "preview_stats": preview_stats,
            "config": config,
            "hide_old_emailing": config["hide_old_emailing"],
        },
    )

# =========================

from decimal import Decimal, InvalidOperation

from decimal import Decimal
from django import forms

from decimal import Decimal
from django import forms


from decimal import Decimal
from django import forms
from django.utils import timezone


from decimal import Decimal
from django import forms
from django.utils import timezone


class PaymentSummaryComposeForm(forms.Form):
    ins_id = forms.IntegerField(widget=forms.HiddenInput())

    recipient_role = forms.ChoiceField(
        choices=(
            ("pilote", "Pilote"),
            ("passager", "Passager"),
        ),
        label="Destinataire",
    )

    sujet = forms.CharField(
        max_length=180,
        label="Sujet",
        widget=forms.TextInput(attrs={
            "placeholder": "Ex : Point sur votre paiement – Corse 2025",
        })
    )

    intro_message = forms.CharField(
        required=False,
        label="Message d’introduction",
        widget=forms.Textarea(attrs={
            "rows": 4,
            "placeholder": (
                "Bonjour,\n\n"
                "Voici votre situation actuelle de paiement concernant votre inscription."
            )
        })
    )

    total_attendu = forms.DecimalField(
        max_digits=12,
        decimal_places=2,
        min_value=0,
        label="Total attendu initial",
    )

    montant_options_ajoutees = forms.DecimalField(
        max_digits=12,
        decimal_places=2,
        min_value=0,
        required=False,
        initial=Decimal("0.00"),
        label="Montant des options ajoutées",
    )

    nouveau_total_attendu = forms.DecimalField(
        max_digits=12,
        decimal_places=2,
        min_value=0,
        required=False,
        label="Nouveau total attendu",
    )

    acompte1_encaisse = forms.DecimalField(
        max_digits=12,
        decimal_places=2,
        min_value=0,
        required=False,
        initial=Decimal("0.00"),
        label="Acompte 1 encaissé",
    )

    acompte2_encaisse = forms.DecimalField(
        max_digits=12,
        decimal_places=2,
        min_value=0,
        required=False,
        initial=Decimal("0.00"),
        label="Acompte 2 encaissé",
    )

    solde_encaisse = forms.DecimalField(
        max_digits=12,
        decimal_places=2,
        min_value=0,
        required=False,
        initial=Decimal("0.00"),
        label="Montant encaissé après acomptes",
    )

    paiement_recu_pour_options = forms.DecimalField(
        max_digits=12,
        decimal_places=2,
        min_value=0,
        required=False,
        initial=Decimal("0.00"),
        label="Paiement reçu pour les nouvelles options",
    )

    date_paiement = forms.DateField(
        required=True,
        label="Date du paiement",
        initial=timezone.localdate,
        widget=forms.DateInput(attrs={
            "type": "date",
            "placeholder": "jj/mm/aaaa",
        }),
        error_messages={
            "required": "Merci de renseigner la date du paiement.",
            "invalid": "Merci de saisir une date valide.",
        }
    )

    montant_restant = forms.DecimalField(
        max_digits=12,
        decimal_places=2,
        min_value=0,
        required=False,
        label="Reste à payer",
    )

    trop_percu = forms.DecimalField(
        max_digits=12,
        decimal_places=2,
        min_value=0,
        required=False,
        label="Trop-perçu",
    )

    situation_label = forms.CharField(
        max_length=120,
        required=False,
        label="Libellé de situation",
        widget=forms.TextInput(attrs={
            "placeholder": "Ex : paiement reçu après ajout d’options"
        })
    )

    inclure_detail_paiements = forms.BooleanField(
        required=False,
        initial=True,
        label="Inclure le détail des paiements enregistrés"
    )

    inclure_infos_paiement = forms.BooleanField(
        required=False,
        initial=True,
        label="Inclure les informations de paiement du circuit"
    )

    infos_paiement_custom = forms.CharField(
        required=False,
        label="Informations de paiement / consignes",
        widget=forms.Textarea(attrs={
            "rows": 5,
            "placeholder": "Coordonnées bancaires, référence à utiliser, consignes manuelles..."
        })
    )

    note_client = forms.CharField(
        required=False,
        label="Message complémentaire",
        widget=forms.Textarea(attrs={
            "rows": 5,
            "placeholder": "Ex : Nous avons bien reçu votre règlement complémentaire lié aux nouvelles options."
        })
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for _, field in self.fields.items():
            widget = field.widget
            css = widget.attrs.get("class", "")

            if isinstance(widget, forms.TextInput):
                widget.attrs["class"] = f"{css} paysum-input".strip()
            elif isinstance(widget, forms.Textarea):
                widget.attrs["class"] = f"{css} paysum-textarea".strip()
            elif isinstance(widget, forms.Select):
                widget.attrs["class"] = f"{css} paysum-select".strip()
            elif isinstance(widget, forms.NumberInput):
                widget.attrs["class"] = f"{css} paysum-input paysum-input--number".strip()
            elif isinstance(widget, forms.CheckboxInput):
                widget.attrs["class"] = f"{css} paysum-checkbox".strip()
            elif isinstance(widget, forms.DateInput):
                widget.attrs["class"] = f"{css} paysum-input".strip()

        for name in (
            "total_attendu",
            "montant_options_ajoutees",
            "nouveau_total_attendu",
            "acompte1_encaisse",
            "acompte2_encaisse",
            "solde_encaisse",
            "paiement_recu_pour_options",
            "montant_restant",
            "trop_percu",
        ):
            self.fields[name].widget.attrs.update({
                "step": "0.01",
                "inputmode": "decimal",
            })

        self.fields["recipient_role"].widget.attrs.update({
            "placeholder": "Choisir le destinataire",
        })
        self.fields["total_attendu"].widget.attrs.update({
            "placeholder": "Ex : 660.00",
        })
        self.fields["montant_options_ajoutees"].widget.attrs.update({
            "placeholder": "Ex : 45.00",
        })
        self.fields["nouveau_total_attendu"].widget.attrs.update({
            "readonly": True,
            "data-auto-computed": "1",
            "placeholder": "Calcul automatique",
        })
        self.fields["acompte1_encaisse"].widget.attrs.update({
            "placeholder": "Ex : 250.00",
        })
        self.fields["acompte2_encaisse"].widget.attrs.update({
            "placeholder": "Ex : 250.00",
        })
        self.fields["solde_encaisse"].widget.attrs.update({
            "placeholder": "Montant encaissé après acomptes, Ex : 160.00",
        })
        self.fields["paiement_recu_pour_options"].widget.attrs.update({
            "placeholder": "Ex : 45.00",
        })
        self.fields["montant_restant"].widget.attrs.update({
            "readonly": True,
            "data-auto-computed": "1",
            "placeholder": "Calcul automatique",
        })
        self.fields["trop_percu"].widget.attrs.update({
            "readonly": True,
            "data-auto-computed": "1",
            "placeholder": "Calcul automatique",
        })

    def clean(self):
        cleaned = super().clean()

        total_attendu = cleaned.get("total_attendu") or Decimal("0")
        montant_options_ajoutees = cleaned.get("montant_options_ajoutees") or Decimal("0")
        acompte1 = cleaned.get("acompte1_encaisse") or Decimal("0")
        acompte2 = cleaned.get("acompte2_encaisse") or Decimal("0")
        solde = cleaned.get("solde_encaisse") or Decimal("0")
        paiement_options = cleaned.get("paiement_recu_pour_options") or Decimal("0")

        nouveau_total = total_attendu + montant_options_ajoutees
        cleaned["nouveau_total_attendu"] = nouveau_total

        total_saisi = acompte1 + acompte2 + solde + paiement_options

        if total_saisi <= Decimal("0"):
            raise forms.ValidationError(
                "Merci de saisir au moins un montant encaissé (acompte 1, acompte 2, montant après acomptes ou paiement lié aux nouvelles options)."
            )

        montant_restant = Decimal("0")
        trop_percu = Decimal("0")
        situation_label = "Soldé"

        if total_saisi < nouveau_total:
            montant_restant = nouveau_total - total_saisi
            situation_label = "Reste à payer"
        elif total_saisi > nouveau_total:
            trop_percu = total_saisi - nouveau_total
            situation_label = "Trop-perçu"

        cleaned["total_deja_paye_calcule"] = total_saisi
        cleaned["montant_restant"] = montant_restant
        cleaned["trop_percu"] = trop_percu

        if not cleaned.get("situation_label"):
            if montant_options_ajoutees > 0 and paiement_options > 0:
                cleaned["situation_label"] = "Paiement reçu après ajout d’options"
            else:
                cleaned["situation_label"] = situation_label

        return cleaned

class PaymentReminderComposeForm(forms.Form):
    ins_id = forms.IntegerField(widget=forms.HiddenInput())

    recipient_role = forms.ChoiceField(
        choices=(
            ("pilote", "Pilote"),
            ("passager", "Passager"),
        ),
        label="Destinataire",
    )

    sujet = forms.CharField(
        max_length=180,
        label="Sujet",
        widget=forms.TextInput(attrs={
            "placeholder": "Ex : Rappel de paiement – Corse 2025",
        })
    )

    intro_message = forms.CharField(
        required=False,
        label="Message d’introduction",
        widget=forms.Textarea(attrs={
            "rows": 4,
            "placeholder": (
                "Bonjour,\n\n"
                "Nous vous rappelons qu’un solde reste à régler pour votre inscription."
            )
        })
    )

    total_attendu = forms.DecimalField(
        max_digits=12,
        decimal_places=2,
        min_value=0,
        label="Total attendu",
        widget=forms.NumberInput(attrs={
            "readonly": True,
            "step": "0.01",
            "class": "paysum-input paysum-input--number",
        }),
    )

    total_paye = forms.DecimalField(
        max_digits=12,
        decimal_places=2,
        min_value=0,
        label="Déjà payé",
        widget=forms.NumberInput(attrs={
            "readonly": True,
            "step": "0.01",
            "class": "paysum-input paysum-input--number",
        }),
    )

    montant_restant = forms.DecimalField(
        max_digits=12,
        decimal_places=2,
        min_value=0,
        label="Reste à payer",
        widget=forms.NumberInput(attrs={
            "readonly": True,
            "step": "0.01",
            "class": "paysum-input paysum-input--number",
        }),
    )

    situation_label = forms.CharField(
        max_length=120,
        required=False,
        label="État du paiement",
        widget=forms.TextInput(attrs={
            "readonly": True,
            "class": "paysum-input",
        })
    )

    inclure_detail_paiements = forms.BooleanField(
        required=False,
        initial=True,
        label="Inclure le détail des paiements enregistrés"
    )

    inclure_infos_paiement = forms.BooleanField(
        required=False,
        initial=True,
        label="Inclure les informations de paiement du circuit"
    )

    infos_paiement_custom = forms.CharField(
        required=False,
        label="Informations de paiement / consignes",
        widget=forms.Textarea(attrs={
            "rows": 5,
            "class": "paysum-textarea",
            "placeholder": "Coordonnées bancaires, référence à utiliser, consignes..."
        })
    )

    note_client = forms.CharField(
        required=False,
        label="Message complémentaire",
        widget=forms.Textarea(attrs={
            "rows": 5,
            "class": "paysum-textarea",
            "placeholder": "Ex : Merci de régulariser votre solde dans les meilleurs délais."
        })
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for _, field in self.fields.items():
            widget = field.widget
            css = widget.attrs.get("class", "")

            if isinstance(widget, forms.TextInput):
                widget.attrs["class"] = f"{css} paysum-input".strip()
            elif isinstance(widget, forms.Textarea):
                widget.attrs["class"] = f"{css} paysum-textarea".strip()
            elif isinstance(widget, forms.Select):
                widget.attrs["class"] = f"{css} paysum-select".strip()
            elif isinstance(widget, forms.NumberInput):
                widget.attrs["class"] = f"{css} paysum-input paysum-input--number".strip()
            elif isinstance(widget, forms.CheckboxInput):
                widget.attrs["class"] = f"{css} paysum-checkbox".strip()

    def clean(self):
        cleaned = super().clean()
        restant = cleaned.get("montant_restant") or Decimal("0")
        if restant <= Decimal("0"):
            raise forms.ValidationError(
                "Ce rappel n’est utile que si un montant reste à payer."
            )
        if not cleaned.get("situation_label"):
            cleaned["situation_label"] = "Reste à payer"
        return cleaned

def _build_payment_summary_defaults(ins, personne):
    paiements = _ordered_payments_for_display(ins)

    total_attendu_courant = Decimal(
        (ins.prix_pilote_unitaire or 0)
        + ((ins.prix_passager_unitaire or 0) if ins.passager_id else 0)
        + sum(sel.prix_total() for sel in ins.selections_options.all())
    )

    total_deja_paye = Decimal(sum((p.montant_encaisse or 0) for p in paiements))
    montant_restant = max(total_attendu_courant - total_deja_paye, Decimal("0"))
    trop_percu = max(total_deja_paye - total_attendu_courant, Decimal("0"))

    if montant_restant > 0:
        situation_label = "Reste à payer"
    elif trop_percu > 0:
        situation_label = "Trop-perçu"
    else:
        situation_label = "Soldé"

    recipient_role = "pilote"
    if ins.passager_id == personne.pk:
        recipient_role = "passager"

    recipient = ins.pilote if recipient_role == "pilote" else ins.passager

    # ---------------------------------------------------------
    # Répartition VISUELLE des montants déjà encaissés
    # dans l'ordre métier : ACOMPTE1 -> ACOMPTE2 -> SOLDE
    # sans modifier la base, sans envoi, sans validation.
    # ---------------------------------------------------------
    acompte1_val = Decimal("0")
    acompte2_val = Decimal("0")
    solde_val = Decimal("0")

    total_deja_paye_affichage = Decimal(sum((p.montant_encaisse or 0) for p in paiements))

    acompte1_du = Decimal("0")
    acompte2_du = Decimal("0")
    solde_du = Decimal("0")

    for p in paiements:
        montant_du = Decimal(p.montant_du or 0)
        if p.libelle == LibellePaiement.ACOMPTE1:
            acompte1_du = montant_du
        elif p.libelle == LibellePaiement.ACOMPTE2:
            acompte2_du = montant_du
        elif p.libelle == LibellePaiement.SOLDE:
            solde_du = montant_du

    remaining = total_deja_paye_affichage

    if acompte1_du > 0 and remaining > 0:
        acompte1_val = min(remaining, acompte1_du)
        remaining = max(remaining - acompte1_val, Decimal("0"))

    if acompte2_du > 0 and remaining > 0:
        acompte2_val = min(remaining, acompte2_du)
        remaining = max(remaining - acompte2_val, Decimal("0"))

    if solde_du > 0 and remaining > 0:
        solde_val = min(remaining, solde_du)
        remaining = max(remaining - solde_val, Decimal("0"))

    # Sécurité : si un reliquat subsiste, on l'ajoute seulement à l'affichage du solde
    if remaining > 0:
        solde_val += remaining

    stored_options_added = Decimal(getattr(ins, "montant_options_ajoutees", 0) or 0)
    options_added = stored_options_added if stored_options_added > 0 else Decimal("0")

    if options_added > total_attendu_courant:
        options_added = Decimal("0")

    total_initial = (
        total_attendu_courant - options_added
        if options_added > 0 else total_attendu_courant
    )

    has_acompte1 = Decimal(ins.circuit.acompte1_par_personne or 0) > 0
    has_acompte2 = Decimal(ins.circuit.acompte2_par_personne or 0) > 0

    return {
        "ins_id": ins.pk,
        "recipient_role": recipient_role,
        "sujet": f"[{ins.circuit.code}] Point sur votre paiement – {ins.circuit.nom}",
        "intro_message": (
            "Bonjour,\n\n"
            "Voici votre récapitulatif de paiement actuel concernant votre inscription."
        ),
        "total_attendu": total_initial,
        "montant_options_ajoutees": options_added,
        "nouveau_total_attendu": total_attendu_courant,

        "acompte1_encaisse": acompte1_val,
        "acompte2_encaisse": acompte2_val,
        "solde_encaisse": solde_val,
        "paiement_recu_pour_options": Decimal(getattr(ins, "paiement_options_recu", 0) or 0),

        "date_paiement": timezone.localdate(),
        "montant_restant": montant_restant,
        "trop_percu": trop_percu,
        "situation_label": situation_label,
        "inclure_detail_paiements": True,
        "inclure_infos_paiement": True,
        "infos_paiement_custom": (getattr(ins.circuit, "infos_paiement", "") or "").strip(),
        "note_client": "",
        "recipient_email": getattr(recipient, "email", "") if recipient else "",
        "recipient_name": getattr(recipient, "nom_complet", "") if recipient else "",
        "paiements": paiements,

        "has_acompte1": has_acompte1,
        "has_acompte2": has_acompte2,
        "has_options_added": options_added > 0,
    }

def _build_email_payment_rows(ins):
    """
    Construit des lignes de paiement pour l'email sans modifier la base.
    Répartit le montant payé sur les échéances dans l'ordre.
    """
    rows = []
    for p in ins.paiements.all().order_by("echeance_le", "pk"):
        attendu = Decimal(p.montant_du or 0)
        encaisse = Decimal(p.montant_encaisse or 0)

        if encaisse >= attendu and attendu > 0:
            statut_label = "Payé"
        elif encaisse > 0:
            statut_label = "Partiellement payé"
        else:
            statut_label = "À payer"

        rows.append({
            "libelle": p.get_libelle_display(),
            "attendu": attendu,
            "encaisse": encaisse,
            "date": p.encaisse_le,
            "statut": statut_label,
        })
    return rows

class CircuitForm(forms.ModelForm):
    # on déclare les champs date pour contrôler rendu/parsing

    image = forms.ImageField(
        label="Image du circuit (optionnelle)",
        required=False,
        widget=forms.ClearableFileInput(attrs={"accept": "image/*"})
    )

    date_debut = forms.DateField(
        label="Date de début",
        widget=forms.DateInput(attrs={"type": "date"}, format="%Y-%m-%d"),
        input_formats=["%Y-%m-%d", "%d/%m/%Y"],
        required=True,
    )
    date_fin = forms.DateField(
        label="Date de fin",
        widget=forms.DateInput(attrs={"type": "date"}, format="%Y-%m-%d"),
        input_formats=["%Y-%m-%d", "%d/%m/%Y"],
        required=True,
    )

    class Meta:
        model = Circuit
        fields = [
            "nom", "code", "image", "date_debut", "date_fin", "description", "infos_rdv",
            "capacite", "devise", "prix_pilote_unitaire", "prix_passager_unitaire",
            "delai_option_jours", "statut",
            "acompte1_par_personne", "acompte2_par_personne",
            "delai_acompte2_jours", "delai_solde_jours",
            "programme", "kilometrages", "prix_comprend", "prix_ne_comprend_pas",
            "echeancier_texte", "infos_paiement",

            # ✅ NOUVEAU : overrides email (optionnels)
            "email_rdv",
            "email_organisation",
            "email_checklist",
        ]
        widgets = {
            "description": forms.Textarea(attrs={"rows": 4, "placeholder": "Résumé clair, programme, conditions…"}),
            "infos_rdv": forms.Textarea(attrs={"rows": 3, "placeholder": "Lieu précis, heure de RDV, contact…"}),

            "infos_paiement": forms.Textarea(
                attrs={
                    "rows": 4,
                    "placeholder": "Ex: IBAN: … / BIC: … / Intitulé: Pulsion Horizon / Réf: {{ID_PUBLIC}} / Rappel: solde J-30…"
                }
            ),

            # ✅ NOUVEAU : champs email dynamiques
            "email_rdv": forms.Textarea(
                attrs={
                    "rows": 3,
                    "placeholder": "Override RDV/Logistique (laisser vide = texte standard)."
                }
            ),
            "email_organisation": forms.Textarea(
                attrs={
                    "rows": 3,
                    "placeholder": "Override Organisation (laisser vide = texte standard)."
                }
            ),
            "email_checklist": forms.Textarea(
                attrs={
                    "rows": 6,
                    "placeholder": "Override Checklist : 1 ligne = 1 item.\n"
                                   "Ex:\nPasseport\nVisa imprimé\nAnti-moustiques\nAdaptateur prise"
                }
            ),
        }

    def clean_image(self):
        img = self.cleaned_data.get("image")
        if img:
            max_mb = 8
            if img.size > max_mb * 1024 * 1024:
                raise forms.ValidationError(f"Image trop volumineuse (max {max_mb} Mo).")
        return img

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # pas de localisation auto sur ces inputs
        for name in ("date_debut", "date_fin"):
            f = self.fields[name]
            f.localize = False
            f.widget.is_localized = False

        # en édition (GET), on pré-remplit en ISO pour l’input date
        if not self.data:
            for name in ("date_debut", "date_fin"):
                val = getattr(self.instance, name, None)
                if isinstance(val, (datetime.date, datetime.datetime)):
                    self.initial[name] = val.strftime("%Y-%m-%d")

    def clean(self):
        cleaned = super().clean()
        dd = cleaned.get("date_debut")
        df = cleaned.get("date_fin")
        cap = cleaned.get("capacite")
        ppu = cleaned.get("prix_pilote_unitaire")
        ppa = cleaned.get("prix_passager_unitaire")
        delai = cleaned.get("delai_option_jours")

        if dd and df and df < dd:
            self.add_error("date_fin", "La date de fin doit être ≥ la date de début.")
        if cap is not None and cap <= 0:
            self.add_error("capacite", "La capacité doit être positive.")
        if ppu is not None and ppu < 0:
            self.add_error("prix_pilote_unitaire", "Prix invalide.")
        if ppa is not None and ppa < 0:
            self.add_error("prix_passager_unitaire", "Prix invalide.")
        if delai is not None and delai < 0:
            self.add_error("delai_option_jours", "Doit être ≥ 0.")

        a1 = cleaned.get("acompte1_par_personne") or 0
        a2 = cleaned.get("acompte2_par_personne") or 0
        d2 = cleaned.get("delai_acompte2_jours") or 0
        ds = cleaned.get("delai_solde_jours") or 0
        if a1 < 0 or a2 < 0:
            if a1 < 0:
                self.add_error("acompte1_par_personne", "Montant invalide.")
            if a2 < 0:
                self.add_error("acompte2_par_personne", "Montant invalide.")
        if d2 < 0 or ds < 0:
            if d2 < 0:
                self.add_error("delai_acompte2_jours", "Doit être ≥ 0.")
            if ds < 0:
                self.add_error("delai_solde_jours", "Doit être ≥ 0.")

        return cleaned
    
class EmailingForm(forms.Form):
    circuit = forms.ModelChoiceField(queryset=Circuit.objects.all().order_by("-date_debut"), required=False, label="Filtrer par circuit")
    statut = forms.ChoiceField(choices=[("", "— Tous statuts —")] + list(StatutInscription.choices), required=False, label="Filtrer par statut d’inscription")
    inclure_passagers = forms.BooleanField(required=False, initial=True, label="Inclure les passagers")
    sujet = forms.CharField(max_length=180, label="Sujet")
    message = forms.CharField(widget=forms.Textarea(attrs={"rows":8}), label="Message (texte)")
    test_only = forms.BooleanField(required=False, initial=False, label="N’envoyer qu’à moi (test)")

    pieces_jointes = forms.FileField(
        label="Pièces jointes",
        required=False,
        widget=MultipleFilesInput(attrs={
            "multiple": True,
            "accept": ".pdf,.doc,.docx,.jpg,.jpeg,.png,.zip",
        }),
        help_text="Vous pouvez sélectionner plusieurs fichiers (max ~25 Mo par email)."
    )

    def clean_pieces_jointes(self):
        files = self.files.getlist("pieces_jointes")
        max_size = 25 * 1024 * 1024  # 25 Mo (par email c’est la limite SMTP courante)
        for f in files:
            if f.size > max_size:
                raise forms.ValidationError(
                    f"“{f.name}” est trop volumineux (> 25 Mo)."
                )
        return files    

@staff_member_required
def dashboard(request):
    now = timezone.localdate()

    # =========================
    # Pagination params
    # =========================
    recents_page = int(request.GET.get("recents_page", 1) or 1)
    docs_page = int(request.GET.get("docs_page", 1) or 1)
    contacts_page = int(request.GET.get("contacts_page", 1) or 1)
    newsletter_page = int(request.GET.get("newsletter_page", 1) or 1)

    ajax_section = (request.GET.get("section") or "").strip()

    # =========================
    # KPIs globaux
    # =========================
    nb_circuits = Circuit.objects.count()
    nb_circuits_a_venir = Circuit.objects.filter(date_debut__gte=now).count()
    nb_inscriptions = Inscription.objects.count()
    nb_inscriptions_validees = Inscription.objects.filter(statut=StatutInscription.VALIDE).count()
    nb_personnes = Personne.objects.count()

    nb_documents_attente = Document.objects.filter(statut=StatutDocument.EN_ATTENTE).count()
    nb_contacts_non_traites = ContactMessage.objects.filter(traite=False).count()
    nb_newsletter_total = NewsletterSubscriber.objects.count()
    nb_newsletter_active = NewsletterSubscriber.objects.filter(is_active=True).count()

    # =========================
    # Balades 1 jour
    # =========================
    nb_balades_total = BaladeJour.objects.count()
    nb_balades_actives = BaladeJour.objects.filter(actif=True).count()
    nb_demandes_balades = DemandeProgrammeBalade.objects.count()
    nb_demandes_balades_non_traitees = DemandeProgrammeBalade.objects.filter(traite=False).count()    

    # Paiements
    paiements_qs = Paiement.objects.all()
    total_attendu = paiements_qs.aggregate(v=Coalesce(Sum("montant_du"), 0))["v"] or 0
    total_encaisse = paiements_qs.aggregate(v=Coalesce(Sum("montant_encaisse"), 0))["v"] or 0
    total_restant = max((total_attendu or 0) - (total_encaisse or 0), 0)

    taux_validation = round((nb_inscriptions_validees / nb_inscriptions) * 100) if nb_inscriptions else 0
    taux_encaissement = round((total_encaisse / total_attendu) * 100) if total_attendu else 0

    # =========================
    # Prochains circuits
    # =========================
    prochains = (
        Circuit.objects
        .filter(date_debut__gte=now, statut=StatutCircuit.PUBLIE)
        .annotate(nb=Count("inscriptions"))
        .order_by("date_debut")[:12]
    )

    # =========================
    # Prochaines balades 1 jour
    # =========================
    prochaines_balades = (
        BaladeJour.objects
        .filter(actif=True, date_debut__gte=now)
        .order_by("date_debut", "ordre", "titre")[:8]
    )

    for b in prochaines_balades:
        if hasattr(b, "date_humaine"):
            b.date_label = b.date_humaine()
        else:
            b.date_label = b.date_affichage or b.date_debut.strftime("%d/%m/%Y")

        b.est_bientot = bool(b.date_debut and (b.date_debut - now).days <= 15)


    for c in prochains:
        cap = c.capacite or 0
        c.pct = int(round((c.nb / cap) * 100)) if cap else 0
        c.places_restantes = max(cap - c.nb, 0) if cap else 0
        c.est_bientot = bool(c.date_debut and (c.date_debut - now).days <= 15)

    # =========================
    # Circuits par statut
    # =========================
    base_circuits = Circuit.objects.annotate(nb=Count("inscriptions"))
    publies = base_circuits.filter(statut=StatutCircuit.PUBLIE).order_by("date_debut")[:8]
    archives = base_circuits.filter(statut=StatutCircuit.ARCHIVE).order_by("-date_debut")[:8]
    brouillons = base_circuits.filter(statut=StatutCircuit.BROUILLON).order_by("-modifie_le")[:8]

    # =========================
    # Inscriptions récentes
    # =========================
    inscriptions_recentes_qs = (
        Inscription.objects
        .select_related("circuit", "pilote", "passager")
        .order_by("-cree_le")
    )
    inscriptions_recentes_page = Paginator(inscriptions_recentes_qs, 6).get_page(recents_page)

    # =========================
    # Documents en attente
    # =========================
    documents_attente_qs = (
        Document.objects
        .filter(statut=StatutDocument.EN_ATTENTE)
        .select_related("inscription", "inscription__circuit", "inscription__pilote")
        .order_by("cree_le")
    )
    documents_attente_page = Paginator(documents_attente_qs, 6).get_page(docs_page)

    # =========================
    # Contacts récents
    # =========================
    contacts_recents_qs = (
        ContactMessage.objects
        .filter(traite=False)
        .select_related("circuit")
        .order_by("-cree_le")
    )
    contacts_recents_page = Paginator(contacts_recents_qs, 6).get_page(contacts_page)

    # =========================
    # Newsletter
    # =========================
    newsletter_qs = (
        NewsletterSubscriber.objects
        .order_by("-created_at", "-id")
    )
    newsletter_page_obj = Paginator(newsletter_qs, 8).get_page(newsletter_page)

    # =========================
    # Alertes
    # =========================
    circuits_complets = 0
    circuits_quasi_complets = 0

    circuits_upcoming_all = (
        Circuit.objects
        .filter(date_debut__gte=now, statut=StatutCircuit.PUBLIE)
        .annotate(nb=Count("inscriptions"))
    )

    for c in circuits_upcoming_all:
        cap = c.capacite or 0
        if not cap:
            continue
        pct = (c.nb / cap) * 100
        if pct >= 100:
            circuits_complets += 1
        elif pct >= 80:
            circuits_quasi_complets += 1

    stats = {
        "nb_circuits": nb_circuits,
        "nb_circuits_a_venir": nb_circuits_a_venir,
        "nb_inscriptions": nb_inscriptions,
        "nb_inscriptions_validees": nb_inscriptions_validees,
        "nb_personnes": nb_personnes,
        "nb_documents_attente": nb_documents_attente,
        "nb_contacts_non_traites": nb_contacts_non_traites,
        "nb_newsletter_total": nb_newsletter_total,
        "nb_newsletter_active": nb_newsletter_active,
        "total_attendu": total_attendu,
        "total_encaisse": total_encaisse,
        "total_restant": total_restant,
        "taux_validation": taux_validation,
        "taux_encaissement": taux_encaissement,
        "circuits_complets": circuits_complets,
        "circuits_quasi_complets": circuits_quasi_complets,
        "nb_balades_total": nb_balades_total,
        "nb_balades_actives": nb_balades_actives,
        "nb_demandes_balades": nb_demandes_balades,
        "nb_demandes_balades_non_traitees": nb_demandes_balades_non_traitees,
    }

    # =========================
    # AJAX partial rendering
    # =========================
    if ajax_section == "recents":
        return render(request, "circuitMoto/admin/partials/dashboard/_inscriptions_recentes.html", {
            "inscriptions_recentes_page": inscriptions_recentes_page,
        })

    if ajax_section == "documents":
        return render(request, "circuitMoto/admin/partials/dashboard/_documents_attente.html", {
            "documents_attente_page": documents_attente_page,
        })

    if ajax_section == "contacts":
        return render(request, "circuitMoto/admin/partials/dashboard/_contacts_recents.html", {
            "contacts_recents_page": contacts_recents_page,
        })

    if ajax_section == "newsletter":
        return render(request, "circuitMoto/admin/partials/dashboard/_newsletter_list.html", {
            "newsletter_page_obj": newsletter_page_obj,
        })

    return render(request, "circuitMoto/admin/dashboard.html", {
        "stats": stats,
        "prochains": prochains,
        "publies": publies,
        "archives": archives,
        "brouillons": brouillons,
        "inscriptions_recentes_page": inscriptions_recentes_page,
        "documents_attente_page": documents_attente_page,
        "contacts_recents_page": contacts_recents_page,
        "newsletter_page_obj": newsletter_page_obj,
        "prochaines_balades": prochaines_balades,
    })

class OptionForm(forms.ModelForm):
    class Meta:
        model = OptionCircuit
        fields = [
            "code", "intitule", "categorie",
            "prix_unitaire", "facture_par_personne",
            "quantite_variable", "quantite_min", "quantite_max", "actif",
        ]
        widgets = {
            "code": forms.TextInput(attrs={"placeholder": "ex: cabine-2lits"}),
            "intitule": forms.TextInput(attrs={"placeholder": "Ex: Cabine 2 lits"}),
            "prix_unitaire": forms.NumberInput(attrs={"step": "1", "min": "0"}),
        }
        labels = {
            "code": "Code option (slug unique par circuit)",
            "intitule": "Intitulé",
            "categorie": "Catégorie",
            "prix_unitaire": "Prix unitaire",
            "facture_par_personne": "Facturer par personne (pilote/passager)",
            "quantite_variable": "Quantité variable (xN)",
            "quantite_min": "Quantité min.",
            "quantite_max": "Quantité max (0 = illimitée)",
        }

    def clean(self):
        cleaned = super().clean()
        code = cleaned.get("code") or ""
        intitule = cleaned.get("intitule") or ""
        qmin = cleaned.get("quantite_min") or 0
        qmax = cleaned.get("quantite_max") or 0

        # Auto-slug du code si vide
        if not code and intitule:
            cleaned["code"] = slugify(intitule)

        # Cohérence min/max (0 = illimité)
        if qmax and qmin and qmax < qmin:
            self.add_error("quantite_max", "La quantité max doit être ≥ min (ou 0 pour illimité).")

        return cleaned

OptionFormSet = inlineformset_factory(
    parent_model=Circuit,
    model=OptionCircuit,
    form=OptionForm,
    fields=[
        "code", "intitule", "categorie",
        "prix_unitaire", "facture_par_personne",
        "quantite_variable", "quantite_min", "quantite_max",
    ],
    extra=0,
    can_delete=True,
    validate_min=False,
    validate_max=False,
)

# ---------- Circuits ----------
@staff_member_required
def circuit_list(request):
    q = request.GET.get("q", "").strip()

    # Tri (mapping depuis UI)
    sort = request.GET.get("sort", "-date")
    sort_map = {
        "-date": "-date_debut",
        "date": "date_debut",
        "nom": "nom",
        "-cap": "-capacite",
        "cap": "capacite",
    }
    order_by = sort_map.get(sort, "-date_debut")

    # Queryset
    qs = (Circuit.objects
          .all()
          .annotate(nb=Count("inscriptions"))
          .order_by(order_by))

    if q:
        qs = qs.filter(
            Q(nom__icontains=q) |
            Q(code__icontains=q) |
            Q(description__icontains=q)
        )

    # Taille de page
    try:
        per = int(request.GET.get("per", 20))
        per_page = max(5, min(per, 100))
    except (TypeError, ValueError):
        per_page = 20

    paginator = Paginator(qs, per_page)
    page = request.GET.get("page", 1)

    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # Querystring commun (sans page) pour les liens de pagination
    params = request.GET.copy()
    params.pop("page", None)
    querystring = params.urlencode()

    return render(request, "circuitMoto/admin/circuits_list.html", {
        "circuits": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "is_paginated": page_obj.has_other_pages(),
        "q": q,
        "sort": sort,
        "per": per_page,
        "querystring": querystring,
    })

@staff_member_required
@transaction.atomic
def circuit_create(request):
    if request.method == "POST":
        form = CircuitForm(request.POST, request.FILES)  # ✅ FILES
        if form.is_valid():
            uploaded_image = form.cleaned_data.get("image")

            circuit = form.save(commit=False)

            # ✅ Ne pas écrire le fichier tant que les options ne sont pas validées
            if uploaded_image:
                circuit.image = None

            circuit.save()

            formset = OptionFormSet(request.POST, instance=circuit, prefix="opts")
            if formset.is_valid():
                formset.save()

                # ✅ Maintenant seulement, on stocke l'image
                if uploaded_image:
                    circuit.image = uploaded_image
                    circuit.save()

                messages.success(request, "Circuit créé avec ses options.")
                return redirect("bo_circuit_list")
            else:
                # options invalides : on supprime le circuit (et aucun fichier n'a été stocké)
                circuit.delete()
        else:
            formset = OptionFormSet(request.POST, prefix="opts")
    else:
        form = CircuitForm()
        formset = OptionFormSet(prefix="opts")

    return render(request, "circuitMoto/admin/circuits_form.html", {
        "form": form, "formset": formset, "mode": "create"
    })

@staff_member_required
def circuit_detail(request, pk: int):
    qs = (Circuit.objects
          .prefetch_related(
              Prefetch("options", queryset=OptionCircuit.objects.order_by("categorie","intitule")),
              Prefetch("inscriptions", queryset=Inscription.objects.select_related("pilote","passager").order_by("-cree_le")),
          )
          .annotate(nb_inscriptions=Count("inscriptions")))
    circuit = get_object_or_404(qs, pk=pk)  # ← au lieu de .get(pk=pk)

    capacite = circuit.capacite or 0
    pct = int(round((circuit.nb_inscriptions / capacite) * 100)) if capacite else 0

    return render(request, "circuitMoto/admin/circuits_detail.html", {"c": circuit, "pct": pct})

@staff_member_required
@transaction.atomic
def circuit_edit(request, pk):
    obj = get_object_or_404(Circuit, pk=pk)

    if request.method == "POST":
        # ✅ IMPORTANT : passer request.FILES
        form = CircuitForm(request.POST, request.FILES, instance=obj)
        formset = OptionFormSet(request.POST, instance=obj, prefix="opts")

        if form.is_valid() and formset.is_valid():
            form.save()  # ✅ sauvegarde aussi l'image
            try:
                formset.save()
            except ProtectedError:
                to_archive_ids = []
                for f in formset.forms:
                    if f.cleaned_data.get("DELETE") and f.instance.pk:
                        to_archive_ids.append(f.instance.pk)
                if to_archive_ids:
                    (OptionCircuit.objects.filter(pk__in=to_archive_ids, circuit=obj)
                        .update(actif=False, archive_le=timezone.now()))
                    messages.warning(
                        request,
                        "Certaines options déjà utilisées ont été ARCHIVÉES (masquées) au lieu d’être supprimées."
                    )

            messages.success(request, "Circuit mis à jour.")
            return redirect("bo_circuit_list")

    else:
        form = CircuitForm(instance=obj)
        formset = OptionFormSet(instance=obj, prefix="opts")

    return render(request, "circuitMoto/admin/circuits_form.html", {
        "form": form, "formset": formset, "mode": "edit", "obj": obj
    })

# ---------- Inscriptions ----------
@staff_member_required
def inscription_list(request):
    """
    Liste des inscriptions + module d'export:
      - Années d'export = uniquement les années où il y a des inscrits
      - Si une année est choisie, Circuits d'export = uniquement circuits avec inscrits sur cette année
      - Prend en charge GET ?year=... & ?circuit=... (noms des champs du formulaire d’export)
      - Recherche multi-termes, tri whitelisté, pagination robuste
    """

    # -------- Helpers --------
    def safe_int(raw, default=None):
        try:
            return int(raw)
        except (TypeError, ValueError):
            return default

    ALLOWED_ORDERS = {
        "-cree_le", "cree_le",
        "pilote__nom", "-pilote__nom",
        "circuit__nom", "-circuit__nom",
        "circuit__code", "-circuit__code",
        "circuit__date_debut", "-circuit__date_debut",
    }

    def ordering_for(o: str) -> List[str]:
        if o in {"circuit__nom", "-circuit__nom"}:
            return [o, "pilote__nom", "-cree_le", "id"]
        if o in {"circuit__code", "-circuit__code"}:
            return [o, "-cree_le", "id"]
        if o in {"circuit__date_debut", "-circuit__date_debut"}:
            return [o, "circuit__code", "-cree_le", "id"]
        if o in {"pilote__nom", "-pilote__nom"}:
            return [o, "cree_le", "id"]
        return [o, "id"]

    # -------- Entrées GET "nettoyées" --------
    q_raw = (request.GET.get("q") or "").strip()
    q = q_raw[:200] if len(q_raw) > 200 else q_raw

    allowed_status = {k for k, _ in getattr(StatutInscription, "choices", [])} or {
        "OPTION", "VALIDE", "ANNULEE", "TERMINEE"
    }
    statut_raw = (request.GET.get("statut") or "").strip().upper()
    statut = statut_raw if statut_raw in allowed_status else ""

    order_param = (request.GET.get("order") or "-cree_le").strip()
    order = order_param if order_param in ALLOWED_ORDERS else "-cree_le"

    page = safe_int(request.GET.get("page"), 1) or 1
    page_size = safe_int(request.GET.get("page_size"), 20)
    if not page_size or page_size <= 0 or page_size > 200:
        page_size = 20

    # -------- Queryset principal --------
    qs = (
        Inscription.objects
        .select_related("circuit", "pilote", "passager", "assurance")
        .order_by(*ordering_for(order))
    )

    if q:
        terms = [t for t in q.split() if t]
        for t in terms:
            qs = qs.filter(
                Q(pilote__nom__icontains=t) |
                Q(pilote__prenom__icontains=t) |
                Q(passager__nom__icontains=t) |
                Q(passager__prenom__icontains=t) |
                Q(circuit__nom__icontains=t) |
                Q(circuit__code__icontains=t) |
                Q(id_public__icontains=t)
            )

    if statut:
        qs = qs.filter(statut=statut)

    paginator = Paginator(qs, page_size)
    page_obj = paginator.get_page(page)

    base_qs = {k: v for k, v in request.GET.items() if k.lower() != "page"}
    preserved = urlencode(base_qs)

    circuits = Circuit.objects.order_by("-date_debut")

    # =====================  MODULE D’EXPORT  =====================
    # Objectif:
    #  - export_years: années avec au moins un inscrit (respecte statut si présent)
    #  - export_circuits: circuits avec au moins un inscrit ; si une année est choisie,
    #    on ne garde que les circuits ayant des inscrits sur cette année.
    #
    # IMPORTANT: on accepte *les deux* familles de paramètres pour compatibilité:
    #   * exp_year / exp_circuit (anciens)
    #   * year / circuit (noms des champs du <form> d’export)
    exp_year_q     = (request.GET.get("exp_year")     or request.GET.get("year")    or "").strip()
    exp_circuit_q  = (request.GET.get("exp_circuit")  or request.GET.get("circuit") or "").strip()

    exp_year_val    = safe_int(exp_year_q, None)
    exp_circuit_val = safe_int(exp_circuit_q, None)

    base_exp = Inscription.objects.all().filter(
        circuit__isnull=False,
        circuit__date_debut__isnull=False,
    )
    if statut:
        base_exp = base_exp.filter(statut=statut)

    # Filtrage croisé
    if exp_year_val is not None:
        base_exp = base_exp.filter(circuit__date_debut__year=exp_year_val)
    if exp_circuit_val is not None:
        base_exp = base_exp.filter(circuit_id=exp_circuit_val)

    # Années avec inscrits (sur base_exp)
    export_years = (
        base_exp
        .annotate(y=ExtractYear("circuit__date_debut"))
        .values_list("y", flat=True)
        .distinct()
        .order_by("-y")
    )

    # Circuits avec inscrits (si année choisie, base_exp est déjà restreinte)
    circuit_ids = base_exp.values_list("circuit_id", flat=True).distinct()
    export_circuits = Circuit.objects.filter(id__in=circuit_ids).order_by("-date_debut")

    # -------- Render --------
    return render(
        request,
        "circuitMoto/admin/inscriptions_list.html",
        {
            "page_obj": page_obj,
            "q": q,
            "statut": statut,
            "order": order,
            "preserved": preserved,
            "circuits": circuits,

            # Données pour l’export
            "export_years": export_years,
            "export_circuits": export_circuits,
            # On renvoie les valeurs qui pilotent l'état sélectionné du <select>
            "exp_year": exp_year_q,          # affichage sélection année
            "exp_circuit": exp_circuit_q,    # affichage sélection circuit
        },
    )

# ---------- Personnes ----------
@staff_member_required
def personne_list(request):
    q = request.GET.get("q","").strip()

    qs = (Personne.objects
          .all()
          .order_by("nom","prenom")
          .prefetch_related(
              "motos",
              "contacts_urgence",
              "inscriptions_pilote__circuit",
              "inscriptions_passager__circuit",
          ))

    if q:
        qs = qs.filter(
            Q(nom__icontains=q) | Q(prenom__icontains=q) |
            Q(email__icontains=q) | Q(telephone__icontains=q)
        )

    # circuits utiles pour l’enrôlement rapide
    circuits = Circuit.objects.order_by("-date_debut")

    return render(
        request,
        "circuitMoto/admin/personnes_list.html",
        {"personnes": qs, "q": q, "circuits": circuits}
    )

def _is_ajax(request):
    return request.headers.get("x-requested-with") == "XMLHttpRequest"

def _json_ok(message="", **payload):
    return JsonResponse({"ok": True, "message": message, **payload})

def _json_err(message="Erreur inconnue", status=400, **payload):
    return JsonResponse({"ok": False, "message": message, **payload}, status=status)



# ---------- Détail d’une personne ----------
from .emails import notify_paiement_resume, notify_paiement_resume_compose, _abs_url

@staff_member_required
def personne_detail(request, pk):
    p = get_object_or_404(Personne, pk=pk)

    # === ACTIONS POST (documents / inscription) ===============================
    if request.method == "POST":
        action = request.POST.get("action")
        is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"

        try:
            # Limiter aux inscriptions de cette personne (pilote OU passager)
            ins_qs = Inscription.objects.filter(Q(pilote=p) | Q(passager=p))

            # ---- Démarrer le processus d'inscription depuis la fiche
            if action == "start_inscription":
                circuit_id = (request.POST.get("circuit_id") or "").strip()
                if not circuit_id:
                    msg = "Sélectionnez un circuit."
                    return _json_err(msg) if is_ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                circuit = Circuit.objects.filter(pk=circuit_id).first()
                if not circuit:
                    msg = "Circuit introuvable."
                    return _json_err(msg) if is_ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                # Redirection vers le wizard d'inscription pré-rempli
                url = f"{reverse('inscription_create_start')}?seed={p.pk}&circuit={circuit.pk}"
                if is_ajax:
                    return _json_ok("Redirection vers l'assistant d'inscription…", redirect_url=url)
                return redirect(url)

            # ---- Documents
            if action in {"doc_validate", "doc_refuse", "doc_reset"}:
                doc_id = request.POST.get("doc_id")
                doc = (
                    Document.objects.filter(pk=doc_id, inscription__in=ins_qs)
                    .select_related("inscription", "verifie_par")
                    .first()
                )
                if not doc:
                    msg = "Document introuvable pour cette personne."
                    return _json_err(msg) if is_ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                if action == "doc_validate":
                    doc.valider(getattr(request, "user", None))
                    JournalAudit.objects.create(
                        inscription=doc.inscription,
                        acteur=request.user,
                        action="DOCUMENT_VALIDE",
                        donnees={"document_id": doc.pk, "type": doc.type_document},
                    )
                    msg = "Document validé."

                elif action == "doc_refuse":
                    raison = request.POST.get("raison", "")
                    notify_document_refuse(doc)
                    doc.refuser(raison, getattr(request, "user", None))
                    JournalAudit.objects.create(
                        inscription=doc.inscription,
                        acteur=request.user,
                        action="DOCUMENT_REFUSE",
                        donnees={"document_id": doc.pk, "raison": raison},
                    )
                    msg = "Document refusé."

                else:  # doc_reset
                    doc.statut = StatutDocument.EN_ATTENTE
                    doc.verifie_par = None
                    doc.verifie_le = None
                    doc.note = ""
                    doc.save(update_fields=["statut", "verifie_par", "verifie_le", "note", "modifie_le"])
                    JournalAudit.objects.create(
                        inscription=doc.inscription,
                        acteur=request.user,
                        action="DOCUMENT_REINITIALISE",
                        donnees={"document_id": doc.pk},
                    )
                    msg = "Document réinitialisé en 'En attente'."

                payload = {
                    "doc": {
                        "id": doc.pk,
                        "statut": doc.statut,
                        "statut_label": doc.get_statut_display(),
                        "note": doc.note or "",
                        "verifie_par": (
                            (doc.verifie_par.get_full_name() or doc.verifie_par.username or doc.verifie_par.email)
                            if doc.verifie_par else ""
                        ),
                        "verifie_le": timezone.localtime(doc.verifie_le).strftime("%d/%m/%Y %H:%M") if doc.verifie_le else "",
                    }
                }
                return _json_ok(msg, **payload) if is_ajax else (messages.success(request, msg) or redirect("bo_personne_detail", pk=p.pk))

            # ---- Inscriptions : valider une inscription OPTION
            if action == "ins_validate":
                ins_id = request.POST.get("ins_id")
                ins = ins_qs.filter(pk=ins_id).select_related("circuit").first()
                if not ins:
                    msg = "Inscription introuvable pour cette personne."
                    return _json_err(msg) if is_ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                ins.statut = StatutInscription.VALIDE
                ins.save(update_fields=["statut", "modifie_le"])
                initialiser_plan_paiement_par_defaut(ins)
                notify_inscription_validee(ins)
                JournalAudit.objects.create(
                    inscription=ins,
                    acteur=request.user,
                    action="INSCRIPTION_VALIDEE",
                    donnees={"inscription_id": ins.pk},
                )
                msg = "Inscription validée."
                payload = {
                    "ins": {"id": ins.pk, "statut": ins.statut, "statut_label": ins.get_statut_display()},
                    "hide_submitter": True,  # ✅ optionnel
                }
                return _json_ok(msg, **payload) if is_ajax else (messages.success(request, msg) or redirect("bo_personne_detail", pk=p.pk))


            # ---- Inscriptions : envoyer un récapitulatif de paiement
            if action == "ins_payment_summary_preview":
                ins_id = request.POST.get("ins_id")
                ins = (
                    ins_qs.filter(pk=ins_id)
                    .select_related("circuit", "pilote", "passager")
                    .prefetch_related("paiements", "selections_options__option")
                    .first()
                )
                if not ins:
                    msg = "Inscription introuvable pour cette personne."
                    return _json_err(msg) if is_ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                defaults = _build_payment_summary_defaults(ins, p)
                form = PaymentSummaryComposeForm(initial=defaults)

                if is_ajax:
                    html = render_to_string(
                        "circuitMoto/admin/partials/_payment_summary_modal_form.html",
                        {
                            "form": form,
                            "ins": ins,
                            "personne": p,
                            "recipient_email": defaults["recipient_email"],
                            "recipient_name": defaults["recipient_name"],
                            "paiements": defaults["paiements"],
                            "has_acompte1": defaults["has_acompte1"],
                            "has_acompte2": defaults["has_acompte2"],
                            "has_options_added": defaults["has_options_added"],
                        },
                        request=request,
                    )
                    return _json_ok("Formulaire chargé.", modal_html=html)

                messages.info(request, "Prévisualisation disponible en AJAX uniquement.")
                return redirect("bo_personne_detail", pk=p.pk)

            # if action == "ins_payment_summary_send":
            if action == "ins_payment_summary_send":
                ins_id = request.POST.get("ins_id")
                ins = (
                    ins_qs.filter(pk=ins_id)
                    .select_related("circuit", "pilote", "passager")
                    .prefetch_related("paiements", "selections_options__option")
                    .first()
                )
                if not ins:
                    msg = "Inscription introuvable pour cette personne."
                    return _json_err(msg) if is_ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                form = PaymentSummaryComposeForm(request.POST)
                if not form.is_valid():
                    if is_ajax:
                        defaults = _build_payment_summary_defaults(ins, p)

                        html = render_to_string(
                            "circuitMoto/admin/partials/_payment_summary_modal_form.html",
                            {
                                "form": form,
                                "ins": ins,
                                "personne": p,
                                "recipient_email": defaults["recipient_email"],
                                "recipient_name": defaults["recipient_name"],
                                "paiements": defaults["paiements"],
                                "has_acompte1": defaults["has_acompte1"],
                                "has_acompte2": defaults["has_acompte2"],
                                "has_options_added": defaults["has_options_added"],
                            },
                            request=request,
                        )
                        return _json_err("Merci de corriger le formulaire.", modal_html=html, status=422)

                    messages.error(request, "Merci de corriger le formulaire.")
                    return redirect("bo_personne_detail", pk=p.pk)

                recipient_role = form.cleaned_data["recipient_role"]
                recipient = ins.pilote if recipient_role == "pilote" else ins.passager

                if not recipient:
                    msg = "Destinataire introuvable."
                    return _json_err(msg) if is_ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                if not recipient.email:
                    msg = "Le destinataire n’a pas d’adresse email."
                    return _json_err(msg) if is_ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                montant_options_ajoutees = form.cleaned_data.get("montant_options_ajoutees") or Decimal("0")

                paiement_options = form.cleaned_data.get("paiement_recu_pour_options") or Decimal("0")
                nouveau_total_attendu = form.cleaned_data.get("nouveau_total_attendu") or Decimal("0")


                if montant_options_ajoutees > 0:
                    ins.montant_options_ajoutees = int(montant_options_ajoutees)

                if paiement_options > 0:
                    ins.paiement_options_recu = int(paiement_options)

                if montant_options_ajoutees > 0 or paiement_options > 0:
                    ins.save(update_fields=[
                        "montant_options_ajoutees",
                        "paiement_options_recu",
                        "modifie_le",
                    ])


                existing_acompte1 = Decimal("0")
                existing_acompte2 = Decimal("0")
                existing_solde = Decimal("0")

                for pmt in ins.paiements.all():
                    enc = Decimal(pmt.montant_encaisse or 0)
                    if pmt.libelle == LibellePaiement.ACOMPTE1:
                        existing_acompte1 = enc
                    elif pmt.libelle == LibellePaiement.ACOMPTE2:
                        existing_acompte2 = enc
                    elif pmt.libelle == LibellePaiement.SOLDE:
                        existing_solde = enc

                # 1) appliquer les paiements saisis à la base
                new_acompte1 = max(
                    Decimal(form.cleaned_data.get("acompte1_encaisse") or 0) - existing_acompte1,
                    Decimal("0")
                )
                new_acompte2 = max(
                    Decimal(form.cleaned_data.get("acompte2_encaisse") or 0) - existing_acompte2,
                    Decimal("0")
                )
                new_solde = max(
                    Decimal(form.cleaned_data.get("solde_encaisse") or 0) - existing_solde,
                    Decimal("0")
                )

                updated_rows, remaining_unallocated = _apply_manual_payments_to_inscription(
                    ins,
                    acompte1_amount=new_acompte1,
                    acompte2_amount=new_acompte2,
                    solde_amount=new_solde + paiement_options,
                    payment_date=form.cleaned_data.get("date_paiement"),
                )

                # 2) recalculer les totaux réels après mise à jour
                ins.refresh_from_db()
                paiements_after = list(ins.paiements.all().order_by("echeance_le", "pk"))

                total_paye = Decimal(sum((p.montant_encaisse or 0) for p in paiements_after))
                reste_a_payer = max(nouveau_total_attendu - total_paye, Decimal("0"))
                trop_percu = max(total_paye - nouveau_total_attendu, Decimal("0"))

                payment_rows = _build_email_payment_rows(ins) if form.cleaned_data["inclure_detail_paiements"] else []

                preview_ctx = {
                    "ins": ins,
                    "recipient": recipient,
                    "dest_name": recipient.prenom or recipient.nom or "client",
                    "circuit": ins.circuit,
                    "payment_rows": payment_rows,
                    "intro_message": form.cleaned_data["intro_message"],
                    "note_client": form.cleaned_data["note_client"],
                    "total_attendu_initial": form.cleaned_data["total_attendu"],
                    "montant_options_ajoutees": montant_options_ajoutees,
                    "nouveau_total_attendu": nouveau_total_attendu,
                    "paiement_recu_pour_options": paiement_options,
                    "total_attendu": nouveau_total_attendu,
                    "total_paye": total_paye,
                    "reste_a_payer": reste_a_payer,
                    "trop_percu": trop_percu,
                    "situation_label": form.cleaned_data["situation_label"] or "",
                    "edit_url": _abs_url(reverse("inscription_edit_start", args=[ins.id_public])),
                    "infos_paiement": form.cleaned_data["infos_paiement_custom"] if form.cleaned_data["inclure_infos_paiement"] else "",
                    "has_infos_paiement": bool(
                        form.cleaned_data["inclure_infos_paiement"] and form.cleaned_data["infos_paiement_custom"]
                    ),
                    "paiement_libelle": "Paiement manuel",
                    "subject": form.cleaned_data["sujet"],

                    # ✅ NOUVEAU : détail du paiement réellement envoyé/confirmé
                    "sent_acompte1_amount": new_acompte1,
                    "sent_acompte2_amount": new_acompte2,
                    "sent_solde_amount": new_solde,
                    "sent_options_amount": paiement_options,

                    "has_sent_acompte1": new_acompte1 > 0,
                    "has_sent_acompte2": new_acompte2 > 0,
                    "has_sent_solde": new_solde > 0,
                    "has_sent_options": paiement_options > 0,
                }

                sent_count = notify_paiement_resume_compose(
                    recipient_email=recipient.email,
                    subject=form.cleaned_data["sujet"],
                    ctx=preview_ctx,
                    role=recipient_role,
                )

                if not sent_count:
                    msg = f"L’email récapitulatif n’a pas été envoyé à {recipient.email}."
                    return (
                        _json_err(msg, status=409)
                        if is_ajax else
                        (messages.warning(request, msg) or redirect("bo_personne_detail", pk=p.pk))
                    )

                JournalAudit.objects.create(
                    inscription=ins,
                    acteur=request.user,
                    action="PAIEMENT_RESUME_ENVOYE",
                    donnees={
                        "inscription_id": ins.pk,
                        "destinataire_id": recipient.pk,
                        "destinataire_email": recipient.email,
                        "destinataire_role": recipient_role,
                        "total_attendu_initial": str(form.cleaned_data.get("total_attendu") or 0),
                        "montant_options_ajoutees": str(montant_options_ajoutees),
                        "nouveau_total_attendu": str(nouveau_total_attendu),
                        "acompte1_encaisse": str(form.cleaned_data.get("acompte1_encaisse") or 0),
                        "acompte2_encaisse": str(form.cleaned_data.get("acompte2_encaisse") or 0),
                        "solde_encaisse": str(form.cleaned_data.get("solde_encaisse") or 0),
                        "paiement_recu_pour_options": str(paiement_options),
                        "date_paiement": str(form.cleaned_data.get("date_paiement")),
                        "total_paye": str(total_paye),
                        "montant_restant": str(reste_a_payer),
                        "trop_percu": str(trop_percu),
                        "reste_non_impute": str(remaining_unallocated),
                    },
                )

                payments_payload = []
                for pmt in paiements_after:
                    payments_payload.append({
                        "id": pmt.pk,
                        "montant_encaisse": str(pmt.montant_encaisse or 0),
                        "encaisse_le": pmt.encaisse_le.strftime("%d/%m/%Y") if pmt.encaisse_le else "—",
                        "statut": pmt.statut,
                        "statut_label": pmt.get_statut_display(),
                    })

                payment_summary_payload = {
                    "ins_id": ins.pk,
                    "options_added": str(montant_options_ajoutees or 0),
                    "total_attendu": str(nouveau_total_attendu),
                    "total_paye": str(total_paye),
                    "reste_a_payer": str(reste_a_payer),
                    "trop_percu": str(trop_percu),
                    "payment_state": (
                        "reste" if reste_a_payer > 0
                        else "trop_percu" if trop_percu > 0
                        else "solde"
                    ),
                    "payment_state_label": (
                        "Reste à payer" if reste_a_payer > 0
                        else "Trop-perçu" if trop_percu > 0
                        else "Soldé"
                    ),
                    "devise": ins.devise,
                }

                msg = f"Récapitulatif envoyé à {recipient.email}."
                return (
                    _json_ok(
                        msg,
                        payments=payments_payload,
                        payment_summary=payment_summary_payload,
                    )
                    if is_ajax else
                    (messages.success(request, msg) or redirect("bo_personne_detail", pk=p.pk))
                )


            # ---- Inscriptions : prévisualiser un rappel de paiement
            if action == "ins_payment_reminder_preview":
                ins_id = request.POST.get("ins_id")
                ins = (
                    ins_qs.filter(pk=ins_id)
                    .select_related("circuit", "pilote", "passager")
                    .prefetch_related("paiements", "selections_options__option")
                    .first()
                )
                if not ins:
                    msg = "Inscription introuvable pour cette personne."
                    return _json_err(msg) if is_ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                defaults = _build_payment_reminder_defaults(ins, p)

                if (defaults["montant_restant"] or Decimal("0")) <= 0:
                    msg = "Le paiement est déjà soldé. Aucun rappel nécessaire."
                    return _json_err(msg) if is_ajax else (messages.warning(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                form = PaymentReminderComposeForm(initial=defaults)

                print("REMINDER DEFAULTS =", defaults)
                print("REMINDER sujet =", defaults.get("sujet"))
                print("REMINDER total_attendu =", defaults.get("total_attendu"))
                print("REMINDER total_paye =", defaults.get("total_paye"))
                print("REMINDER montant_restant =", defaults.get("montant_restant"))

                form = PaymentReminderComposeForm(initial=defaults)

                print("FORM sujet =", form["sujet"].value())
                print("FORM total_attendu =", form["total_attendu"].value())
                print("FORM total_paye =", form["total_paye"].value())
                print("FORM montant_restant =", form["montant_restant"].value())
                print("FORM situation_label =", form["situation_label"].value())

                if is_ajax:
                    html = render_to_string(
                        "circuitMoto/admin/partials/_payment_reminder_modal_form.html",
                        {
                            "form": form,
                            "ins": ins,
                            "personne": p,
                            "recipient_email": defaults["recipient_email"],
                            "recipient_name": defaults["recipient_name"],
                            "paiements": defaults["paiements"],
                        },
                        request=request,
                    )
                    return _json_ok("Formulaire chargé.", modal_html=html)

                messages.info(request, "Prévisualisation disponible en AJAX uniquement.")
                return redirect("bo_personne_detail", pk=p.pk)


            # ---- Inscriptions : envoyer un rappel de paiement
            if action == "ins_payment_reminder_send":
                ins_id = request.POST.get("ins_id")
                ins = (
                    ins_qs.filter(pk=ins_id)
                    .select_related("circuit", "pilote", "passager")
                    .prefetch_related("paiements", "selections_options__option")
                    .first()
                )
                if not ins:
                    msg = "Inscription introuvable pour cette personne."
                    return _json_err(msg) if is_ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                form = PaymentReminderComposeForm(request.POST)
                if not form.is_valid():
                    if is_ajax:
                        defaults = _build_payment_reminder_defaults(ins, p)
                        html = render_to_string(
                            "circuitMoto/admin/partials/_payment_reminder_modal_form.html",
                            {
                                "form": form,
                                "ins": ins,
                                "personne": p,
                                "recipient_email": defaults["recipient_email"],
                                "recipient_name": defaults["recipient_name"],
                                "paiements": defaults["paiements"],
                            },
                            request=request,
                        )
                        return _json_err("Merci de corriger le formulaire.", modal_html=html, status=422)

                    messages.error(request, "Merci de corriger le formulaire.")
                    return redirect("bo_personne_detail", pk=p.pk)

                recipient_role = form.cleaned_data["recipient_role"]
                recipient = ins.pilote if recipient_role == "pilote" else ins.passager

                if not recipient:
                    msg = "Destinataire introuvable."
                    return _json_err(msg) if is_ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                if not recipient.email:
                    msg = "Le destinataire n’a pas d’adresse email."
                    return _json_err(msg) if is_ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                paiements_after = list(ins.paiements.all().order_by("echeance_le", "pk"))
                total_paye = Decimal(sum((p.montant_encaisse or 0) for p in paiements_after))
                total_attendu = Decimal(form.cleaned_data.get("total_attendu") or 0)
                reste_a_payer = max(total_attendu - total_paye, Decimal("0"))
                trop_percu = max(total_paye - total_attendu, Decimal("0"))

                if reste_a_payer <= 0:
                    msg = "Le paiement est déjà soldé. Aucun rappel nécessaire."
                    return _json_err(msg) if is_ajax else (messages.warning(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                payment_rows = _build_email_payment_rows(ins) if form.cleaned_data["inclure_detail_paiements"] else []

                preview_ctx = {
                    "ins": ins,
                    "recipient": recipient,
                    "dest_name": recipient.prenom or recipient.nom or "client",
                    "circuit": ins.circuit,
                    "payment_rows": payment_rows,
                    "intro_message": form.cleaned_data["intro_message"],
                    "note_client": form.cleaned_data["note_client"],

                    # compatibilité avec le template commun recap/rappel
                    "total_attendu_initial": total_attendu,
                    "montant_options_ajoutees": Decimal(getattr(ins, "montant_options_ajoutees", 0) or 0),
                    "nouveau_total_attendu": total_attendu,

                    "total_attendu": total_attendu,
                    "total_paye": total_paye,
                    "reste_a_payer": reste_a_payer,
                    "trop_percu": trop_percu,
                    "situation_label": form.cleaned_data["situation_label"] or "Reste à payer",

                    "edit_url": _abs_url(reverse("inscription_edit_start", args=[ins.id_public])),

                    "infos_paiement": (
                        form.cleaned_data["infos_paiement_custom"]
                        if form.cleaned_data["inclure_infos_paiement"] else ""
                    ),
                    "has_infos_paiement": bool(
                        form.cleaned_data["inclure_infos_paiement"] and form.cleaned_data["infos_paiement_custom"]
                    ),

                    "paiement_libelle": "Virement bancaire",
                    "subject": form.cleaned_data["sujet"],
                    "is_reminder": True,

                    # flags attendus par le template email
                    "sent_acompte1_amount": Decimal("0"),
                    "sent_acompte2_amount": Decimal("0"),
                    "sent_solde_amount": Decimal("0"),
                    "sent_options_amount": Decimal("0"),

                    "has_sent_acompte1": False,
                    "has_sent_acompte2": False,
                    "has_sent_solde": False,
                    "has_sent_options": False,
                }

                print("REMINDER SEND -> to =", recipient.email)
                print("REMINDER SEND -> role =", recipient_role)
                print("REMINDER SEND -> subject =", form.cleaned_data["sujet"])
                print("REMINDER SEND -> ctx keys =", sorted(preview_ctx.keys()))

                try:
                    sent_count = notify_paiement_resume_compose(
                        recipient_email=recipient.email,
                        subject=form.cleaned_data["sujet"],
                        ctx=preview_ctx,
                        role=recipient_role,
                    )
                    print("REMINDER SEND -> sent_count =", sent_count)
                except Exception as e:
                    print("REMINDER SEND ERROR =", repr(e))
                    if is_ajax:
                        return _json_err(f"Erreur envoi email rappel: {e}", status=500)
                    messages.error(request, f"Erreur envoi email rappel: {e}")
                    return redirect("bo_personne_detail", pk=p.pk)

                if not sent_count:
                    msg = f"L’email de rappel n’a pas été envoyé à {recipient.email}."
                    return (
                        _json_err(msg, status=409)
                        if is_ajax else
                        (messages.warning(request, msg) or redirect("bo_personne_detail", pk=p.pk))
                    )

                JournalAudit.objects.create(
                    inscription=ins,
                    acteur=request.user,
                    action="PAIEMENT_RAPPEL_ENVOYE",
                    donnees={
                        "inscription_id": ins.pk,
                        "destinataire_id": recipient.pk,
                        "destinataire_email": recipient.email,
                        "destinataire_role": recipient_role,
                        "total_attendu": str(total_attendu),
                        "total_paye": str(total_paye),
                        "montant_restant": str(reste_a_payer),
                        "trop_percu": str(trop_percu),
                        "situation_label": form.cleaned_data["situation_label"] or "Reste à payer",
                    },
                )

                payment_summary_payload = {
                    "ins_id": ins.pk,
                    "options_added": str(getattr(ins, "montant_options_ajoutees", 0) or 0),
                    "total_attendu": str(total_attendu),
                    "total_paye": str(total_paye),
                    "reste_a_payer": str(reste_a_payer),
                    "trop_percu": str(trop_percu),
                    "payment_state": "reste" if reste_a_payer > 0 else "solde",
                    "payment_state_label": "Reste à payer" if reste_a_payer > 0 else "Soldé",
                    "devise": ins.devise,
                }

                msg = f"Rappel de paiement envoyé à {recipient.email}."
                return (
                    _json_ok(
                        msg,
                        payment_summary=payment_summary_payload,
                    )
                    if is_ajax else
                    (messages.success(request, msg) or redirect("bo_personne_detail", pk=p.pk))
                )


            # ---- MAJ du numéro d'urgence d'assurance
            if action == "assurance_set_tel_urgence":
                ins_id = request.POST.get("ins_id")
                tel = (request.POST.get("telephone_urgence") or "").strip()

                ins = ins_qs.filter(pk=ins_id).select_related("assurance").first()
                if not ins:
                    msg = "Inscription introuvable pour cette personne."
                    return _json_err(msg) if is_ajax else (messages.error(request, msg) or redirect("bo_personne_detail", pk=p.pk))

                # créer l'assurance si absente
                from .models import Assurance
                if not ins.assurance:
                    ins.assurance = Assurance.objects.create(inscription=ins)

                ins.assurance.telephone_urgence = tel
                ins.assurance.save(update_fields=["telephone_urgence", "modifie_le"])

                JournalAudit.objects.create(
                    inscription=ins,
                    acteur=request.user,
                    action="ASSURANCE_TEL_URGENCE_MAJ",
                    donnees={"telephone_urgence": tel},
                )

                msg = "Numéro d’urgence mis à jour."
                payload = {"assurance": {"ins_id": ins.pk, "telephone_urgence": ins.assurance.telephone_urgence}}
                return _json_ok(msg, **payload) if is_ajax else (messages.success(request, msg) or redirect("bo_personne_detail", pk=p.pk))

        except Exception as e:
            print("PERSONNE_DETAIL ERROR =", repr(e))
            if is_ajax:
                return _json_err(f"Erreur lors du traitement: {e}", status=500)
            messages.error(request, f"Erreur lors du traitement: {e}")
            return redirect("bo_personne_detail", pk=p.pk)
    # ==========================================================================

    # Préfetch pour limiter les requêtes
    selections_prefetch = Prefetch(
        "selections_options",
        queryset=SelectionOption.objects.select_related("option", "inscription"),
    )
    documents_prefetch = Prefetch(
        "documents",
        queryset=Document.objects.select_related("verifie_par").order_by("type_document", "-cree_le"),
    )

    ins_pilote = (
        Inscription.objects
        .filter(pilote=p)
        .select_related("circuit", "assurance", "decharge", "pilote", "passager")
        .prefetch_related(selections_prefetch, documents_prefetch, "paiements", "journaux_rappel", "journaux_audit")
        .order_by("-cree_le")
    )

    ins_passager = (
        Inscription.objects
        .filter(passager=p)
        .select_related("circuit", "assurance", "decharge", "pilote", "passager")
        .prefetch_related(selections_prefetch, documents_prefetch, "paiements", "journaux_rappel", "journaux_audit")
        .order_by("-cree_le")
    )

    # Petits totaux pratiques (base/options/attendu)
    def enrich(ins):
        base = (ins.prix_pilote_unitaire or 0) + ((ins.prix_passager_unitaire or 0) if ins.passager_id else 0)
        opts = sum(sel.prix_total() for sel in ins.selections_options.all())
        total_paye = sum((pmt.montant_encaisse or 0) for pmt in ins.paiements.all())

        ins.total_base = base
        ins.total_options = opts
        ins.total_attendu = base + opts
        ins.total_paye = total_paye
        ins.reste_a_payer = max(ins.total_attendu - total_paye, 0)
        ins.trop_percu = max(total_paye - ins.total_attendu, 0)
        ins.options_added = getattr(ins, "montant_options_ajoutees", 0) or 0
        ins.paiements_display = _ordered_payments_for_display(ins)

        if ins.reste_a_payer > 0:
            ins.payment_state = "reste"
        elif ins.trop_percu > 0:
            ins.payment_state = "trop_percu"
        else:
            ins.payment_state = "solde"

        return ins

    ins_pilote = [enrich(i) for i in ins_pilote]
    ins_passager = [enrich(i) for i in ins_passager]

    # Circuits pour le menu "Nouvelle inscription" dans personne_detail
    circuits = Circuit.objects.order_by("-date_debut")

    ctx = {
        "personne": p,
        "infos_medicales": getattr(p, "infos_medicales", None),
        "motos": list(p.motos.all()),
        "contacts": list(p.contacts_urgence.all()),
        "ins_pilote": ins_pilote,
        "ins_passager": ins_passager,
        "circuits": circuits,
    }
    return render(request, "circuitMoto/admin/personnes_detail.html", ctx)

# ---------- Emailing ----------
@staff_member_required
def emailing(request):
    if request.method == "POST":
        form = EmailingForm(request.POST, request.FILES)   # 👈 IMPORTANT
        if form.is_valid():
            circuit     = form.cleaned_data["circuit"]
            statut      = form.cleaned_data["statut"]
            incl_ps     = form.cleaned_data["inclure_passagers"]
            sujet       = form.cleaned_data["sujet"]
            message     = form.cleaned_data["message"]
            test_only   = form.cleaned_data["test_only"]
            files       = request.FILES.getlist("pieces_jointes")  # 👈

            ins_qs = Inscription.objects.select_related("pilote","passager")
            if circuit:
                ins_qs = ins_qs.filter(circuit=circuit)
            if statut:
                ins_qs = ins_qs.filter(statut=statut)

            emails = set()
            for ins in ins_qs:
                if ins.pilote.email:
                    emails.add(ins.pilote.email)
                if incl_ps and ins.passager and ins.passager.email:
                    emails.add(ins.passager.email)

            if test_only:
                emails = {request.user.email} if request.user.email else set()

            if not emails:
                messages.warning(request, "Aucun destinataire.")
            else:
                # Préparer les PJ (lues en mémoire une fois)
                attachments = []
                total_bytes = 0
                for f in files:
                    data = f.read()
                    total_bytes += len(data)
                    attachments.append((f.name, data, f.content_type or "application/octet-stream"))

                # Petit garde-fou: éviter d’envoyer 25Mo * N destinataires sans le vouloir
                if total_bytes > 25 * 1024 * 1024 and len(emails) > 5:
                    messages.warning(
                        request,
                        "Beaucoup de pièces jointes volumineuses et de destinataires : "
                        "pense à tester d’abord ou à réduire la taille."
                    )

                from django.conf import settings
                sent = 0
                # Réutilise la même connexion SMTP pour toutes les expéditions
                with get_connection(fail_silently=False) as conn:
                    for to in emails:
                        msg = EmailMultiAlternatives(
                            subject=sujet,
                            body=strip_tags(message),  # corps texte brut (fallback)
                            from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
                            to=[to],
                            connection=conn,
                        )
                        # Si tu saisis du HTML dans “message”, on l’ajoute en alternatif
                        if "<" in message and ">" in message:
                            msg.attach_alternative(message, "text/html")
                        # Attacher les PJ
                        for name, data, ctype in attachments:
                            msg.attach(name, data, ctype)
                        msg.send()
                        sent += 1

                messages.success(request, f"Email envoyé à {sent} destinataire(s).")
                return redirect("bo_emailing")
    else:
        form = EmailingForm()

    return render(request, "circuitMoto/admin/emailing.html", {"form": form})

# ---------- Formulaire Personne ----------
class PersonneForm(forms.ModelForm):
    class Meta:
        model = Personne
        fields = [
            "prenom", "nom", "email", "telephone",
            "date_naissance", "numero_carte_identite",
            "adresse", "code_postal", "localite", "pays",
        ]
        widgets = {
            "date_naissance": forms.DateInput(attrs={"type": "date"}),
            "adresse": forms.Textarea(attrs={"rows": 2}),
        }

# ---------- Créer une personne (avec option d’inscription immédiate) ----------
# ---------- Formulaire Personne (identité/contact) ----------
class PersonneForm(forms.ModelForm):
    class Meta:
        model = Personne
        fields = [
            "prenom", "nom", "email", "telephone",
            "date_naissance", "numero_carte_identite",
            "adresse", "code_postal", "localite", "pays",
        ]
        widgets = {
            "date_naissance": forms.DateInput(attrs={"type": "date"}),
            "adresse": forms.Textarea(attrs={"rows": 2, "placeholder": "Rue, N°, compléments…"}),
        }

# ---------- Infos médicales (OneToOne) ----------
class InfosMedicalesForm(forms.ModelForm):
    class Meta:
        model = InfosMedicales
        exclude = ("personne",)
        widgets = {
            "notes": forms.Textarea(attrs={"rows": 3, "placeholder": "Allergies, traitements, remarques utiles…"}),
        }

# ---------- Motos (formset simple, pas inline) ----------
class SimpleMotoForm(forms.Form):
    marque = forms.CharField(label="Marque", max_length=80, required=False)
    modele = forms.CharField(label="Modèle", max_length=80, required=False)
    immatriculation = forms.CharField(label="Immatriculation", max_length=32, required=False)

    def has_any_data(self):
        return any(self.cleaned_data.get(k) for k in ["marque", "modele", "immatriculation"])

MotoFormSet = formset_factory(SimpleMotoForm, extra=1, can_delete=True)

# ---------- Contacts d’urgence (formset simple) ----------
class SimpleContactUrgenceForm(forms.Form):
    nom = forms.CharField(label="Nom", max_length=120, required=False)
    lien_parente = forms.CharField(label="Lien de parenté", max_length=80, required=False)
    telephone = forms.CharField(label="Téléphone", max_length=32, required=False)

    def has_any_data(self):
        return any(self.cleaned_data.get(k) for k in ["nom", "telephone", "lien_parente"])

ContactUrgenceFormSet = formset_factory(SimpleContactUrgenceForm, extra=2, can_delete=True)

# ---------- Créer une personne (dossier complet) ----------
@staff_member_required
@transaction.atomic
def personne_create(request):
    circuits = Circuit.objects.order_by("-date_debut")

    if request.method == "POST":
        form = PersonneForm(request.POST)
        med_form = InfosMedicalesForm(request.POST, prefix="med")
        motos = MotoFormSet(request.POST, prefix="motos")
        urg = ContactUrgenceFormSet(request.POST, prefix="urg")
        circuit_id = request.POST.get("circuit_id")  # optionnel

        all_valid = form.is_valid() and med_form.is_valid() and motos.is_valid() and urg.is_valid()
        if all_valid:
            # 1) Personne
            personne = form.save()

            # 2) Infos médicales (crée si au moins un champ saisi)
            med_cd = med_form.cleaned_data
            if any(med_cd.get(k) for k in med_cd.keys()):
                InfosMedicales.objects.update_or_create(
                    personne=personne,
                    defaults=med_cd
                )

            # 3) Motos
            for f in motos.forms:
                if not f.cleaned_data or f.cleaned_data.get("DELETE"):
                    continue
                if f.has_any_data():
                    Moto.objects.create(
                        proprietaire=personne,
                        marque=f.cleaned_data.get("marque") or "",
                        modele=f.cleaned_data.get("modele") or "",
                        immatriculation=f.cleaned_data.get("immatriculation") or "",
                    )

            # 4) Contacts d’urgence (garde seulement les lignes renseignées)
            for f in urg.forms:
                if not f.cleaned_data or f.cleaned_data.get("DELETE"):
                    continue
                if f.has_any_data():
                    ContactUrgence.objects.create(
                        personne=personne,
                        nom=f.cleaned_data.get("nom") or "",
                        lien_parente=f.cleaned_data.get("lien_parente") or "",
                        telephone=f.cleaned_data.get("telephone") or "",
                    )

            # 5) Enrôlement direct sur un circuit (optionnel)
            if circuit_id:
                circuit = get_object_or_404(Circuit, pk=circuit_id)
                ins = Inscription.objects.create(
                    circuit=circuit,
                    pilote=personne,
                    devise=circuit.devise,
                    prix_pilote_unitaire=circuit.prix_pilote_unitaire,
                    prix_passager_unitaire=0,
                )
                messages.success(request, "Dossier créé et inscrit sur le circuit sélectionné. Complétez l’inscription.")
                return redirect("inscription_edit_start", id_public=ins.id_public)

            messages.success(request, "Dossier personne créé.")
            return redirect("bo_personne_list")

        # (invalid) → on retombe en bas pour re-afficher les erreurs
    else:
        form = PersonneForm()
        med_form = InfosMedicalesForm(prefix="med")
        motos = MotoFormSet(prefix="motos")
        urg = ContactUrgenceFormSet(prefix="urg")

    return render(
        request,
        "circuitMoto/admin/personnes_form.html",
        {"form": form, "med_form": med_form, "motos": motos, "urg": urg, "circuits": circuits}
    )

# ---------- Inscrire une personne existante dans un circuit ----------
@staff_member_required
@transaction.atomic
def inscrire_personne(request, pk):
    # Appel via POST depuis la liste des personnes
    if request.method != "POST":
        return redirect("bo_personne_list")

    personne = get_object_or_404(Personne, pk=pk)
    circuit = get_object_or_404(Circuit, pk=request.POST.get("circuit_id"))

    # Ne tente pas de dupliquer : récupère si déjà présent
    try:
        ins, created = Inscription.objects.get_or_create(
            circuit=circuit,
            pilote=personne,
            defaults={
                "devise": circuit.devise,
                "prix_pilote_unitaire": circuit.prix_pilote_unitaire,
                "prix_passager_unitaire": 0,
            },
        )
    except IntegrityError:
        # Garde-fou en cas de course, on retombe sur l’existante
        ins = Inscription.objects.get(circuit=circuit, pilote=personne)
        created = False

    if created:
        messages.success(request, f"{personne.prenom} {personne.nom} inscrit sur {circuit.nom}.")
        return redirect("inscription_edit_start", id_public=ins.id_public)

    # Déjà inscrit : on informe et on renvoie sur la fiche
    messages.info(
        request,
        f"{personne.prenom} {personne.nom} est déjà inscrit sur « {circuit.nom} ». "
        "Ouverte la fiche existante."
    )
    return redirect("inscription_edit_start", id_public=ins.id_public)

@staff_member_required
@transaction.atomic
def inscription_delete(request, pk: int):
    if request.method != "POST":
        return redirect("bo_inscription_list")
    ins = get_object_or_404(Inscription, pk=pk)
    ref = ins.id_public
    ins.delete()
    messages.success(request, f"Inscription {ref} supprimée.")
    return redirect("bo_inscription_list")

def _first_day_of_month(d: date) -> date:
    return date(d.year, d.month, 1)

def _add_month(d: date) -> date:
    return date(d.year + (1 if d.month == 12 else 0), 1 if d.month == 12 else d.month + 1, 1)

def _sub_months(d: date, n: int) -> date:
    y = d.year
    m = d.month - n
    while m <= 0:
        m += 12
        y -= 1
    return date(y, m, 1)

@staff_member_required
def stats(request):
    # --- Fenêtre temporelle (par défaut : 12 derniers mois) ---
    today = timezone.now().date()
    default_end = _first_day_of_month(today)
    default_start = _sub_months(default_end, 11)

    try:
        start_q = request.GET.get("start")  # "YYYY-MM" ou "YYYY-MM-DD"
        end_q = request.GET.get("end")
        start = date.fromisoformat(start_q + "-01") if start_q and len(start_q) == 7 else date.fromisoformat(start_q) if start_q else default_start
        end   = date.fromisoformat(end_q + "-01")   if end_q and len(end_q) == 7   else date.fromisoformat(end_q)   if end_q   else default_end
    except Exception:
        start, end = default_start, default_end

    start = _first_day_of_month(start)
    end   = _first_day_of_month(end)

    # Séquence des mois (labels complets + clés YYYY-MM)
    months = []
    keys = []
    cur = start
    while cur <= end:
        months.append(cur.strftime("%b %Y").capitalize())  # ex: "janv. 2025"
        keys.append(cur.strftime("%Y-%m"))
        cur = _add_month(cur)

    # --- KPIs globaux ---
    total_circuits = Circuit.objects.count()
    total_personnes = Personne.objects.count()
    total_inscriptions = Inscription.objects.count()
    total_validees = Inscription.objects.filter(statut=StatutInscription.VALIDE).count()

    # Finances globales (tous temps)
    from .models import Paiement  # évite import circulaire dans certains IDE
    total_attendu = Paiement.objects.aggregate(s=Coalesce(Sum("montant_du"), 0))["s"]
    total_encaisse = Paiement.objects.aggregate(s=Coalesce(Sum("montant_encaisse"), 0))["s"]
    taux_encaissement = int(round((total_encaisse / total_attendu) * 100)) if total_attendu else 0

    # --- Séries temporelles sur la fenêtre ---
    # 1) Inscriptions créées par mois
    ins_series = (
        Inscription.objects
        .filter(cree_le__date__gte=start, cree_le__date__lt=_add_month(end))
        .annotate(m=TruncMonth("cree_le"))
        .values("m")
        .annotate(c=Count("id"))
        .order_by("m")
    )
    ins_map = {row["m"].strftime("%Y-%m"): row["c"] for row in ins_series}

    # 2) Montants attendus (par échéance) vs encaissés (par date de paiement)
    due_series = (
        Paiement.objects
        .filter(echeance_le__gte=start, echeance_le__lt=_add_month(end))
        .annotate(m=TruncMonth("echeance_le"))
        .values("m")
        .annotate(s=Coalesce(Sum("montant_du"), 0))
        .order_by("m")
    )
    due_map = {row["m"].strftime("%Y-%m"): row["s"] for row in due_series}

    paid_series = (
        Paiement.objects
        .filter(encaisse_le__isnull=False, encaisse_le__gte=start, encaisse_le__lt=_add_month(end))
        .annotate(m=TruncMonth("encaisse_le"))
        .values("m")
        .annotate(s=Coalesce(Sum("montant_encaisse"), 0))
        .order_by("m")
    )
    paid_map = {row["m"].strftime("%Y-%m"): row["s"] for row in paid_series}

    # Vecteurs alignés à la liste des mois
    serie_ins = [ins_map.get(k, 0) for k in keys]
    serie_due = [due_map.get(k, 0) for k in keys]
    serie_paid = [paid_map.get(k, 0) for k in keys]

    # --- Répartitions ---
    # Inscription par statut
    rep_ins = list(
        Inscription.objects.values("statut").annotate(c=Count("id")).order_by("statut")
    )
    # Documents par statut
    rep_docs = list(
        Document.objects.values("statut").annotate(c=Count("id")).order_by("statut")
    )
    # Circuits par statut
    rep_circuits = list(
        Circuit.objects.values("statut").annotate(c=Count("id")).order_by("statut")
    )

    # --- Occupation prochains circuits ---
    upcoming = (
        Circuit.objects
        .filter(date_debut__gte=today)
        .annotate(nb=Count("inscriptions"))
        .order_by("date_debut")[:10]
    )
    upcoming_ctx = []
    for c in upcoming:
        cap = c.capacite or 0
        pct = int(round((c.nb / cap) * 100)) if cap else 0
        upcoming_ctx.append({
            "id": c.id,
            "nom": c.nom,
            "code": c.code,
            "date_debut": c.date_debut.isoformat(),
            "capacite": cap,
            "nb": c.nb,
            "pct": pct,
        })

    # --- Série personnes créées (bonus) ---
    people_series = (
        Personne.objects
        .filter(cree_le__date__gte=start, cree_le__date__lt=_add_month(end))
        .annotate(m=TruncMonth("cree_le"))
        .values("m")
        .annotate(c=Count("id"))
        .order_by("m")
    )
    people_map = {row["m"].strftime("%Y-%m"): row["c"] for row in people_series}
    serie_people = [people_map.get(k, 0) for k in keys]

    # Contexte JSON pour le front (Chart.js)
    data = {
        "range": {
            "start": start.isoformat(),
            "end": end.isoformat(),
            "labels": months,
            "keys": keys,
        },
        "kpis": {
            "circuits": total_circuits,
            "personnes": total_personnes,
            "inscriptions": total_inscriptions,
            "validees": total_validees,
            "attendu": total_attendu,
            "encaisse": total_encaisse,
            "taux": taux_encaissement,
        },
        "series": {
            "inscriptions": serie_ins,
            "attendu": serie_due,
            "encaisse": serie_paid,
            "personnes": serie_people,
        },
        "repartitions": {
            "inscriptions": rep_ins,
            "documents": rep_docs,
            "circuits": rep_circuits,
        },
        "upcoming": upcoming_ctx,
    }

    return render(request, "circuitMoto/admin/stats.html", {
        "data": data,
        "start": start,
        "end": end,
    })

# ---------- Création d'utilisateurs staff ----------
class StaffUserCreateForm(forms.ModelForm):
    """Formulaire minimal et clair pour créer un staff complet."""
    is_superuser = forms.BooleanField(required=False, initial=False, label="Administrateur (superuser)")

    class Meta:
        model = get_user_model()
        fields = ["username", "email", "first_name", "last_name"]

    def clean_email(self):
        e = (self.cleaned_data.get("email") or "").strip()
        if not e:
            raise forms.ValidationError("Email requis.")
        if get_user_model().objects.filter(email__iexact=e).exists():
            raise forms.ValidationError("Un utilisateur avec cet email existe déjà.")
        return e

def generate_strong_password(length: int = 12) -> str:
    """Génère un mot de passe fort: maj, min, chiffres, symboles."""
    U = string.ascii_uppercase
    L = string.ascii_lowercase
    D = string.digits
    S = "!@#$%^&*()-_=+[]{};:,.?/"

    # au moins 1 de chaque
    base = [secrets.choice(U), secrets.choice(L), secrets.choice(D), secrets.choice(S)]
    pool = U + L + D + S
    base += [secrets.choice(pool) for _ in range(max(8, length) - 4)]
    secrets.SystemRandom().shuffle(base)
    return "".join(base)

@staff_member_required
def user_create(request):
    from .models import AccountFlags

    if not request.user.has_perm("auth.add_user"):
        messages.error(request, "Vous n'avez pas la permission de créer des utilisateurs.")
        return redirect("bo_dashboard")

    if request.method == "POST":
        form = StaffUserCreateForm(request.POST)
        if form.is_valid():
            User = get_user_model()
            tmp_pwd = generate_strong_password(12)

            user = User(
                username=form.cleaned_data["username"],
                email=form.cleaned_data["email"],
                first_name=form.cleaned_data["first_name"],
                last_name=form.cleaned_data["last_name"],
                is_active=True,
                is_staff=True,
                is_superuser=form.cleaned_data.get("is_superuser") or False,
            )
            user.set_password(tmp_pwd)
            user.save()

            # Flag de première connexion + mémorisation du hash provisoire
            AccountFlags.objects.update_or_create(
                user=user,
                defaults={
                    "must_change_password": True,
                    "initial_password_hash": user.password,
                }
            )

            # On affiche le mot de passe provisoire une seule fois
            messages.success(request, f"Utilisateur « {user.get_full_name() or user.username} » créé.")
            return render(request, "circuitMoto/admin/users_created_show_pwd.html", {"user_obj": user, "tmp_pwd": tmp_pwd})
    else:
        form = StaffUserCreateForm()

    return render(request, "circuitMoto/admin/users_create.html", {"form": form})

class FirstPasswordChangeForm(forms.Form):
    password1 = forms.CharField(widget=forms.PasswordInput(attrs={"autocomplete":"new-password", "class":"pw-input", "placeholder":"Nouveau mot de passe"}), label="Nouveau mot de passe")
    password2 = forms.CharField(widget=forms.PasswordInput(attrs={"autocomplete":"new-password", "class":"pw-input", "placeholder":"Confirmer le mot de passe"}), label="Confirmer")

    def clean(self):
        c = super().clean()
        p1 = c.get("password1") or ""
        p2 = c.get("password2") or ""

        if p1 != p2:
            self.add_error("password2", "Les deux mots de passe ne correspondent pas.")
        # règles de base
        if len(p1) < 8:
            self.add_error("password1", "Au moins 8 caractères.")
        if not any(ch.islower() for ch in p1): self.add_error("password1", "Doit contenir une minuscule.")
        if not any(ch.isupper() for ch in p1): self.add_error("password1", "Doit contenir une majuscule.")
        if not any(ch.isdigit() for ch in p1): self.add_error("password1", "Doit contenir un chiffre.")
        if not any(ch in "!@#$%^&*()-_=+[]{};:,.?/" for ch in p1): self.add_error("password1", "Doit contenir un symbole.")
        # validators Django (optionnel)
        try:
            validate_password(p1)
        except PwValidationError as e:
            self.add_error("password1", " ".join(e.messages))
        return c

@login_required
def password_change_first(request):
    """Forcé au premier login ; interdit de reprendre le mot de passe provisoire."""
    from .models import AccountFlags

    flags = getattr(request.user, "flags", None)
    if not flags or not flags.must_change_password:
        return redirect("bo_dashboard")

    if request.method == "POST":
        form = FirstPasswordChangeForm(request.POST)
        if form.is_valid():
            new_pwd = form.cleaned_data["password1"]

            # Interdit de reprendre le provisoire
            if flags.initial_password_hash and check_password(new_pwd, flags.initial_password_hash):
                form.add_error("password1", "Le nouveau mot de passe ne peut pas être identique au mot de passe provisoire.")
            else:
                request.user.set_password(new_pwd)
                request.user.save(update_fields=["password"])
                flags.must_change_password = False
                flags.initial_password_hash = ""
                flags.save(update_fields=["must_change_password", "initial_password_hash"])

                # Reconnecte l'utilisateur automatiquement
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, request.user)

                messages.success(request, "Mot de passe mis à jour. Bienvenue !")
                return redirect("bo_dashboard")
    else:
        form = FirstPasswordChangeForm()

    return render(request, "circuitMoto/auth/first_change_password.html", {"form": form})

class SelfPasswordChangeForm(forms.Form):
    current_password = forms.CharField(
        label="Mot de passe actuel",
        widget=forms.PasswordInput(attrs={"autocomplete": "current-password", "class": "pw-input", "placeholder": "Mot de passe actuel"})
    )
    new_password1 = forms.CharField(
        label="Nouveau mot de passe",
        widget=forms.PasswordInput(attrs={"autocomplete": "new-password", "class": "pw-input", "placeholder": "Nouveau mot de passe"})
    )
    new_password2 = forms.CharField(
        label="Confirmer le mot de passe",
        widget=forms.PasswordInput(attrs={"autocomplete": "new-password", "class": "pw-input", "placeholder": "Confirmer"})
    )

    def __init__(self, user, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.user = user

    def clean_current_password(self):
        cur = self.cleaned_data.get("current_password") or ""
        if not self.user.check_password(cur):
            raise forms.ValidationError("Mot de passe actuel incorrect.")
        return cur

    def clean(self):
        c = super().clean()
        cur = c.get("current_password") or ""
        p1  = c.get("new_password1") or ""
        p2  = c.get("new_password2") or ""

        if p1 != p2:
            self.add_error("new_password2", "Les deux mots de passe ne correspondent pas.")
        if len(p1) < 8:
            self.add_error("new_password1", "Au moins 8 caractères.")
        if not any(ch.islower() for ch in p1):
            self.add_error("new_password1", "Doit contenir une minuscule.")
        if not any(ch.isupper() for ch in p1):
            self.add_error("new_password1", "Doit contenir une majuscule.")
        if not any(ch.isdigit() for ch in p1):
            self.add_error("new_password1", "Doit contenir un chiffre.")
        if not any(ch in "!@#$%^&*()-_=+[]{};:,.?/" for ch in p1):
            self.add_error("new_password1", "Doit contenir un symbole.")
        if p1 and cur and p1 == cur:
            self.add_error("new_password1", "Le nouveau mot de passe doit être différent de l’actuel.")
        try:
            validate_password(p1, user=self.user)
        except PwValidationError as e:
            self.add_error("new_password1", " ".join(e.messages))
        return c

@login_required
def password_change_self(request):
    """Changement volontaire de mot de passe par un utilisateur connecté."""
    if request.method == "POST":
        form = SelfPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            new_pwd = form.cleaned_data["new_password1"]
            request.user.set_password(new_pwd)
            request.user.save(update_fields=["password"])
            update_session_auth_hash(request, request.user)  # reste connecté
            messages.success(request, "Votre mot de passe a bien été mis à jour.")
            # staff -> dashboard, sinon -> home
            return redirect("bo_dashboard" if request.user.is_staff else "home")
    else:
        form = SelfPasswordChangeForm(request.user)

    return render(request, "circuitMoto/auth/change_password.html", {"form": form})

from django.contrib.auth.decorators import user_passes_test
@user_passes_test(lambda u: u.is_active and u.is_superuser)
@require_POST
def email_flags_set(request):
    """
    Active/désactive une pause d’email.
    key: all | pilote | passager
    value: "1" (pause ON) ou "0" (pause OFF)
    """
    key = (request.POST.get("key") or "").strip()
    raw = (request.POST.get("value") or "").strip().lower()
    value = raw in {"1", "true", "on", "yes"}

    mapping = {
        "all": "emails_pause_all",
        "pilote": "emails_pause_pilote",
        "passager": "emails_pause_passager",
    }
    if key not in mapping:
        return JsonResponse({"ok": False, "message": "Clé invalide."}, status=400)

    prefs = SitePrefs.get()
    setattr(prefs, mapping[key], value)
    prefs.save(update_fields=[mapping[key], "modifie_le"])

    # invalide le cache côté emails.py (voir plus bas)
    cache.delete("site_email_flags_v1")

    return JsonResponse({
        "ok": True,
        "key": key,
        "paused": value,
        "flags": SitePrefs.as_flags(),
        "message": ("Pause activée" if value else "Pause désactivée"),
    })

def _safe_str(v):
    if v is None: return ""
    if isinstance(v, bool): return "Oui" if v else "Non"
    return str(v)

def _reservations_export_queryset(request):
    year = (request.GET.get("year") or "").strip()
    circuit_id = (request.GET.get("circuit") or "").strip()
    statut = (request.GET.get("statut") or "").strip()

    qs = (
        Inscription.objects
        .select_related("circuit", "pilote", "passager", "assurance")
        .prefetch_related(
            Prefetch("selections_options", queryset=SelectionOption.objects.select_related("option")),
            "documents",
            "paiements",
        )
        .order_by("circuit__date_debut", "circuit__code", "pilote__nom", "pilote__prenom", "id")
    )

    if year:
        qs = qs.filter(circuit__date_debut__year=year)
    if circuit_id:
        qs = qs.filter(circuit_id=circuit_id)
    if statut:
        qs = qs.filter(statut=statut)

    return qs, year, circuit_id, statut


def _sheet_title(raw_title, used_titles):
    title = re.sub(r"[:\\/*?\[\]]+", "-", (raw_title or "Circuit")).strip() or "Circuit"
    title = title[:31]
    base = title
    idx = 2
    while title in used_titles:
        suffix = f" {idx}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        idx += 1
    used_titles.add(title)
    return title


def _date_for_excel(value):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return value


def _money_for_excel(value):
    value = Decimal(value or 0)
    if value <= 0:
        return ""
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def _join_unique(values):
    out = []
    seen = set()
    for value in values:
        value = (value or "").strip()
        if not value:
            continue
        key = value.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(value)
    return " / ".join(out)


def _short_option_value(option, quantity=1):
    hay = slugify(f"{getattr(option, 'code', '')} {getattr(option, 'intitule', '')} {getattr(option, 'description', '')}")

    if "single" in hay or "singel" in hay or "singe" in hay:
        value = "single"
    elif "double" in hay:
        value = "double"
    elif "twin" in hay:
        value = "twin"
    elif "triple" in hay:
        value = "triple"
    elif "quadruple" in hay or "quandruple" in hay:
        value = "quadruple"
    else:
        value = "oui"

    quantity = quantity or 1
    if quantity > 1:
        return f"{value} x{quantity}"
    return value


def _selection_applies_to_role(selection, role):
    option = getattr(selection, "option", None)
    if not option:
        return False
    if getattr(option, "facture_par_personne", False):
        return bool(selection.pour_passager) == (role == "passager")
    return True


def _reservation_option_columns(inscription, role):
    cols = {
        "chambre": [],
        "cabine": [],
        "contrex": [],
        "grenoble": [],
        "montbrison": [],
        "troys": [],
        "divers": [],
    }

    selections = list(getattr(inscription, "selections_options", []).all())
    for selection in selections:
        if not _selection_applies_to_role(selection, role):
            continue

        option = selection.option
        hay = slugify(f"{option.code} {option.intitule} {option.description}")
        value = _short_option_value(option, selection.quantite or 1)
        categorie = getattr(option, "categorie", "")

        if categorie == "HEBERGEMENT" or "chambre" in hay:
            cols["chambre"].append(value if value != "oui" else option.intitule)
        elif categorie == "CABINE" or "cabine" in hay:
            cols["cabine"].append(value if value != "oui" else option.intitule)
        elif categorie == "NUIT_ETAPE" or "nuit" in hay:
            if "contrex" in hay:
                cols["contrex"].append(value)
            elif "grenoble" in hay:
                cols["grenoble"].append(value)
            elif "montbrison" in hay:
                cols["montbrison"].append(value)
            elif "troyes" in hay or "troys" in hay:
                cols["troys"].append(value)
            else:
                cols["divers"].append(option.intitule)
        else:
            cols["divers"].append(option.intitule)

    return {key: _join_unique(values) for key, values in cols.items()}


def _reservation_document_label(inscription):
    docs = list(getattr(inscription, "documents", []).all())
    if not docs:
        return ""

    total = len(docs)
    valides = sum(1 for doc in docs if doc.statut == StatutDocument.VALIDE)
    refuses = sum(1 for doc in docs if doc.statut == StatutDocument.REFUSE)
    attente = total - valides - refuses

    if total and valides == total:
        return "ok"

    parts = []
    if valides:
        parts.append(f"{valides}/{total} valides")
    if attente:
        parts.append(f"{attente} attente")
    if refuses:
        parts.append(f"{refuses} refuse")
    return ", ".join(parts)


def _payment_map(inscription):
    return {paiement.libelle: paiement for paiement in getattr(inscription, "paiements", []).all()}


def _payment_amount(payments, libelle):
    paiement = payments.get(libelle)
    if not paiement:
        return ""
    return _money_for_excel(paiement.montant_encaisse)


def _payment_date(payments, libelle):
    paiement = payments.get(libelle)
    if not paiement or not paiement.montant_encaisse:
        return ""
    return _date_for_excel(paiement.encaisse_le)


def _payment_notes(payments):
    notes = []
    for libelle in (LibellePaiement.ACOMPTE1, LibellePaiement.ACOMPTE2, LibellePaiement.SOLDE):
        paiement = payments.get(libelle)
        if not paiement:
            continue
        fragments = []
        if getattr(paiement, "methode", ""):
            try:
                fragments.append(paiement.get_methode_display())
            except Exception:
                fragments.append(paiement.methode)
        if getattr(paiement, "reference", ""):
            fragments.append(paiement.reference)
        if fragments:
            notes.append(f"{paiement.get_libelle_display()}: {' - '.join(fragments)}")
    return notes


def _reservation_total(inscription):
    base = (inscription.prix_pilote_unitaire or 0) + (
        (inscription.prix_passager_unitaire or 0) if inscription.passager_id else 0
    )
    options = sum(selection.prix_total() for selection in inscription.selections_options.all())
    options_ajoutees = getattr(inscription, "montant_options_ajoutees", 0) or 0
    return Decimal(base + options + options_ajoutees)


def _append_reservation_sheet(wb, circuit, inscriptions, used_titles):
    title = _sheet_title(getattr(circuit, "code", "") or getattr(circuit, "nom", "") or "Circuit", used_titles)
    ws = wb.create_sheet(title=title)

    date_label = ""
    if getattr(circuit, "date_debut", None) and getattr(circuit, "date_fin", None):
        date_label = f"{circuit.date_debut:%d/%m/%Y} au {circuit.date_fin:%d/%m/%Y}"

    ws.append([
        "",
        getattr(circuit, "nom", "") or title,
        date_label,
        "doc",
        "mode",
        "chambre",
        "cabine",
        "contrex",
        "grenoble",
        "montbrison",
        "troys",
        "1 er acompte",
        "date",
        "2 em acompte",
        "date",
        "solde",
        "date",
        "total",
        "divers",
    ])

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True, color="111827")
    thin = Side(style="thin", color="E5E7EB")
    border = Border(bottom=thin)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border

    ws.freeze_panes = "A2"

    participant_no = 1
    total_acompte1 = Decimal("0")
    total_acompte2 = Decimal("0")
    total_solde = Decimal("0")
    total_attendu = Decimal("0")
    total_paye = Decimal("0")

    for inscription in inscriptions:
        payments = _payment_map(inscription)
        inscription_total = _reservation_total(inscription)
        inscription_paid = Decimal(sum((p.montant_encaisse or 0) for p in payments.values()))
        total_attendu += inscription_total
        total_paye += inscription_paid
        total_acompte1 += Decimal(getattr(payments.get(LibellePaiement.ACOMPTE1), "montant_encaisse", 0) or 0)
        total_acompte2 += Decimal(getattr(payments.get(LibellePaiement.ACOMPTE2), "montant_encaisse", 0) or 0)
        total_solde += Decimal(getattr(payments.get(LibellePaiement.SOLDE), "montant_encaisse", 0) or 0)

        participants = [("pilote", inscription.pilote)]
        if inscription.passager_id:
            participants.append(("passager", inscription.passager))

        for role, person in participants:
            if not person:
                continue

            is_primary = role == "pilote"
            option_cols = _reservation_option_columns(inscription, role)

            divers_parts = []
            if getattr(inscription, "notes", ""):
                divers_parts.append(inscription.notes)
            if option_cols.get("divers"):
                divers_parts.append(option_cols["divers"])
            if is_primary:
                divers_parts.extend(_payment_notes(payments))
                reste = max(inscription_total - inscription_paid, Decimal("0"))
                trop_percu = max(inscription_paid - inscription_total, Decimal("0"))
                if reste > 0:
                    divers_parts.append(f"reste a payer: {_money_for_excel(reste)}")
                elif trop_percu > 0:
                    divers_parts.append(f"trop-percu: {_money_for_excel(trop_percu)}")
                if getattr(inscription, "paiement_options_recu", 0):
                    divers_parts.append(f"paiement options recu: {inscription.paiement_options_recu}")

            ws.append([
                participant_no,
                f"{(person.nom or '').upper()} {person.prenom or ''}".strip(),
                person.email or "",
                _reservation_document_label(inscription) if is_primary else "",
                "",
                option_cols["chambre"],
                option_cols["cabine"],
                option_cols["contrex"],
                option_cols["grenoble"],
                option_cols["montbrison"],
                option_cols["troys"],
                _payment_amount(payments, LibellePaiement.ACOMPTE1) if is_primary else "",
                _payment_date(payments, LibellePaiement.ACOMPTE1) if is_primary else "",
                _payment_amount(payments, LibellePaiement.ACOMPTE2) if is_primary else "",
                _payment_date(payments, LibellePaiement.ACOMPTE2) if is_primary else "",
                _payment_amount(payments, LibellePaiement.SOLDE) if is_primary else "",
                _payment_date(payments, LibellePaiement.SOLDE) if is_primary else "",
                _money_for_excel(inscription_total) if is_primary else "",
                " | ".join(_safe_str(part) for part in divers_parts if part),
            ])

            current_row = ws.max_row
            for cell in ws[current_row]:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if role == "passager":
                    cell.fill = PatternFill("solid", fgColor="F9FAFB")
            participant_no += 1

    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=2, value="TOTAL")
    ws.cell(row=total_row, column=3, value=f"{participant_no - 1} pax")
    ws.cell(row=total_row, column=12, value=_money_for_excel(total_acompte1))
    ws.cell(row=total_row, column=14, value=_money_for_excel(total_acompte2))
    ws.cell(row=total_row, column=16, value=_money_for_excel(total_solde))
    ws.cell(row=total_row, column=18, value=_money_for_excel(total_attendu))
    ws.cell(row=total_row, column=19, value=f"encaisse: {_money_for_excel(total_paye)}")

    for cell in ws[total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FEF3C7")
        cell.border = border

    for col_idx in (13, 15, 17):
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = "DD/MM/YYYY"

    for col_idx in (12, 14, 16, 18):
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = "#,##0"

    widths = {
        "A": 5, "B": 28, "C": 30, "D": 18, "E": 22, "F": 16, "G": 16,
        "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 13,
        "N": 14, "O": 13, "P": 14, "Q": 13, "R": 12, "S": 46,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.auto_filter.ref = f"A1:S{max(ws.max_row, 1)}"
    return ws



PAYMENT_EXCEL_HEADERS = [
    "ID_INSCRIPTION",
    "REFERENCE",
    "CIRCUIT",
    "DATE_INSCRIPTION",
    "NOM",
    "PRENOM",
    "EMAIL",
    "TELEPHONE",
    "STATUT_DOSSIER",
    "MONTANT_TOTAL",
    "TOTAL_DEJA_ENREGISTRE",
    "RESTE_A_PAYER",
    "ACOMPTE_1_EXISTANT",
    "DATE_ACOMPTE_1",
    "ACOMPTE_2_EXISTANT",
    "DATE_ACOMPTE_2",
    "SOLDE_EXISTANT",
    "DATE_SOLDE",
    "ETAPE_ACTUELLE_A_ENCODER",
    "MONTANT_A_ENCODER",
    "DATE_PAIEMENT_A_ENCODER",
    "MODE_PAIEMENT",
    "ENVOYER_EMAIL",
    "NOTE_INTERNE",
]

PAYMENT_STAGE_ORDER = [
    LibellePaiement.ACOMPTE1,
    LibellePaiement.ACOMPTE2,
    LibellePaiement.SOLDE,
]

PAYMENT_STAGE_LABELS = {
    LibellePaiement.ACOMPTE1: "ACOMPTE_1",
    LibellePaiement.ACOMPTE2: "ACOMPTE_2",
    LibellePaiement.SOLDE: "SOLDE",
}

PAYMENT_STAGE_REVERSE = {v: k for k, v in PAYMENT_STAGE_LABELS.items()}

PAYMENT_METHOD_NORMALIZATION = {
    "ESPECES": MethodePaiement.ESPECES,
    "ESPÈCES": MethodePaiement.ESPECES,
    "VIREMENT": MethodePaiement.VIREMENT,
    "VIREMENT BANCAIRE": MethodePaiement.VIREMENT,
    "CHEQUE": MethodePaiement.CHEQUE,
    "CHÈQUE": MethodePaiement.CHEQUE,
    "MOBILE_MONEY": MethodePaiement.MOBILE_MONEY,
    "MOBILE MONEY": MethodePaiement.MOBILE_MONEY,
    "PAYPAL": MethodePaiement.VIREMENT,
}


def _payment_stage_payload(ins):
    payment_map = {p.libelle: p for p in ins.paiements.all()}
    rows = []
    current_label = "TERMINE"
    current_payment = None

    for libelle in PAYMENT_STAGE_ORDER:
        p = payment_map.get(libelle)
        expected = Decimal(getattr(p, "montant_du", 0) or 0)
        collected = Decimal(getattr(p, "montant_encaisse", 0) or 0)
        remaining = max(expected - collected, Decimal("0"))
        if not current_payment and remaining > 0:
            current_payment = p
            current_label = PAYMENT_STAGE_LABELS[libelle]
        rows.append({
            "libelle": libelle,
            "label": PAYMENT_STAGE_LABELS[libelle],
            "payment": p,
            "expected": expected,
            "collected": collected,
            "remaining": remaining,
            "date": getattr(p, "encaisse_le", None),
        })

    total_paid = sum((row["collected"] for row in rows), Decimal("0"))
    total_due = Decimal(ins.total_attendu() or 0)
    remaining_total = max(total_due - total_paid, Decimal("0"))

    return {
        "rows": rows,
        "current_label": current_label,
        "current_payment": current_payment,
        "total_paid": total_paid,
        "total_due": total_due,
        "remaining_total": remaining_total,
    }


def _payment_row_for_export(ins):
    payload = _payment_stage_payload(ins)
    payment_lookup = {row["libelle"]: row for row in payload["rows"]}

    def _date_for(libelle):
        row = payment_lookup.get(libelle) or {}
        return row.get("date")

    def _amount_for(libelle):
        row = payment_lookup.get(libelle) or {}
        return row.get("collected", Decimal("0"))

    pilot = ins.pilote
    return {
        "ID_INSCRIPTION": ins.pk,
        "REFERENCE": str(ins.id_public),
        "CIRCUIT": ins.circuit.nom,
        "DATE_INSCRIPTION": timezone.localtime(ins.cree_le).date() if ins.cree_le else None,
        "NOM": getattr(pilot, "nom", "") or "",
        "PRENOM": getattr(pilot, "prenom", "") or "",
        "EMAIL": getattr(pilot, "email", "") or "",
        "TELEPHONE": getattr(pilot, "telephone", "") or "",
        "STATUT_DOSSIER": ins.get_statut_display() if hasattr(ins, "get_statut_display") else ins.statut,
        "MONTANT_TOTAL": payload["total_due"],
        "TOTAL_DEJA_ENREGISTRE": payload["total_paid"],
        "RESTE_A_PAYER": payload["remaining_total"],
        "ACOMPTE_1_EXISTANT": _amount_for(LibellePaiement.ACOMPTE1),
        "DATE_ACOMPTE_1": _date_for(LibellePaiement.ACOMPTE1),
        "ACOMPTE_2_EXISTANT": _amount_for(LibellePaiement.ACOMPTE2),
        "DATE_ACOMPTE_2": _date_for(LibellePaiement.ACOMPTE2),
        "SOLDE_EXISTANT": _amount_for(LibellePaiement.SOLDE),
        "DATE_SOLDE": _date_for(LibellePaiement.SOLDE),
        "ETAPE_ACTUELLE_A_ENCODER": payload["current_label"],
        "MONTANT_A_ENCODER": None,
        "DATE_PAIEMENT_A_ENCODER": None,
        "MODE_PAIEMENT": "",
        "ENVOYER_EMAIL": "NON",
        "NOTE_INTERNE": "",
    }




def _payment_excel_base_queryset(circuit_id=None):
    qs = (
        Inscription.objects
        .select_related("circuit", "pilote", "passager")
        .prefetch_related("paiements", "selections_options__option")
        .order_by("circuit__date_debut", "circuit__code", "pilote__nom", "pilote__prenom", "id")
    )
    if circuit_id:
        qs = qs.filter(circuit_id=circuit_id)
    return qs


def _payment_excel_circuits_queryset():
    return (
        Circuit.objects
        .annotate(nb_inscriptions=Count("inscriptions"))
        .filter(nb_inscriptions__gt=0)
        .order_by("date_debut", "code", "nom")
    )


def _payment_excel_sheet_name(circuit, used_titles=None):
    used_titles = used_titles if used_titles is not None else set()
    base = f"{circuit.code or circuit.nom or circuit.pk} - {circuit.nom or ''}".strip(" -")
    sanitized = re.sub(r'[\\/*?:\[\]]', '-', base).strip() or f"Circuit {circuit.pk}"
    sanitized = sanitized[:31]
    candidate = sanitized
    idx = 2
    while candidate in used_titles:
        suffix = f"-{idx}"
        candidate = f"{sanitized[:31-len(suffix)]}{suffix}"
        idx += 1
    used_titles.add(candidate)
    return candidate


def _payment_excel_intro_sheet(ws, circuits):
    ws.title = "Accueil"
    ws["A1"] = "Classeur multi-circuits — paiements"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Généré le"
    ws["B2"] = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")
    ws["A4"] = "Mode d’emploi"
    ws["A4"].font = Font(bold=True)
    instructions = [
        "1. Chaque onglet après ‘Accueil’ correspond à un circuit.",
        "2. Ne modifiez que les colonnes jaunes : montant, date, mode, email, note.",
        "3. Une seule étape est ouverte par ligne : ACOMPTE_1, ACOMPTE_2, SOLDE ou TERMINE.",
        "4. À l’import, le système relit toutes les feuilles du classeur et applique toutes les lignes valides.",
        "5. Les lignes déjà soldées ou incohérentes sont ignorées ou signalées en erreur.",
    ]
    for idx, line in enumerate(instructions, start=5):
        ws.cell(row=idx, column=1, value=line)
    start = 12
    headers = ["Feuille", "Circuit ID", "Code", "Circuit", "Date début", "Inscriptions"]
    used_titles = set(["Accueil"])
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, circuit in enumerate(circuits, start=start + 1):
        ws.cell(row=row_idx, column=1, value=_payment_excel_sheet_name(circuit, used_titles))
        ws.cell(row=row_idx, column=2, value=circuit.pk)
        ws.cell(row=row_idx, column=3, value=circuit.code)
        ws.cell(row=row_idx, column=4, value=circuit.nom)
        ws.cell(row=row_idx, column=5, value=circuit.date_debut)
        ws.cell(row=row_idx, column=6, value=getattr(circuit, "nb_inscriptions", ""))
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    for row in range(start + 1, ws.max_row + 1):
        ws[f"E{row}"].number_format = "DD/MM/YYYY"


def _parse_excel_date(value):
    if not value:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return dt.datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Date invalide: {raw}")


def _parse_decimal_excel(value):
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    raw = str(value).strip().replace(" ", " ").replace(" ", "")
    if not raw:
        return Decimal("0")
    raw = raw.replace(",", ".")
    return Decimal(raw)


def _normalize_payment_method(value):
    raw = (str(value or "").strip() or "VIREMENT").upper()
    return PAYMENT_METHOD_NORMALIZATION.get(raw, "")


def _normalize_yes_no(value):
    raw = (str(value or "").strip()).upper()
    return raw in {"OUI", "YES", "Y", "TRUE", "1"}


def _editable_payment_columns():
    return {
        "MONTANT_A_ENCODER",
        "DATE_PAIEMENT_A_ENCODER",
        "MODE_PAIEMENT",
        "ENVOYER_EMAIL",
        "NOTE_INTERNE",
    }


def _send_payment_excel_email(*, ins, request_user, stage_label, amount, payment_date, recipient=None):
    recipient = recipient or ins.pilote or ins.passager
    if not recipient or not getattr(recipient, "email", ""):
        return False, "Aucun destinataire email disponible."

    role = "pilote" if ins.pilote_id and recipient.pk == ins.pilote_id else "passager"
    payment_rows = _build_email_payment_rows(ins)

    total_attendu = Decimal(ins.total_attendu() or 0)
    total_paye = Decimal(sum((p.montant_encaisse or 0) for p in ins.paiements.all()))
    reste_a_payer = max(total_attendu - total_paye, Decimal("0"))
    trop_percu = max(total_paye - total_attendu, Decimal("0"))
    montant_options_ajoutees = Decimal(getattr(ins, "montant_options_ajoutees", 0) or 0)

    stage_text = {
        "ACOMPTE_1": "Acompte 1",
        "ACOMPTE_2": "Acompte 2",
        "SOLDE": "Solde",
    }.get(stage_label, stage_label)

    ctx = {
        "ins": ins,
        "recipient": recipient,
        "dest_name": recipient.prenom or recipient.nom or "client",
        "circuit": ins.circuit,
        "payment_rows": payment_rows,
        "intro_message": f"Bonjour,\n\nNous confirmons l’enregistrement de votre paiement ({stage_text.lower()}).",
        "note_client": "Merci. Ce paiement a été enregistré via notre suivi administratif.",
        "total_attendu_initial": total_attendu,
        "montant_options_ajoutees": montant_options_ajoutees,
        "nouveau_total_attendu": total_attendu,
        "total_attendu": total_attendu,
        "total_paye": total_paye,
        "reste_a_payer": reste_a_payer,
        "trop_percu": trop_percu,
        "situation_label": "Soldé" if reste_a_payer <= 0 else "Reste à payer",
        "edit_url": _abs_url(reverse("inscription_edit_start", args=[ins.id_public])),
        "infos_paiement": (getattr(ins.circuit, "infos_paiement", "") or "").strip(),
        "has_infos_paiement": bool((getattr(ins.circuit, "infos_paiement", "") or "").strip()),
        "paiement_libelle": "Paiement manuel",
        "subject": f"[{ins.circuit.code}] Paiement enregistré – {ins.circuit.nom}",
        "sent_acompte1_amount": amount if stage_label == "ACOMPTE_1" else Decimal("0"),
        "sent_acompte2_amount": amount if stage_label == "ACOMPTE_2" else Decimal("0"),
        "sent_solde_amount": amount if stage_label == "SOLDE" else Decimal("0"),
        "sent_options_amount": Decimal("0"),
        "has_sent_acompte1": stage_label == "ACOMPTE_1",
        "has_sent_acompte2": stage_label == "ACOMPTE_2",
        "has_sent_solde": stage_label == "SOLDE",
        "has_sent_options": False,
        "date_paiement": payment_date,
        "is_reminder": False,
    }

    sent = notify_paiement_resume_compose(
        recipient_email=recipient.email,
        subject=ctx["subject"],
        ctx=ctx,
        role=role,
    )
    return bool(sent), recipient.email


def _apply_excel_payment_stage(*, ins, stage_label, amount, payment_date, method, note, actor):
    libelle = PAYMENT_STAGE_REVERSE.get(stage_label)
    if not libelle:
        raise ValidationError("Étape de paiement inconnue.")

    mapping = {
        LibellePaiement.ACOMPTE1: {"acompte1_amount": amount, "acompte2_amount": Decimal("0"), "solde_amount": Decimal("0")},
        LibellePaiement.ACOMPTE2: {"acompte1_amount": Decimal("0"), "acompte2_amount": amount, "solde_amount": Decimal("0")},
        LibellePaiement.SOLDE: {"acompte1_amount": Decimal("0"), "acompte2_amount": Decimal("0"), "solde_amount": amount},
    }[libelle]

    _apply_manual_payments_to_inscription(ins, payment_date=payment_date, **mapping)
    p = ins.paiements.filter(libelle=libelle).first()
    if not p:
        raise ValidationError("Ligne de paiement introuvable sur l’inscription.")

    fields = []
    if method:
        p.methode = method
        fields.append("methode")
    if note:
        p.reference = (note or "")[:120]
        fields.append("reference")
    if fields:
        fields.append("modifie_le")
        p.save(update_fields=fields)

    JournalAudit.objects.create(
        inscription=ins,
        acteur=actor,
        action="import_excel_paiement",
        donnees={
            "stage": stage_label,
            "amount": str(amount),
            "payment_date": str(payment_date),
            "method": method or "",
            "note": note or "",
        },
    )
    return p


def _payment_excel_error_row(*, sheet_name, circuit_label, row_number, reference="", client="", stage="", amount="", payment_date="", method="", send_email=False, message=""):
    return {
        "sheet_name": sheet_name,
        "circuit_label": circuit_label,
        "row_number": row_number,
        "reference": reference,
        "client": client,
        "stage": stage,
        "amount": str(amount) if amount not in (None, "") else "",
        "payment_date": payment_date,
        "method": method,
        "send_email": send_email,
        "status": "error",
        "message": message,
        "inscription_id": None,
    }


def _validate_payment_excel_sheet(ws):
    header_row_index = 5
    headers = [ws.cell(row=header_row_index, column=i).value for i in range(1, len(PAYMENT_EXCEL_HEADERS) + 1)]
    normalized_headers = [str(h).strip() if h is not None else "" for h in headers]
    if normalized_headers != PAYMENT_EXCEL_HEADERS:
        raise ValidationError("Le fichier importé ne correspond pas au modèle attendu.")


def _build_payment_excel_preview(*, workbook_file, request_user):
    wb = load_workbook(workbook_file, data_only=True)
    circuits = list(_payment_excel_circuits_queryset())
    circuits_by_id = {c.pk: c for c in circuits}
    preview_rows = []
    actionable_rows = []
    stats = {"ready": 0, "errors": 0, "ignored": 0, "emails": 0, "circuits": 0}
    sheet_names_seen = []
    ins_cache = {}

    for ws in wb.worksheets:
        sheet_name = ws.title
        if sheet_name == "Accueil":
            continue
        sheet_names_seen.append(sheet_name)

        try:
            _validate_payment_excel_sheet(ws)
        except Exception as exc:
            preview_rows.append(_payment_excel_error_row(
                sheet_name=sheet_name,
                circuit_label="",
                row_number="-",
                message=f"Onglet invalide: {exc}",
            ))
            stats["errors"] += 1
            continue

        meta_circuit_id = ws["B2"].value
        try:
            circuit_id = int(str(meta_circuit_id).strip())
        except Exception:
            circuit_id = None
        circuit = circuits_by_id.get(circuit_id)
        if not circuit:
            preview_rows.append(_payment_excel_error_row(
                sheet_name=sheet_name,
                circuit_label="",
                row_number="-",
                message="Circuit introuvable pour cet onglet.",
            ))
            stats["errors"] += 1
            continue

        circuit_label = f"{circuit.code} — {circuit.nom}"
        stats["circuits"] += 1
        if circuit.pk not in ins_cache:
            ins_cache[circuit.pk] = {ins.pk: ins for ins in _payment_excel_base_queryset(circuit.pk)}
        ins_by_id = ins_cache[circuit.pk]

        for row_idx in range(6, ws.max_row + 1):
            values = {PAYMENT_EXCEL_HEADERS[col_idx - 1]: ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(PAYMENT_EXCEL_HEADERS) + 1)}
            ins_id_raw = values.get("ID_INSCRIPTION")
            reference = str(values.get("REFERENCE") or "").strip()
            amount = _parse_decimal_excel(values.get("MONTANT_A_ENCODER"))
            note = str(values.get("NOTE_INTERNE") or "").strip()
            wants_email = _normalize_yes_no(values.get("ENVOYER_EMAIL"))
            client = f"{values.get('NOM') or ''} {values.get('PRENOM') or ''}".strip()

            if not ins_id_raw and not reference and amount <= 0 and not note and not wants_email:
                continue

            row_result = {
                "sheet_name": sheet_name,
                "circuit_label": circuit_label,
                "row_number": row_idx,
                "reference": reference,
                "client": client,
                "stage": str(values.get("ETAPE_ACTUELLE_A_ENCODER") or "").strip(),
                "amount": str(amount),
                "payment_date": "",
                "method": "",
                "send_email": wants_email,
                "status": "ignored",
                "message": "",
                "inscription_id": None,
            }

            try:
                ins_id = int(ins_id_raw)
            except (TypeError, ValueError):
                ins_id = None

            if not ins_id or ins_id not in ins_by_id:
                row_result["status"] = "error"
                row_result["message"] = "Inscription introuvable pour ce circuit."
                preview_rows.append(row_result)
                stats["errors"] += 1
                continue

            ins = ins_by_id[ins_id]
            row_result["inscription_id"] = ins.pk

            if str(ins.id_public) != reference:
                row_result["status"] = "error"
                row_result["message"] = "La référence du fichier ne correspond pas à l’inscription."
                preview_rows.append(row_result)
                stats["errors"] += 1
                continue

            if amount <= 0:
                row_result["status"] = "ignored"
                row_result["message"] = "Aucun montant saisi : ligne ignorée."
                preview_rows.append(row_result)
                stats["ignored"] += 1
                continue

            payload = _payment_stage_payload(ins)
            current_stage = payload["current_label"]
            row_result["stage"] = current_stage

            if current_stage == "TERMINE":
                row_result["status"] = "error"
                row_result["message"] = "Cette inscription est déjà soldée."
                preview_rows.append(row_result)
                stats["errors"] += 1
                continue

            file_stage = str(values.get("ETAPE_ACTUELLE_A_ENCODER") or "").strip()
            if file_stage and file_stage != current_stage:
                row_result["status"] = "error"
                row_result["message"] = f"Le fichier est obsolète : étape attendue {current_stage}."
                preview_rows.append(row_result)
                stats["errors"] += 1
                continue

            current_payment = payload["current_payment"]
            stage_remaining = Decimal(current_payment.montant_du or 0) - Decimal(current_payment.montant_encaisse or 0)
            if amount > stage_remaining:
                row_result["status"] = "error"
                row_result["message"] = f"Montant supérieur au reste attendu pour {current_stage.lower().replace('_', ' ')} ({stage_remaining})."
                preview_rows.append(row_result)
                stats["errors"] += 1
                continue

            try:
                payment_date = _parse_excel_date(values.get("DATE_PAIEMENT_A_ENCODER")) or timezone.localdate()
            except ValueError as exc:
                row_result["status"] = "error"
                row_result["message"] = str(exc)
                preview_rows.append(row_result)
                stats["errors"] += 1
                continue

            method = _normalize_payment_method(values.get("MODE_PAIEMENT"))
            if not method:
                row_result["status"] = "error"
                row_result["message"] = "Mode de paiement invalide. Utilisez espèces, virement, chèque ou mobile money."
                preview_rows.append(row_result)
                stats["errors"] += 1
                continue

            row_result["payment_date"] = payment_date.strftime("%Y-%m-%d")
            row_result["method"] = method
            row_result["status"] = "ready"
            row_result["message"] = f"{current_stage.replace('_', ' ').title()} prêt à être enregistré."
            preview_rows.append(row_result)
            stats["ready"] += 1
            if wants_email:
                stats["emails"] += 1

            actionable_rows.append({
                "sheet_name": sheet_name,
                "circuit_id": circuit.pk,
                "inscription_id": ins.pk,
                "reference": reference,
                "stage": current_stage,
                "amount": str(amount),
                "payment_date": payment_date.isoformat(),
                "method": method,
                "send_email": wants_email,
                "note": note,
                "row_number": row_idx,
            })

    token = secrets.token_urlsafe(16)
    return {
        "token": token,
        "rows": preview_rows,
        "actionable_rows": actionable_rows,
        "stats": stats,
        "sheet_names": sheet_names_seen,
        "generated_at": timezone.now().isoformat(),
        "actor_id": getattr(request_user, "pk", None),
    }


def _payment_import_session_key(token):
    return f"payment_excel_preview::{token}"


@staff_member_required
def payment_excel_template_xlsx(request):
    circuits = list(_payment_excel_circuits_queryset())
    wb = Workbook()
    intro_ws = wb.active
    _payment_excel_intro_sheet(intro_ws, circuits)

    used_titles = set(["Accueil"])
    editable_fill = PatternFill("solid", fgColor="FEF3C7")
    done_fill = PatternFill("solid", fgColor="DCFCE7")
    a1_fill = PatternFill("solid", fgColor="FEF3C7")
    a2_fill = PatternFill("solid", fgColor="DBEAFE")
    solde_fill = PatternFill("solid", fgColor="FCE7F3")
    meta_fill = PatternFill("solid", fgColor="E0F2FE")

    for circuit in circuits:
        inscriptions = list(_payment_excel_base_queryset(circuit.pk))
        ws = wb.create_sheet(title=_payment_excel_sheet_name(circuit, used_titles))
        ws["A1"] = "Circuit"
        ws["B1"] = f"{circuit.code} — {circuit.nom}"
        ws["A2"] = "Circuit ID"
        ws["B2"] = str(circuit.pk)
        ws["A3"] = "Généré le"
        ws["B3"] = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")
        for cell_ref in ("A1", "A2", "A3"):
            ws[cell_ref].font = Font(bold=True)
            ws[cell_ref].fill = meta_fill

        header_row = 5
        for idx, header in enumerate(PAYMENT_EXCEL_HEADERS, start=1):
            cell = ws.cell(row=header_row, column=idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D4ED8")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, ins in enumerate(inscriptions, start=6):
            row_data = _payment_row_for_export(ins)
            for col_idx, header in enumerate(PAYMENT_EXCEL_HEADERS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
                if header in _editable_payment_columns():
                    cell.fill = editable_fill
                    cell.protection = Protection(locked=False)
                else:
                    cell.protection = Protection(locked=True)
            stage = row_data["ETAPE_ACTUELLE_A_ENCODER"]
            fill = {"TERMINE": done_fill, "ACOMPTE_1": a1_fill, "ACOMPTE_2": a2_fill, "SOLDE": solde_fill}.get(stage)
            if fill:
                for col_idx in range(1, len(PAYMENT_EXCEL_HEADERS) + 1):
                    header = PAYMENT_EXCEL_HEADERS[col_idx - 1]
                    if header not in _editable_payment_columns():
                        ws.cell(row=row_idx, column=col_idx).fill = fill

        ws.freeze_panes = "A6"
        ws.auto_filter.ref = f"A5:X{max(ws.max_row, 5)}"
        ws.column_dimensions["A"].hidden = True
        for col in ("D", "N", "P", "R", "U"):
            for row in range(6, ws.max_row + 1):
                ws[f"{col}{row}"].number_format = "DD/MM/YYYY"
        for col in ("J", "K", "L", "M", "O", "Q", "T"):
            for row in range(6, ws.max_row + 1):
                ws[f"{col}{row}"].number_format = "#,##0.00"
        widths = {
            "A": 5, "B": 40, "C": 30, "D": 15, "E": 18, "F": 18, "G": 28, "H": 18,
            "I": 18, "J": 14, "K": 18, "L": 14, "M": 15, "N": 14, "O": 15, "P": 14,
            "Q": 15, "R": 14, "S": 20, "T": 18, "U": 18, "V": 18, "W": 14, "X": 26,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.protection.sheet = True
        ws.protection.enable()

    filename = f"paiements_multi_circuits_{timezone.now().strftime('%Y%m%d-%H%M')}.xlsx"
    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)
    response = HttpResponse(
        buff.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@staff_member_required
def payment_excel_import(request):
    circuits = list(_payment_excel_circuits_queryset())
    preview = None

    if request.method == "POST":
        action = (request.POST.get("action") or "preview").strip()

        if action == "preview":
            if "excel_file" not in request.FILES:
                messages.error(request, "Ajoutez le fichier Excel multi-circuits exporté depuis le système.")
            else:
                try:
                    preview = _build_payment_excel_preview(
                        workbook_file=request.FILES["excel_file"],
                        request_user=request.user,
                    )
                    request.session[_payment_import_session_key(preview["token"])] = preview
                    request.session.modified = True
                    if not preview["rows"]:
                        messages.warning(request, "Le fichier ne contient aucune ligne exploitable.")
                    elif preview["stats"]["errors"]:
                        messages.warning(request, "Prévisualisation générée avec anomalies. Corrige les lignes en erreur avant validation.")
                    else:
                        messages.success(request, "Prévisualisation prête. Vérifie les lignes puis applique l’import global.")
                except Exception as exc:
                    messages.error(request, f"Import impossible: {exc}")

        elif action == "apply":
            token = (request.POST.get("preview_token") or "").strip()
            session_key = _payment_import_session_key(token)
            preview = request.session.get(session_key)
            if not preview:
                messages.error(request, "La prévisualisation a expiré. Réimporte le fichier Excel.")
            elif preview.get("actor_id") != request.user.pk:
                messages.error(request, "Cette prévisualisation n’appartient pas à votre session.")
                preview = None
            else:
                actionable_rows = preview.get("actionable_rows") or []
                if not actionable_rows:
                    messages.warning(request, "Aucune ligne prête à être appliquée.")
                elif any(r.get("status") == "error" for r in preview.get("rows", [])):
                    messages.error(request, "Corrige d’abord les lignes en erreur puis regénère la prévisualisation.")
                else:
                    applied_count = 0
                    emailed_count = 0
                    email_errors = []
                    try:
                        rows_by_circuit = {}
                        for row in actionable_rows:
                            rows_by_circuit.setdefault(row["circuit_id"], []).append(row)
                        email_queue = []
                        with transaction.atomic():
                            ins_by_circuit = {
                                circuit_id: {ins.pk: ins for ins in _payment_excel_base_queryset(circuit_id)}
                                for circuit_id in rows_by_circuit.keys()
                            }
                            for circuit_id, rows in rows_by_circuit.items():
                                ins_by_id = ins_by_circuit[circuit_id]
                                for row in rows:
                                    ins = ins_by_id.get(row["inscription_id"])
                                    if not ins:
                                        raise ValidationError(f"Inscription introuvable pour la ligne {row['row_number']}")
                                    fresh_payload = _payment_stage_payload(ins)
                                    if fresh_payload["current_label"] != row["stage"]:
                                        raise ValidationError(
                                            f"La ligne {row['row_number']} ({row['sheet_name']}) a changé depuis la prévisualisation (étape attendue: {fresh_payload['current_label']})."
                                        )
                                    _apply_excel_payment_stage(
                                        ins=ins,
                                        stage_label=row["stage"],
                                        amount=Decimal(row["amount"]),
                                        payment_date=dt.date.fromisoformat(row["payment_date"]),
                                        method=row["method"],
                                        note=row.get("note") or "",
                                        actor=request.user,
                                    )
                                    applied_count += 1
                                    ins.refresh_from_db()
                                    if row.get("send_email"):
                                        email_queue.append((ins.pk, row))
                        for ins_id, row in email_queue:
                            try:
                                ins = _payment_excel_base_queryset().get(pk=ins_id)
                                sent, info = _send_payment_excel_email(
                                    ins=ins,
                                    request_user=request.user,
                                    stage_label=row["stage"],
                                    amount=Decimal(row["amount"]),
                                    payment_date=dt.date.fromisoformat(row["payment_date"]),
                                )
                                if sent:
                                    emailed_count += 1
                                else:
                                    email_errors.append(f"{row['sheet_name']} ligne {row['row_number']}: {info}")
                            except Exception as exc:
                                email_errors.append(f"{row['sheet_name']} ligne {row['row_number']}: {exc}")
                        request.session.pop(session_key, None)
                        request.session.modified = True
                        if email_errors:
                            messages.warning(request, f"Import appliqué: {applied_count} paiement(s) enregistrés, {emailed_count} email(s) envoyés. Erreurs email: {' | '.join(email_errors[:3])}{' …' if len(email_errors) > 3 else ''}")
                        else:
                            messages.success(request, f"Import global appliqué: {applied_count} paiement(s) enregistrés, {emailed_count} email(s) envoyés.")
                        return redirect(reverse("bo_payment_excel_import"))
                    except Exception as exc:
                        messages.error(request, f"Application annulée: {exc}")
                        preview = request.session.get(session_key)

    elif request.GET.get("preview_token"):
        preview = request.session.get(_payment_import_session_key(request.GET.get("preview_token")))

    return render(request, "circuitMoto/admin/payment_excel_import.html", {
        "circuits": circuits,
        "preview": preview,
    })

@staff_member_required
def reservations_export_xlsx(request):
    """
    Export de suivi interne proche du fichier de reservation historique:
    une feuille par circuit, une ligne par participant, paiements sur la ligne pilote.
    """
    qs, year, circuit_id, statut = _reservations_export_queryset(request)
    inscriptions = list(qs)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    used_titles = set()
    by_circuit = {}
    for inscription in inscriptions:
        by_circuit.setdefault(inscription.circuit_id, {"circuit": inscription.circuit, "items": []})
        by_circuit[inscription.circuit_id]["items"].append(inscription)

    if by_circuit:
        ordered_groups = sorted(
            by_circuit.values(),
            key=lambda item: (
                item["circuit"].date_debut or date.max,
                item["circuit"].code or "",
            )
        )
        for group in ordered_groups:
            _append_reservation_sheet(wb, group["circuit"], group["items"], used_titles)
    else:
        ws = wb.create_sheet(title="Reservations")
        ws.append(["Aucune inscription pour les filtres selectionnes"])

    now = timezone.now().strftime("%Y%m%d-%H%M")
    part_year = year or "toutes-annees"
    circuit_code = ""
    if circuit_id:
        circuit = Circuit.objects.filter(pk=circuit_id).only("code").first()
        circuit_code = circuit.code if circuit and circuit.code else f"id{circuit_id}"
    part_circuit = slugify(circuit_code or "tous-circuits") or "tous-circuits"
    part_statut = f"_{slugify(statut)}" if statut else ""
    filename = f"suivi_reservations_{part_year}_{part_circuit}{part_statut}_{now}.xlsx"

    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)

    response = HttpResponse(
        buff.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@staff_member_required
def personnes_export_xlsx(request):
    """
    Excel: 1 ligne par personne présente dans une inscription (pilote + passager s'il existe),
    filtrable par année (?year) OU circuit (?circuit) et statut (?statut).
    Enrichi avec: infos médicales, assurance, 2 contacts d'urgence, options.
    """
    year       = (request.GET.get("year") or "").strip()
    circuit_id = (request.GET.get("circuit") or "").strip()
    statut     = (request.GET.get("statut") or "").strip()

    # --- Préchargements costauds pour éviter le N+1 --------------------------------
    qs = (
        Inscription.objects
        .select_related(
            "circuit",
            "assurance",
            "pilote", "passager",
            "pilote__infos_medicales", "passager__infos_medicales",
        )
        .prefetch_related(
            Prefetch("selections_options", queryset=SelectionOption.objects.select_related("option")),
            Prefetch("pilote__contacts_urgence"),
            Prefetch("passager__contacts_urgence"),
        )
        .order_by("circuit__date_debut", "pilote__nom", "pilote__prenom")
    )

    if year:
        qs = qs.filter(circuit__date_debut__year=year)
    if circuit_id:
        qs = qs.filter(circuit_id=circuit_id)
    if statut:
        qs = qs.filter(statut=statut)

    # --- Définition des colonnes ---------------------------------------------------
    # 1) Identité de la personne
    person_cols = [
        ("nom",                     "Nom"),
        ("prenom",                  "Prénom"),
        ("email",                   "Email"),
        ("telephone",               "Téléphone"),
        ("date_naissance",          "Date de naissance"),
        ("age",                     "Âge"),
        ("numero_carte_identite",   "N° carte d’identité"),
        ("adresse",                 "Adresse"),
        ("code_postal",             "Code postal"),
        ("localite",                "Localité"),
        ("pays",                    "Pays"),
    ]

    # 2) Contexte inscription
    context_cols = [
        ("role",            "Rôle"),
        ("statut",          "Statut inscription"),
        ("annee",           "Année"),
        ("circuit_code",    "Code circuit"),
        ("circuit_nom",     "Nom circuit"),
        ("circuit_debut",   "Début"),
        ("circuit_fin",     "Fin"),
        ("inscription_id",  "ID interne"),
        ("id_public",       "ID public"),
        ("cree_le",         "Inscription créée le"),
    ]

    # 3) Infos médicales (attachées à la Personne)
    medical_cols = [
        ("med_groupe",      "Groupe sanguin"),
        ("med_hta",         "HTA"),
        ("med_asthme",      "Asthme"),
        ("med_epilepsie",   "Épilepsie"),
        ("med_peau",        "Problèmes de peau"),
        ("med_vertiges",    "Vertiges"),
        ("med_notes",       "Notes médicales"),
    ]

    # 4) Assurance (attachée à l'Inscription)
    assurance_cols = [
        ("ass_type",        "Assurance - Type"),
        ("ass_compagnie",   "Assurance - Compagnie"),
        ("ass_police",      "Assurance - N° police"),
        ("ass_valide_du",   "Ass. valable du"),
        ("ass_valide_au",   "Ass. valable au"),
        ("ass_tel_urg",     "Ass. N° d’urgence"),
    ]

    # 5) Contacts d'urgence (on exporte les 2 premiers)
    MAX_CONTACTS = 2
    contact_cols = []
    for i in range(1, MAX_CONTACTS+1):
        contact_cols += [
            (f"cu{i}_nom",   f"Contact {i} - Nom"),
            (f"cu{i}_lien",  f"Contact {i} - Lien"),
            (f"cu{i}_tel",   f"Contact {i} - Téléphone"),
        ]

    # 6) Options (attachées à l'Inscription)
    options_cols = [
        ("opts_synthese",   "Options (code×qte)"),
        ("opts_detail",     "Options (intitulés)"),
        ("opts_total",      "Montant options"),
    ]

    # --- Workbook/Feuille ----------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Inscrits"

    header = [label for _, label in (person_cols + context_cols + medical_cols + assurance_cols + contact_cols + options_cols)]
    ws.append(header)
    header_fill = PatternFill("solid", fgColor="EEF2FF")
    header_font = Font(bold=True)
    for col_idx in range(1, len(header) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    def _as_date(v):
        """Retourne un objet date (naïf) pour Excel."""
        if v is None:
            return None
        if isinstance(v, dt.datetime):
            return v.date()
        if isinstance(v, dt.date):
            return v
        return v

    def _medical_vals(person):
        m = getattr(person, "infos_medicales", None)
        return {
            "med_groupe":   (m.groupe_sanguin if m and m.groupe_sanguin else ""),
            "med_hta":      bool(m and m.hta),
            "med_asthme":   bool(m and m.asthme),
            "med_epilepsie":bool(m and m.epilepsie),
            "med_peau":     bool(m and m.problemes_peau),
            "med_vertiges": bool(m and m.vertiges),
            "med_notes":    (m.notes or "") if m else "",
        }

    def _assurance_vals(ins):
        a = getattr(ins, "assurance", None)
        return {
            "ass_type":      (a.get_type_display() if a else ""),
            "ass_compagnie": (a.compagnie if a else ""),
            "ass_police":    (a.numero_police if a else ""),
            "ass_valide_du": _as_date(a.valide_du) if a else None,
            "ass_valide_au": _as_date(a.valide_au) if a else None,
            "ass_tel_urg":   (a.telephone_urgence if a else ""),
        }

    def _contact_vals(person):
        rows = {"cu1_nom":"", "cu1_lien":"", "cu1_tel":"", "cu2_nom":"", "cu2_lien":"", "cu2_tel":""}
        if not person:
            return rows
        contacts = list(getattr(person, "contacts_urgence", []).all()) if hasattr(person, "contacts_urgence") else []
        # garde-fou: ordonner par -cree_le si dispo
        try:
            contacts.sort(key=lambda c: getattr(c, "cree_le", dt.datetime.min), reverse=True)
        except Exception:
            pass
        for idx, c in enumerate(contacts[:MAX_CONTACTS], start=1):
            rows[f"cu{idx}_nom"]  = c.nom or ""
            rows[f"cu{idx}_lien"] = c.lien_parente or ""
            rows[f"cu{idx}_tel"]  = c.telephone or ""
        return rows

    def _options_vals(ins):
        sels = list(getattr(ins, "selections_options", []).all()) if hasattr(ins, "selections_options") else []
        if not sels:
            return {"opts_synthese":"", "opts_detail":"", "opts_total":0}
        synth = []
        detail = []
        total = 0
        for s in sels:
            code = s.option.code if s.option_id else ""
            intitule = s.option.intitule if s.option_id else ""
            q = s.quantite or 0
            synth.append(f"{code}×{q}")
            who = " (passager)" if s.pour_passager else ""
            detail.append(f"{intitule}×{q}{who}")
            total += (s.prix_unitaire_fige or 0) * q
        return {
            "opts_synthese": " | ".join(synth)[:500],
            "opts_detail":   " | ".join(detail)[:500],
            "opts_total":    total,
        }

    # --- Construction d'une ligne --------------------------------------------------
    def add_row(ins, person, role: str):
        if not person:
            return
        circ = ins.circuit

        # PERSONNE
        pvals = {
            "nom": person.nom or "",
            "prenom": person.prenom or "",
            "email": person.email or "",
            "telephone": person.telephone or "",
            "date_naissance": _as_date(getattr(person, "date_naissance", None)),
            "age": getattr(person, "age", None) or "",
            "numero_carte_identite": person.numero_carte_identite or "",
            "adresse": person.adresse or "",
            "code_postal": person.code_postal or "",
            "localite": person.localite or "",
            "pays": person.pays or "",
        }

        # CONTEXTE
        cvals = {
            "role": role,
            "statut": ins.statut,
            "annee": circ.date_debut.year if circ and circ.date_debut else "",
            "circuit_code": getattr(circ, "code", ""),
            "circuit_nom": getattr(circ, "nom", ""),
            "circuit_debut": _as_date(getattr(circ, "date_debut", None)),
            "circuit_fin": _as_date(getattr(circ, "date_fin", None)),
            "inscription_id": ins.id,
            "id_public": str(ins.id_public),
            "cree_le": _as_date(getattr(ins, "cree_le", None)),
        }

        mvals = _medical_vals(person)
        avals = _assurance_vals(ins)
        cuvals = _contact_vals(person)
        ovals = _options_vals(ins)

        row = []
        for key, _ in person_cols:   row.append(pvals.get(key, ""))
        for key, _ in context_cols:  row.append(cvals.get(key, ""))
        for key, _ in medical_cols:  row.append(_safe_str(mvals.get(key, "")))
        for key, _ in assurance_cols:
            val = avals.get(key, "")
            row.append(val if "valide_" not in key else _as_date(val))
        for key, _ in contact_cols:  row.append(cuvals.get(key, ""))
        for key, _ in options_cols:  row.append(ovals.get(key, ""))

        ws.append(row)

    # --- Alimentation : 1 ligne par personne de l’inscription ---------------------
    for ins in qs:
        add_row(ins, ins.pilote, "Pilote")
        if ins.passager_id:
            add_row(ins, ins.passager, "Passager")

    # --- Formats & largeurs --------------------------------------------------------
    date_headers = {
        "Date de naissance", "Début", "Fin", "Inscription créée le",
        "Ass. valable du", "Ass. valable au",
    }
    header_map = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}

    # Dates
    for h in date_headers:
        idx = header_map.get(h)
        if idx:
            for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                r[0].number_format = "DD/MM/YYYY"

    # Montants (options)
    idx_opts_total = header_map.get("Montant options")
    if idx_opts_total:
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx_opts_total, max_col=idx_opts_total):
            r[0].number_format = "#,##0"

    # Largeur auto raisonnable
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for cell in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            v = cell[0].value
            max_len = max(max_len, len(_safe_str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 50)

    # --- Nom du fichier ------------------------------------------------------------
    now = timezone.now().strftime("%Y%m%d-%H%M")
    part_year = year or "toutes-annees"
    circuit_code = ""
    if circuit_id:
        c = Circuit.objects.filter(pk=circuit_id).only("code").first()
        circuit_code = (c.code if c and c.code else f"id{circuit_id}")
    part_circ = circuit_code or "tous-circuits"
    part_statut = f"_{statut.lower()}" if statut else ""
    filename = f"inscrits_{part_year}_{part_circ}{part_statut}_{now}.xlsx"

    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)

    resp = HttpResponse(
        buff.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

@staff_member_required
def bo_backup_download(request):
    if not request.user.is_superuser:
        raise Http404("Page introuvable.")

    if request.method != "POST":
        raise Http404("Méthode non autorisée.")

    try:
        archive_path, archive_name = create_full_backup_zip()

        JournalAudit.objects.create(
            acteur=request.user,
            action="backup_download",
            donnees={
                "filename": archive_name,
                "created_at": timezone.now().isoformat(),
            },
        )

        response = FileResponse(
            open(archive_path, "rb"),
            as_attachment=True,
            filename=archive_name,
        )

        return response

    except BackupError as exc:
        messages.error(request, f"Sauvegarde impossible : {exc}")
        raise Http404(str(exc))

    except Exception:
        messages.error(request, "Une erreur inattendue est survenue lors de la génération de la sauvegarde.")
        raise Http404("Erreur de sauvegarde.")
    

# =========================================================
# BALADES 1 JOUR — BACK OFFICE
# =========================================================

class BaladeJourForm(forms.ModelForm):
    class Meta:
        model = BaladeJour
        fields = [
            "date_debut",
            "date_fin",
            "date_affichage",
            "titre",
            "guide",
            "actif",
            "ordre",
        ]
        widgets = {
            "date_debut": forms.DateInput(attrs={"type": "date", "class": "form-control"}),
            "date_fin": forms.DateInput(attrs={"type": "date", "class": "form-control"}),
            "date_affichage": forms.TextInput(attrs={
                "class": "form-control",
                "placeholder": "Ex: 9 et 10/5",
            }),
            "titre": forms.TextInput(attrs={
                "class": "form-control",
                "placeholder": "Ex: Moto Zenith Huy : Namur - le Condroz",
            }),
            "guide": forms.TextInput(attrs={
                "class": "form-control",
                "placeholder": "Ex: Thomas / Jean Louis / libre",
            }),
            "actif": forms.CheckboxInput(attrs={"class": "form-check-input"}),
            "ordre": forms.NumberInput(attrs={
                "class": "form-control",
                "min": "0",
                "step": "1",
            }),
        }
        labels = {
            "date_debut": "Date de début",
            "date_fin": "Date de fin",
            "date_affichage": "Date affichée personnalisée",
            "titre": "Titre / circuit",
            "guide": "Guide",
            "actif": "Balade active",
            "ordre": "Ordre d’affichage",
        }
        help_texts = {
            "date_affichage": "Optionnel. Exemple : 9 et 10/5. Si vide, la date est générée automatiquement.",
            "date_fin": "Laissez vide pour une balade sur une seule journée.",
            "ordre": "Permet d’ajuster manuellement l’ordre d’affichage à date égale.",
        }

    def clean(self):
        cleaned = super().clean()
        date_debut = cleaned.get("date_debut")
        date_fin = cleaned.get("date_fin")
        titre = (cleaned.get("titre") or "").strip()

        if not titre:
            self.add_error("titre", "Le titre de la balade est obligatoire.")

        if date_debut and date_fin and date_fin < date_debut:
            self.add_error("date_fin", "La date de fin doit être postérieure ou égale à la date de début.")

        return cleaned


@staff_member_required
def balade_list(request):
    q = (request.GET.get("q") or "").strip()
    actif = (request.GET.get("actif") or "").strip()
    sort = (request.GET.get("sort") or "date").strip()

    qs = BaladeJour.objects.all()

    if q:
        qs = qs.filter(
            Q(titre__icontains=q) |
            Q(guide__icontains=q) |
            Q(date_affichage__icontains=q)
        )

    if actif == "1":
        qs = qs.filter(actif=True)
    elif actif == "0":
        qs = qs.filter(actif=False)

    sort_map = {
        "date": ("date_debut", "ordre", "titre"),
        "-date": ("-date_debut", "-ordre", "titre"),
        "titre": ("titre", "date_debut"),
        "-titre": ("-titre", "date_debut"),
        "guide": ("guide", "date_debut"),
        "-guide": ("-guide", "date_debut"),
        "ordre": ("ordre", "date_debut", "titre"),
        "-ordre": ("-ordre", "date_debut", "titre"),
    }
    qs = qs.order_by(*sort_map.get(sort, ("date_debut", "ordre", "titre")))

    paginator = Paginator(qs, 12)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    base_qs = request.GET.copy()
    if "page" in base_qs:
        del base_qs["page"]
    qs_no_page = base_qs.urlencode()

    stats = {
        "total": BaladeJour.objects.count(),
        "actives": BaladeJour.objects.filter(actif=True).count(),
        "inactives": BaladeJour.objects.filter(actif=False).count(),
        "a_venir": BaladeJour.objects.filter(date_debut__gte=timezone.localdate()).count(),
    }

    return render(request, "circuitMoto/admin/balades_list.html", {
        "page_obj": page_obj,
        "q": q,
        "actif": actif,
        "sort": sort,
        "qs_no_page": qs_no_page,
        "stats": stats,
    })


@staff_member_required
@transaction.atomic
def balade_create(request):
    if request.method == "POST":
        form = BaladeJourForm(request.POST)
        if form.is_valid():
            obj = form.save()
            messages.success(request, f"Balade créée : {obj.titre}")
            return redirect("bo_balade_list")
    else:
        form = BaladeJourForm(initial={
            "actif": True,
            "ordre": 0,
        })

    return render(request, "circuitMoto/admin/balades_form.html", {
        "form": form,
        "mode": "create",
        "title_page": "Nouvelle balade 1 jour",
    })


@staff_member_required
@transaction.atomic
def balade_edit(request, pk: int):
    obj = get_object_or_404(BaladeJour, pk=pk)

    if request.method == "POST":
        form = BaladeJourForm(request.POST, instance=obj)
        if form.is_valid():
            obj = form.save()
            messages.success(request, f"Balade mise à jour : {obj.titre}")
            return redirect("bo_balade_list")
    else:
        form = BaladeJourForm(instance=obj)

    return render(request, "circuitMoto/admin/balades_form.html", {
        "form": form,
        "mode": "edit",
        "obj": obj,
        "title_page": f"Modifier la balade — {obj.titre}",
    })


@staff_member_required
def demande_balade_list(request):
    q = (request.GET.get("q") or "").strip()
    traite = (request.GET.get("traite") or "").strip()
    sort = (request.GET.get("sort") or "-date").strip()
    balade_id = (request.GET.get("balade") or "").strip()

    qs = (
        DemandeProgrammeBalade.objects
        .prefetch_related("balades")
        .order_by("-cree_le")
    )

    if q:
        qs = qs.filter(
            Q(prenom__icontains=q) |
            Q(nom__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone__icontains=q) |
            Q(commentaire__icontains=q) |
            Q(balades__titre__icontains=q)
        ).distinct()

    if traite == "1":
        qs = qs.filter(traite=True)
    elif traite == "0":
        qs = qs.filter(traite=False)

    if balade_id.isdigit():
        qs = qs.filter(balades__id=int(balade_id)).distinct()

    sort_map = {
        "-date": ("-cree_le",),
        "date": ("cree_le",),
        "nom": ("nom", "prenom", "-cree_le"),
        "-nom": ("-nom", "-prenom", "-cree_le"),
        "email": ("email", "-cree_le"),
        "-email": ("-email", "-cree_le"),
    }
    qs = qs.order_by(*sort_map.get(sort, ("-cree_le",)))

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        selected_ids = request.POST.getlist("ids")

        if action in {"mark_done", "mark_undone"} and selected_ids:
            selected_qs = DemandeProgrammeBalade.objects.filter(id__in=selected_ids)

            if action == "mark_done":
                updated = selected_qs.update(traite=True)
                messages.success(request, f"{updated} demande(s) marquée(s) comme traitée(s).")
            elif action == "mark_undone":
                updated = selected_qs.update(traite=False)
                messages.success(request, f"{updated} demande(s) remise(s) en non traitée(s).")

            url = reverse("bo_demande_balade_list")
            qs_params = request.GET.urlencode()
            return redirect(f"{url}?{qs_params}" if qs_params else url)

    paginator = Paginator(qs, 12)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    base_qs = request.GET.copy()
    if "page" in base_qs:
        del base_qs["page"]
    qs_no_page = base_qs.urlencode()

    balades = BaladeJour.objects.filter(actif=True).order_by("date_debut", "ordre", "titre")

    stats = {
        "total": DemandeProgrammeBalade.objects.count(),
        "non_traitees": DemandeProgrammeBalade.objects.filter(traite=False).count(),
        "traitees": DemandeProgrammeBalade.objects.filter(traite=True).count(),
    }

    balade_selected = None
    if balade_id.isdigit():
        balade_selected = BaladeJour.objects.filter(id=int(balade_id)).first()

    return render(request, "circuitMoto/admin/demandes_balades_list.html", {
        "page_obj": page_obj,
        "q": q,
        "traite": traite,
        "sort": sort,
        "qs_no_page": qs_no_page,
        "stats": stats,
        "balades": balades,
        "balade_id": balade_id,
        "balade_selected": balade_selected,
    })


@staff_member_required
def demande_balade_par_balade(request):
    q = (request.GET.get("q") or "").strip()
    actif = (request.GET.get("actif") or "").strip()
    sort = (request.GET.get("sort") or "-demandes").strip()

    balades_qs = BaladeJour.objects.all()

    if q:
        balades_qs = balades_qs.filter(
            Q(titre__icontains=q) |
            Q(guide__icontains=q) |
            Q(date_affichage__icontains=q)
        )

    if actif == "1":
        balades_qs = balades_qs.filter(actif=True)
    elif actif == "0":
        balades_qs = balades_qs.filter(actif=False)

    balades_qs = balades_qs.annotate(
        total_demandes=Count("demandes_programme", distinct=True),
        total_non_traitees=Count(
            "demandes_programme",
            filter=Q(demandes_programme__traite=False),
            distinct=True,
        ),
        total_traitees=Count(
            "demandes_programme",
            filter=Q(demandes_programme__traite=True),
            distinct=True,
        ),
        total_emails_uniques=Count("demandes_programme__email", distinct=True),
    ).prefetch_related(
        Prefetch(
            "demandes_programme",
            queryset=DemandeProgrammeBalade.objects.order_by("-cree_le"),
        )
    )

    sort_map = {
        "-demandes": ("-total_demandes", "date_debut", "ordre", "titre"),
        "demandes": ("total_demandes", "date_debut", "ordre", "titre"),
        "date": ("date_debut", "ordre", "titre"),
        "-date": ("-date_debut", "-ordre", "titre"),
        "titre": ("titre",),
        "-titre": ("-titre",),
        "emails": ("-total_emails_uniques", "-total_demandes", "titre"),
    }
    balades_qs = balades_qs.order_by(*sort_map.get(sort, ("-total_demandes", "date_debut", "titre")))

    paginator = Paginator(balades_qs, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    base_qs = request.GET.copy()
    if "page" in base_qs:
        del base_qs["page"]
    qs_no_page = base_qs.urlencode()

    stats = {
        "total_balades": BaladeJour.objects.count(),
        "balades_avec_demandes": BaladeJour.objects.filter(demandes_programme__isnull=False).distinct().count(),
        "total_demandes": DemandeProgrammeBalade.objects.count(),
        "total_non_traitees": DemandeProgrammeBalade.objects.filter(traite=False).count(),
    }

    return render(request, "circuitMoto/admin/demandes_balades_par_balade.html", {
        "page_obj": page_obj,
        "q": q,
        "actif": actif,
        "sort": sort,
        "qs_no_page": qs_no_page,
        "stats": stats,
    })

@staff_member_required
def demande_balade_detail(request, pk: int):
    balade = get_object_or_404(BaladeJour, pk=pk)

    q = (request.GET.get("q") or "").strip()
    traite = (request.GET.get("traite") or "").strip()
    sort = (request.GET.get("sort") or "-date").strip()

    qs = (
        DemandeProgrammeBalade.objects
        .filter(balades=balade)
        .prefetch_related("balades")
        .distinct()
    )

    if q:
        qs = qs.filter(
            Q(prenom__icontains=q) |
            Q(nom__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone__icontains=q) |
            Q(commentaire__icontains=q)
        )

    if traite == "1":
        qs = qs.filter(traite=True)
    elif traite == "0":
        qs = qs.filter(traite=False)

    sort_map = {
        "-date": ("-cree_le",),
        "date": ("cree_le",),
        "nom": ("nom", "prenom", "-cree_le"),
        "-nom": ("-nom", "-prenom", "-cree_le"),
        "email": ("email", "-cree_le"),
        "-email": ("-email", "-cree_le"),
    }
    qs = qs.order_by(*sort_map.get(sort, ("-cree_le",)))

    paginator = Paginator(qs, 12)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    base_qs = request.GET.copy()
    if "page" in base_qs:
        del base_qs["page"]
    qs_no_page = base_qs.urlencode()

    stats = {
        "total": DemandeProgrammeBalade.objects.filter(balades=balade).distinct().count(),
        "non_traitees": DemandeProgrammeBalade.objects.filter(balades=balade, traite=False).distinct().count(),
        "traitees": DemandeProgrammeBalade.objects.filter(balades=balade, traite=True).distinct().count(),
        "emails_uniques": (
            DemandeProgrammeBalade.objects
            .filter(balades=balade)
            .values("email")
            .distinct()
            .count()
        ),
    }

    return render(request, "circuitMoto/admin/demande_balade_detail.html", {
        "balade": balade,
        "page_obj": page_obj,
        "q": q,
        "traite": traite,
        "sort": sort,
        "qs_no_page": qs_no_page,
        "stats": stats,
    })


# =========================
# PRE PROGRAMME 2027 - BACK OFFICE
# =========================

def _programme_2027_selection_prefetch():
    return Prefetch(
        "selections",
        queryset=Programme2027Selection.objects.select_related("circuit").order_by("circuit__ordre", "circuit__titre"),
    )


@staff_member_required
def programme_2027_interest_list(request):
    q = (request.GET.get("q") or "").strip()
    traite = (request.GET.get("traite") or "").strip()
    circuit_id = (request.GET.get("circuit") or "").strip()
    niveau = (request.GET.get("niveau") or "").strip()
    sort = (request.GET.get("sort") or "-date").strip()

    qs = (
        Programme2027Interest.objects
        .prefetch_related(_programme_2027_selection_prefetch())
        .order_by("-cree_le")
    )

    if q:
        qs = qs.filter(
            Q(prenom__icontains=q) |
            Q(nom__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone__icontains=q) |
            Q(localite__icontains=q) |
            Q(pays__icontains=q) |
            Q(vehicule__icontains=q) |
            Q(commentaire__icontains=q) |
            Q(selections__circuit__titre__icontains=q)
        ).distinct()

    if traite == "1":
        qs = qs.filter(traite=True)
    elif traite == "0":
        qs = qs.filter(traite=False)

    if circuit_id.isdigit():
        qs = qs.filter(selections__circuit_id=int(circuit_id)).distinct()

    valid_levels = {choice[0] for choice in NiveauInteret2027.choices}
    if niveau in valid_levels:
        qs = qs.filter(selections__niveau_interet=niveau).distinct()

    sort_map = {
        "-date": ("-cree_le",),
        "date": ("cree_le",),
        "nom": ("nom", "prenom", "-cree_le"),
        "-nom": ("-nom", "-prenom", "-cree_le"),
        "email": ("email", "-cree_le"),
        "-email": ("-email", "-cree_le"),
        "decision": ("delai_decision", "-cree_le"),
    }
    qs = qs.order_by(*sort_map.get(sort, ("-cree_le",)))

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        selected_ids = request.POST.getlist("ids")

        if action in {"mark_done", "mark_undone"} and selected_ids:
            selected_qs = Programme2027Interest.objects.filter(id__in=selected_ids)

            if action == "mark_done":
                updated = selected_qs.update(traite=True)
                messages.success(request, f"{updated} interesse(s) marque(s) comme traite(s).")
            elif action == "mark_undone":
                updated = selected_qs.update(traite=False)
                messages.success(request, f"{updated} interesse(s) remis en non traite(s).")

            url = reverse("bo_programme_2027_interests")
            qs_params = request.GET.urlencode()
            return redirect(f"{url}?{qs_params}" if qs_params else url)

    paginator = Paginator(qs, 12)
    page_obj = paginator.get_page(request.GET.get("page"))

    base_qs = request.GET.copy()
    if "page" in base_qs:
        del base_qs["page"]
    qs_no_page = base_qs.urlencode()

    circuits = Programme2027Circuit.objects.filter(actif=True).order_by("ordre", "titre")
    circuit_selected = Programme2027Circuit.objects.filter(id=int(circuit_id)).first() if circuit_id.isdigit() else None

    stats = {
        "total": Programme2027Interest.objects.count(),
        "non_traites": Programme2027Interest.objects.filter(traite=False).count(),
        "traites": Programme2027Interest.objects.filter(traite=True).count(),
        "emails_uniques": Programme2027Interest.objects.values("email").distinct().count(),
        "selections": Programme2027Selection.objects.count(),
        "prets": Programme2027Selection.objects.filter(niveau_interet=NiveauInteret2027.PRET_A_RESERVER).count(),
    }

    return render(request, "circuitMoto/admin/programme_2027_interests.html", {
        "page_obj": page_obj,
        "q": q,
        "traite": traite,
        "circuit_id": circuit_id,
        "niveau": niveau,
        "sort": sort,
        "qs_no_page": qs_no_page,
        "stats": stats,
        "circuits": circuits,
        "circuit_selected": circuit_selected,
        "niveau_choices": NiveauInteret2027.choices,
    })


@staff_member_required
def programme_2027_by_circuit(request):
    q = (request.GET.get("q") or "").strip()
    actif = (request.GET.get("actif") or "1").strip()
    categorie = (request.GET.get("categorie") or "").strip()
    sort = (request.GET.get("sort") or "-interesses").strip()

    qs = Programme2027Circuit.objects.all()

    if q:
        qs = qs.filter(
            Q(titre__icontains=q) |
            Q(mois__icontains=q) |
            Q(duree__icontains=q) |
            Q(formule__icontains=q)
        )

    if actif == "1":
        qs = qs.filter(actif=True)
    elif actif == "0":
        qs = qs.filter(actif=False)

    valid_categories = {choice[0] for choice in Programme2027Circuit._meta.get_field("categorie").choices}
    if categorie in valid_categories:
        qs = qs.filter(categorie=categorie)

    qs = qs.annotate(
        total_interesses=Count("selections__demande", distinct=True),
        total_emails_uniques=Count("selections__demande__email", distinct=True),
        total_non_traites=Count(
            "selections__demande",
            filter=Q(selections__demande__traite=False),
            distinct=True,
        ),
        total_curieux=Count(
            "selections",
            filter=Q(selections__niveau_interet=NiveauInteret2027.CURIEUX),
            distinct=True,
        ),
        total_interesse=Count(
            "selections",
            filter=Q(selections__niveau_interet=NiveauInteret2027.INTERESSE),
            distinct=True,
        ),
        total_tres_interesse=Count(
            "selections",
            filter=Q(selections__niveau_interet=NiveauInteret2027.TRES_INTERESSE),
            distinct=True,
        ),
        total_pret=Count(
            "selections",
            filter=Q(selections__niveau_interet=NiveauInteret2027.PRET_A_RESERVER),
            distinct=True,
        ),
    ).prefetch_related(
        Prefetch(
            "selections",
            queryset=Programme2027Selection.objects.select_related("demande").order_by("-cree_le"),
        )
    )

    sort_map = {
        "-interesses": ("-total_interesses", "ordre", "titre"),
        "interesses": ("total_interesses", "ordre", "titre"),
        "-prets": ("-total_pret", "-total_tres_interesse", "ordre", "titre"),
        "ordre": ("ordre", "titre"),
        "-ordre": ("-ordre", "titre"),
        "titre": ("titre",),
        "-titre": ("-titre",),
    }
    qs = qs.order_by(*sort_map.get(sort, ("-total_interesses", "ordre", "titre")))

    paginator = Paginator(qs, 12)
    page_obj = paginator.get_page(request.GET.get("page"))

    base_qs = request.GET.copy()
    if "page" in base_qs:
        del base_qs["page"]
    qs_no_page = base_qs.urlencode()

    stats = {
        "total_circuits": Programme2027Circuit.objects.count(),
        "circuits_actifs": Programme2027Circuit.objects.filter(actif=True).count(),
        "circuits_avec_interesses": Programme2027Circuit.objects.filter(selections__isnull=False).distinct().count(),
        "total_interesses": Programme2027Interest.objects.count(),
        "total_selections": Programme2027Selection.objects.count(),
        "total_non_traites": Programme2027Interest.objects.filter(traite=False).count(),
    }

    return render(request, "circuitMoto/admin/programme_2027_by_circuit.html", {
        "page_obj": page_obj,
        "q": q,
        "actif": actif,
        "categorie": categorie,
        "sort": sort,
        "qs_no_page": qs_no_page,
        "stats": stats,
        "categorie_choices": Programme2027Circuit._meta.get_field("categorie").choices,
    })


@staff_member_required
def programme_2027_circuit_detail(request, pk: int):
    circuit = get_object_or_404(Programme2027Circuit, pk=pk)

    q = (request.GET.get("q") or "").strip()
    traite = (request.GET.get("traite") or "").strip()
    niveau = (request.GET.get("niveau") or "").strip()
    sort = (request.GET.get("sort") or "-date").strip()

    qs = (
        Programme2027Selection.objects
        .filter(circuit=circuit)
        .select_related("demande", "circuit")
    )

    if q:
        qs = qs.filter(
            Q(demande__prenom__icontains=q) |
            Q(demande__nom__icontains=q) |
            Q(demande__email__icontains=q) |
            Q(demande__telephone__icontains=q) |
            Q(demande__localite__icontains=q) |
            Q(demande__pays__icontains=q) |
            Q(demande__vehicule__icontains=q) |
            Q(demande__commentaire__icontains=q) |
            Q(mode_prefere__icontains=q) |
            Q(commentaire__icontains=q)
        )

    if traite == "1":
        qs = qs.filter(demande__traite=True)
    elif traite == "0":
        qs = qs.filter(demande__traite=False)

    valid_levels = {choice[0] for choice in NiveauInteret2027.choices}
    if niveau in valid_levels:
        qs = qs.filter(niveau_interet=niveau)

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        selected_ids = request.POST.getlist("ids")

        if action in {"mark_done", "mark_undone"} and selected_ids:
            demande_ids = list(
                Programme2027Selection.objects
                .filter(id__in=selected_ids, circuit=circuit)
                .values_list("demande_id", flat=True)
            )
            selected_qs = Programme2027Interest.objects.filter(id__in=demande_ids)

            if action == "mark_done":
                updated = selected_qs.update(traite=True)
                messages.success(request, f"{updated} interesse(s) marque(s) comme traite(s).")
            elif action == "mark_undone":
                updated = selected_qs.update(traite=False)
                messages.success(request, f"{updated} interesse(s) remis en non traite(s).")

            url = reverse("bo_programme_2027_circuit_detail", args=[circuit.pk])
            qs_params = request.GET.urlencode()
            return redirect(f"{url}?{qs_params}" if qs_params else url)

    sort_map = {
        "-date": ("-demande__cree_le",),
        "date": ("demande__cree_le",),
        "nom": ("demande__nom", "demande__prenom", "-demande__cree_le"),
        "-nom": ("-demande__nom", "-demande__prenom", "-demande__cree_le"),
        "email": ("demande__email", "-demande__cree_le"),
        "-email": ("-demande__email", "-demande__cree_le"),
        "niveau": ("niveau_interet", "-demande__cree_le"),
        "-niveau": ("-niveau_interet", "-demande__cree_le"),
    }
    qs = qs.order_by(*sort_map.get(sort, ("-demande__cree_le",)))

    paginator = Paginator(qs, 12)
    page_obj = paginator.get_page(request.GET.get("page"))

    base_qs = request.GET.copy()
    if "page" in base_qs:
        del base_qs["page"]
    qs_no_page = base_qs.urlencode()

    all_for_circuit = Programme2027Selection.objects.filter(circuit=circuit)
    stats = {
        "total": all_for_circuit.count(),
        "emails_uniques": all_for_circuit.values("demande__email").distinct().count(),
        "non_traites": all_for_circuit.filter(demande__traite=False).count(),
        "traites": all_for_circuit.filter(demande__traite=True).count(),
        "curieux": all_for_circuit.filter(niveau_interet=NiveauInteret2027.CURIEUX).count(),
        "interesses": all_for_circuit.filter(niveau_interet=NiveauInteret2027.INTERESSE).count(),
        "tres_interesses": all_for_circuit.filter(niveau_interet=NiveauInteret2027.TRES_INTERESSE).count(),
        "prets": all_for_circuit.filter(niveau_interet=NiveauInteret2027.PRET_A_RESERVER).count(),
    }

    return render(request, "circuitMoto/admin/programme_2027_circuit_detail.html", {
        "circuit": circuit,
        "page_obj": page_obj,
        "q": q,
        "traite": traite,
        "niveau": niveau,
        "sort": sort,
        "qs_no_page": qs_no_page,
        "stats": stats,
        "niveau_choices": NiveauInteret2027.choices,
    })
